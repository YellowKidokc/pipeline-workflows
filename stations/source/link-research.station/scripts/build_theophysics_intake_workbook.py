from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sys

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"

if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from link_research_engine.modules.theophysics_seed import load_theophysics_seed_index


METADATA_JSON = Path(
    r"C:\Users\lowes\Desktop\Conm\cons\hTML eX[ORTS\theophysics-deploy\site-lib\metadata.json"
)
OUT_XLSX = ROOT / "data" / "workbooks" / "theophysics_link_intake.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
ACCENT_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_body(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BOX
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            lengths = max(len(part) for part in str(cell.value).splitlines())
            widths[cell.column] = min(max(widths.get(cell.column, 0), lengths + 2), 42)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_summary_sheet(wb: Workbook, page_count: int, edge_count: int, internal_count: int, external_count: int) -> None:
    ws = wb.active
    ws.title = "RUN_SUMMARY"
    ws["A1"] = "THEOPHYSICS LINK INTAKE"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Seed source"
    ws["B2"] = str(METADATA_JSON)
    ws["A4"] = "Metric"
    ws["B4"] = "Value"
    style_header(ws[4])

    rows = [
        ("Pages discovered", page_count),
        ("Total link edges", edge_count),
        ("Internal links", internal_count),
        ("External links", external_count),
        ("Output workbook", str(OUT_XLSX)),
    ]
    start = 5
    for offset, (label, value) in enumerate(rows):
        ws.cell(start + offset, 1, label)
        ws.cell(start + offset, 2, value)

    ws["D2"] = "Orientation"
    ws["D2"].font = BOLD_FONT
    ws["D3"] = (
        "This workbook is the first live link-intake pass for Theophysics. "
        "It treats the deployed metadata export as a seed corpus and turns it into "
        "reviewable pages plus source-target link edges."
    )
    ws["D3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["D3"].fill = ACCENT_FILL
    ws.merge_cells("D3:H6")
    style_body(ws)
    autosize(ws)
    ws.freeze_panes = "A4"


def build_pages_sheet(wb: Workbook, pages) -> None:
    ws = wb.create_sheet("SEED_PAGES")
    headers = [
        "page_id",
        "title",
        "export_path",
        "source_path",
        "full_url",
        "link_count",
        "backlink_count",
        "source_size",
        "modified_time",
        "show_in_tree",
        "page_type",
    ]
    ws.append(headers)
    style_header(ws[1])
    for page in pages:
        ws.append(
            [
                page.page_id,
                page.title,
                page.export_path,
                page.source_path,
                page.full_url,
                page.link_count,
                page.backlink_count,
                page.source_size,
                page.modified_time,
                "Y" if page.show_in_tree else "N",
                page.page_type,
            ]
        )
    style_body(ws)
    autosize(ws)
    ws.freeze_panes = "A2"


def build_edges_sheet(wb: Workbook, edges) -> None:
    ws = wb.create_sheet("LINK_EDGES")
    headers = [
        "source_page_id",
        "source_title",
        "target_url",
        "target_type",
        "target_domain",
    ]
    ws.append(headers)
    style_header(ws[1])
    for edge in edges:
        ws.append(
            [
                edge.source_page_id,
                edge.source_title,
                edge.target_url,
                edge.target_type,
                edge.target_domain,
            ]
        )
    style_body(ws)
    autosize(ws)
    ws.freeze_panes = "A2"


def build_domains_sheet(wb: Workbook, edges) -> None:
    ws = wb.create_sheet("DOMAIN_COUNTS")
    counts = Counter(edge.target_domain for edge in edges)
    ws.append(["target_domain", "link_count"])
    style_header(ws[1])
    for domain, count in counts.most_common():
        ws.append([domain, count])
    style_body(ws)
    autosize(ws)
    ws.freeze_panes = "A2"


def build_page_rollup_sheet(wb: Workbook, pages) -> None:
    ws = wb.create_sheet("PAGE_ROLLUP")
    ws.append(["title", "link_count", "backlink_count", "source_size", "export_path"])
    style_header(ws[1])
    sorted_pages = sorted(pages, key=lambda p: (-p.link_count, p.title.lower()))
    for page in sorted_pages:
        ws.append([page.title, page.link_count, page.backlink_count, page.source_size, page.export_path])
    style_body(ws)
    autosize(ws)
    ws.freeze_panes = "A2"


def main() -> None:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    pages, edges = load_theophysics_seed_index(str(METADATA_JSON))
    internal_count = sum(1 for edge in edges if edge.target_type == "internal_page")
    external_count = sum(1 for edge in edges if edge.target_type == "external")

    wb = Workbook()
    build_summary_sheet(wb, len(pages), len(edges), internal_count, external_count)
    build_pages_sheet(wb, pages)
    build_edges_sheet(wb, edges)
    build_domains_sheet(wb, edges)
    build_page_rollup_sheet(wb, pages)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True

    wb.save(OUT_XLSX)
    print(f"Saved workbook to {OUT_XLSX}")
    print(f"Pages: {len(pages)}")
    print(f"Edges: {len(edges)}")
    print(f"Internal links: {internal_count}")
    print(f"External links: {external_count}")


if __name__ == "__main__":
    main()
