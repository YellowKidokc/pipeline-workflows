from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:  # pragma: no cover - optional export
    Workbook = None

SHARED = Path(__file__).resolve().parents[1] / "_shared"
if str(SHARED) not in sys.path:
    sys.path.insert(0, str(SHARED))
from canonical_lexicon import regex_from_terms, semantic_terms, sheet_terms  # noqa: E402


PERSONAL_BELIEF_RE = re.compile(
    r"\b(i believe|i think|my belief|personally|i don't have scientific evidence|"
    r"i do not have scientific evidence|my own belief|i figure|i figured)\b",
    re.I,
)
QUESTION_RE = re.compile(r"\?\s*$|^\s*(what|why|how|was|were|is|are|could|should|would)\b", re.I)
BOILERPLATE_RE = re.compile(
    r"\b(read aloud|deep dive|podcast|theophysics institute|scrape_mode|source_sha256|"
    r"source_path|deploy_relative|audio narration|0:00)\b",
    re.I,
)
EVIDENCE_RE = regex_from_terms(
    sheet_terms("EVIDENCE_TERMS")
    | {"data", "rate", "observed", "measured", "baseline", "control group", "correlation", "egeland", "statistical"}
    | {r"\d+(?:\.\d+)?\s*%", r"\d+(?:\.\d+)?\s*sigma", r"p\s*[<=>]"},
    r"\b(\d+(?:\.\d+)?\s*%|\d+(?:\.\d+)?\s*sigma|p\s*[<=>]|data|rate|study|"
    r"observed|measured|evidence|baseline|control group|correlation|Egeland|statistical)\b",
)
MECHANISM_RE = regex_from_terms(
    sheet_terms("DEPENDENCY_TERMS") | sheet_terms("BRIDGE_TERMS") | {"mechanism", "operator", "constraint", "entropy", "coherence", "if .+ then"},
    r"\b(because|therefore|thus|implies|requires|mechanism|operator|constraint|entropy|coherence|if .+ then)\b",
)
BOUNDARY_RE = re.compile(r"\b(if|when|under|within|in this paper|for this series|control group|compared with|versus|vs)\b", re.I)
PREDICTION_RE = regex_from_terms(
    sheet_terms("FALSIFY_TERMS") | {"predict", "prediction", "should", "would expect", "if .+ then", "testable", "falsifiable", "kill condition"},
    r"\b(predict|should|would expect|if .+ then|testable|falsifiable|kill condition)\b",
)
GROUND_RE = regex_from_terms(
    sheet_terms("EVIDENCE_TERMS") | sheet_terms("DEPENDENCY_TERMS") | semantic_terms(buckets={"CLAIM_STRENGTH"}) | {"axiom", "ground", "scripture", "equation", "model", "egeland"},
    r"\b(axiom|source|ground|because|depends|derived|scripture|study|data|equation|model|Egeland)\b",
)
THEOLOGICAL_RE = re.compile(r"\b(god|christ|jesus|sin|grace|adam|scripture|theological|salvation|atonement)\b", re.I)
EQUATION_RE = re.compile(r"(=|χ|\\chi|→|->|∇|Σ|\\frac|\\sum|\\int)")


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def classify_kind(text: str, original_classification: str, has_equation: bool) -> tuple[str, str]:
    t = compact(text)
    low = t.lower()
    word_count = len(re.findall(r"\b[\w'-]+\b", t))

    if BOILERPLATE_RE.search(t) or t.startswith("--- "):
        return "BOILERPLATE", "DROP"
    if PERSONAL_BELIEF_RE.search(t):
        return "PERSONAL_BELIEF", "HOLD_PERSONAL_NOT_PAPER_CLAIM"
    if word_count < 5 and not has_equation and not EVIDENCE_RE.search(t):
        return "FRAGMENT", "DROP"
    if QUESTION_RE.search(t) and not MECHANISM_RE.search(t):
        return "QUESTION", "BACKGROUND"
    if has_equation or original_classification == "EQUATION" or EQUATION_RE.search(t):
        return "EQUATION_CLAIM", "REVIEW"
    if EVIDENCE_RE.search(t):
        return "EVIDENCE_CLAIM", "REVIEW"
    if THEOLOGICAL_RE.search(t) or original_classification == "THEOLOGICAL":
        if "i believe" in low or "my belief" in low:
            return "PERSONAL_BELIEF", "HOLD_PERSONAL_NOT_PAPER_CLAIM"
        return "THEOLOGICAL_CLAIM", "REVIEW"
    if MECHANISM_RE.search(t):
        return "PAPER_CLAIM", "REVIEW"
    return "PAPER_CLAIM", "REVIEW_LOW_CONFIDENCE"


def q_value(condition: bool, good: str, weak: str) -> str:
    return good if condition else weak


def seven_q(text: str, kind: str, status: str) -> dict[str, str]:
    t = compact(text)
    word_count = len(re.findall(r"\b[\w'-]+\b", t))
    return {
        "Q0_posture": "personal-belief-boundary" if kind == "PERSONAL_BELIEF" else "paper-claim-review",
        "Q1_identity": q_value(kind not in {"FRAGMENT", "BOILERPLATE"}, kind, "not-a-claim"),
        "Q2_scope": q_value(bool(BOUNDARY_RE.search(t)), "bounded", "broad-or-implicit"),
        "Q3_commitment": q_value(word_count >= 8 and status.startswith("REVIEW"), "reviewable assertion", "weak/fragmentary"),
        "Q4_support": q_value(bool(EVIDENCE_RE.search(t)), "explicit support marker", "support not explicit in sentence"),
        "Q5_ground": q_value(bool(GROUND_RE.search(t)), "ground marker present", "ground missing/implicit"),
        "Q6_prediction": q_value(bool(PREDICTION_RE.search(t)), "prediction/test implied", "no prediction in sentence"),
        "Q7_destroy": destroy_prompt(t, kind, status),
    }


def destroy_prompt(text: str, kind: str, status: str) -> str:
    if status == "DROP":
        return "Do not death-test; not a paper claim."
    if kind == "PERSONAL_BELIEF":
        return "Boundary test: keep as personal belief unless the paper supplies public evidence or formal argument."
    if kind == "EVIDENCE_CLAIM":
        return "Kill if the cited statistic/source is false, cherry-picked, or not comparable to the control group."
    if kind == "EQUATION_CLAIM":
        return "Kill if variables are undefined, direction is wrong, or the equation does not map to the prose claim."
    if kind == "THEOLOGICAL_CLAIM":
        return "Kill or narrow if it exceeds the stated theological authority or confuses doctrine with measurement."
    return "Kill if the pattern appears without the proposed mechanism or if a simpler explanation fits the same data."


def boundary_note(kind: str) -> str:
    if kind == "PERSONAL_BELIEF":
        return "Personal belief / theological intuition. Do not present as paper evidence."
    if kind == "BOILERPLATE":
        return "Page metadata or chrome. Exclude from claim audit."
    if kind == "FRAGMENT":
        return "Fragment. Merge with neighboring text or drop."
    if kind == "QUESTION":
        return "Question/background unless later answered as a claim."
    return "Paper-facing claim candidate; send through review/proof/evidence gates."


def load_claims(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def enrich(data: dict[str, Any]) -> dict[str, Any]:
    enriched_docs = []
    counts: Counter[str] = Counter()
    status_counts: Counter[str] = Counter()
    for doc in data.get("documents", []):
        new_claims = []
        for idx, claim in enumerate(doc.get("claims", []), start=1):
            text = compact(claim.get("text", ""))
            kind, status = classify_kind(text, claim.get("classification", ""), bool(claim.get("has_equation")))
            q = seven_q(text, kind, status)
            counts[kind] += 1
            status_counts[status] += 1
            new_claims.append(
                {
                    **claim,
                    "claim_number": idx,
                    "claim_kind": kind,
                    "review_status": status,
                    "boundary_note": boundary_note(kind),
                    **q,
                }
            )
        enriched_docs.append({**doc, "claims": new_claims})
    return {
        **data,
        "schema": "claim_extractor.7q_enriched.v1",
        "claim_kind_distribution": dict(counts),
        "review_status_distribution": dict(status_counts),
        "documents": enriched_docs,
    }


def iter_claim_rows(data: dict[str, Any]):
    for doc in data.get("documents", []):
        for c in doc.get("claims", []):
            yield c


def write_csv(path: Path, data: dict[str, Any]) -> None:
    fields = [
        "claim_number", "claim_id", "source_file", "section", "claim_kind", "review_status",
        "classification", "confidence", "text", "boundary_note",
        "Q0_posture", "Q1_identity", "Q2_scope", "Q3_commitment", "Q4_support",
        "Q5_ground", "Q6_prediction", "Q7_destroy",
    ]
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in iter_claim_rows(data):
            writer.writerow(row)


def write_xlsx(path: Path, data: dict[str, Any]) -> None:
    if Workbook is None:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "7Q_Claims"
    headers = [
        "claim_number", "claim_kind", "review_status", "section", "text", "boundary_note",
        "Q0_posture", "Q1_identity", "Q2_scope", "Q3_commitment", "Q4_support",
        "Q5_ground", "Q6_prediction", "Q7_destroy",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")
        cell.alignment = Alignment(wrap_text=True)
    for c in iter_claim_rows(data):
        ws.append([c.get(h, "") for h in headers])
    widths = [14, 22, 26, 28, 70, 38, 24, 24, 22, 24, 28, 26, 24, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["total_claims", data.get("total_claims", "")])
    for k, v in data.get("claim_kind_distribution", {}).items():
        summary.append([f"kind:{k}", v])
    for k, v in data.get("review_status_distribution", {}).items():
        summary.append([f"status:{k}", v])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 16
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Add belief/claim taxonomy and deterministic 7Q fields to claim extractor JSON.")
    parser.add_argument("claims_json", type=Path)
    parser.add_argument("--out-dir", type=Path)
    args = parser.parse_args()

    data = enrich(load_claims(args.claims_json))
    out_dir = args.out_dir or args.claims_json.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = args.claims_json.stem
    json_path = out_dir / f"{stem}.7q-enriched.json"
    csv_path = out_dir / f"{stem}.7q-enriched.csv"
    xlsx_path = out_dir / f"{stem}.7q-enriched.xlsx"
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(csv_path, data)
    write_xlsx(xlsx_path, data)
    print(json.dumps({
        "json": str(json_path),
        "csv": str(csv_path),
        "xlsx": str(xlsx_path) if Workbook is not None else "",
        "claim_kind_distribution": data["claim_kind_distribution"],
        "review_status_distribution": data["review_status_distribution"],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
