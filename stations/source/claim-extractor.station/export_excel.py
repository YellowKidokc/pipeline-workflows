r"""
08_CLAIMS: Excel Exporter
Takes a claims JSON from extract.py and produces a review-ready Excel workbook.

Usage:
    python export_excel.py <claims_json> [--output <path>]

Examples:
    python export_excel.py claims_20260502.json
    python export_excel.py claims_20260502.json --output REVIEW.xlsx
"""
import sys, json, argparse, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

ILLEGAL_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
def sanitize(text):
    if not isinstance(text, str):
        return text
    text = ILLEGAL_CHARS.sub('', text)
    if len(text) > 500:
        text = text[:500] + "..."
    return text

COLOR_MAP = {
    "PRE_SYSTEM":             ("FFDDDD", "FF4444"),
    "PRESUPPOSITION":         ("FFF0DD", "FF8800"),
    "AXIOM":                  ("FFFFDD", "CC9900"),
    "DEFINITION":             ("DDEDFF", "4A9EFF"),
    "THEOREM":                ("DDFFD0", "22C55E"),
    "COROLLARY":              ("E8FFE0", "44AA44"),
    "EQUATION":               ("E0E0FF", "6666CC"),
    "PREDICTION":             ("DDFFF8", "00AAAA"),
    "EVIDENCE":               ("E0FFF0", "008800"),
    "THEOLOGICAL_POSTULATE":  ("FFF5DD", "D4AF37"),
    "THEOLOGICAL":            ("FFF5DD", "D4AF37"),
    "CONJECTURE":             ("FFF0FF", "AA44AA"),
    "RESTATEMENT":            ("F0F0F0", "888888"),
    "UNCLASSIFIED":           ("FFFFFF", "333333"),
    "DROP":                   ("FFCCCC", "FF0000"),
}

def main():
    parser = argparse.ArgumentParser(description="Export claims JSON to review Excel")
    parser.add_argument("json_file", help="Path to claims JSON from extract.py")
    parser.add_argument("--output", "-o", help="Output Excel path (default: same name .xlsx)")
    args = parser.parse_args()
    
    json_path = Path(args.json_file)
    if not json_path.exists():
        print(f"ERROR: {json_path} not found")
        sys.exit(1)
    
    data = json.loads(json_path.read_text(encoding="utf-8"))
    
    if args.output:
        out_path = Path(args.output)
    else:
        out_path = json_path.with_suffix(".xlsx")
    
    wb = Workbook()
    
    # === SUMMARY SHEET ===
    ws_sum = wb.active
    ws_sum.title = "Summary"
    
    title_font = Font(name="Arial", bold=True, size=16, color="D4AF37")
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    body_font = Font(name="Arial", size=11)
    small_font = Font(name="Arial", size=10, color="888888")
    header_fill = PatternFill("solid", fgColor="333333")
    
    ws_sum["A1"] = "CLAIM EXTRACTION REVIEW"
    ws_sum["A1"].font = title_font
    ws_sum["A2"] = f"Run: {data['run_id']} | Source: {data['source_folder']}"
    ws_sum["A2"].font = small_font
    ws_sum["A3"] = f"Files: {data['files_processed']} | Total claims: {data['total_claims']}"
    ws_sum["A3"].font = body_font
    
    ws_sum["A5"] = "Classification"
    ws_sum["B5"] = "Count"
    ws_sum["A5"].font = header_font
    ws_sum["B5"].font = header_font
    ws_sum["A5"].fill = header_fill
    ws_sum["B5"].fill = header_fill
    
    row = 6
    for cls, count in sorted(data["classification_distribution"].items(), key=lambda x: -x[1]):
        bg, fg = COLOR_MAP.get(cls, ("FFFFFF", "333333"))
        ws_sum.cell(row=row, column=1, value=cls).font = Font(name="Arial", bold=True, color=fg)
        ws_sum.cell(row=row, column=1).fill = PatternFill("solid", fgColor=bg)
        ws_sum.cell(row=row, column=2, value=count).font = body_font
        row += 1
    
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 10
    
    # === ALL CLAIMS SHEET ===
    ws = wb.create_sheet("All Claims")
    headers = ["#", "Source File", "Section", "Claim Text", "Classification", 
                "Confidence", "Has Equation", "Has Scripture", "REVIEW (your call)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    claim_num = 0
    for doc in data["documents"]:
        source = Path(doc["source_file"]).name
        for claim in doc["claims"]:
            claim_num += 1
            cls = claim["classification"]
            bg, fg = COLOR_MAP.get(cls, ("FFFFFF", "333333"))
            fill = PatternFill("solid", fgColor=bg)
            
            row_data = [
                claim_num,
                sanitize(source),
                sanitize(claim.get("section", "")),
                sanitize(claim["text"]),
                cls,
                round(claim.get("confidence", 0), 2),
                "Y" if claim.get("has_equation") else "",
                "Y" if claim.get("has_scripture_ref") else "",
                ""  # empty review column for David
            ]
            
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=claim_num + 1, column=c, value=val)
                cell.font = Font(name="Arial", size=9)
                cell.fill = fill
                if c == 4:  # claim text
                    cell.alignment = Alignment(wrap_text=True)
                if c == 5:  # classification
                    cell.font = Font(name="Arial", size=9, bold=True, color=fg)
                if c == 9:  # review column
                    cell.fill = PatternFill("solid", fgColor="FFFFF0")
                    cell.font = Font(name="Arial", size=10, color="0000FF")
    
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 20
    
    ws.auto_filter.ref = f"A1:I{claim_num + 1}"
    ws.freeze_panes = "A2"
    
    # === PER-FILE SHEET ===
    ws_files = wb.create_sheet("By File")
    ws_files.cell(row=1, column=1, value="File").font = header_font
    ws_files.cell(row=1, column=1).fill = header_fill
    ws_files.cell(row=1, column=2, value="Claims").font = header_font
    ws_files.cell(row=1, column=2).fill = header_fill
    ws_files.cell(row=1, column=3, value="Top Classification").font = header_font
    ws_files.cell(row=1, column=3).fill = header_fill
    
    for i, doc in enumerate(sorted(data["documents"], key=lambda d: -d["total_claims"]), 2):
        fname = Path(doc["source_file"]).name
        ws_files.cell(row=i, column=1, value=fname).font = Font(name="Arial", size=10)
        ws_files.cell(row=i, column=2, value=doc["total_claims"]).font = body_font
        
        # Find dominant classification
        cls_count = {}
        for c in doc["claims"]:
            cls_count[c["classification"]] = cls_count.get(c["classification"], 0) + 1
        if cls_count:
            top = max(cls_count, key=cls_count.get)
            ws_files.cell(row=i, column=3, value=top).font = body_font
    
    ws_files.column_dimensions["A"].width = 50
    ws_files.column_dimensions["B"].width = 10
    ws_files.column_dimensions["C"].width = 25
    
    # === PER-TYPE SHEETS ===
    type_claims = {}
    for doc in data["documents"]:
        for claim in doc["claims"]:
            cls = claim["classification"]
            if cls not in type_claims:
                type_claims[cls] = []
            type_claims[cls].append((Path(doc["source_file"]).name, claim))
    
    for cls in sorted(type_claims.keys()):
        items = type_claims[cls]
        sheet_name = cls[:28]  # Excel 31 char limit
        ws_t = wb.create_sheet(sheet_name)
        bg, fg = COLOR_MAP.get(cls, ("FFFFFF", "333333"))
        
        for c, h in enumerate(["Source", "Section", "Claim", "Confidence", "REVIEW"], 1):
            cell = ws_t.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for i, (source, claim) in enumerate(items, 2):
            ws_t.cell(row=i, column=1, value=sanitize(source)).font = Font(name="Arial", size=9)
            ws_t.cell(row=i, column=2, value=sanitize(claim.get("section", ""))).font = Font(name="Arial", size=9)
            ws_t.cell(row=i, column=3, value=sanitize(claim["text"])).font = Font(name="Arial", size=9)
            ws_t.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
            ws_t.cell(row=i, column=4, value=round(claim.get("confidence", 0), 2)).font = Font(name="Arial", size=9)
            ws_t.cell(row=i, column=5).fill = PatternFill("solid", fgColor="FFFFF0")
        
        ws_t.column_dimensions["A"].width = 30
        ws_t.column_dimensions["B"].width = 25
        ws_t.column_dimensions["C"].width = 60
        ws_t.column_dimensions["D"].width = 10
        ws_t.column_dimensions["E"].width = 20
    
    wb.save(out_path)
    print(f"Saved review workbook: {out_path}")
    print(f"Total claims: {claim_num}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"\nOpen in Excel, review the 'REVIEW' column, then we push to Postgres.")

if __name__ == "__main__":
    main()
