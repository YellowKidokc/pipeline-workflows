"""
Lowe Standard Full Scorer v2.0
==============================
Scores documents using OpenAI via Cloudflare AI Gateway.

Two modes:
  --mode prosecution  : Score theories against 8-criteria prosecution standard
  --mode tsr          : Score Theophysics papers against TSR100 rubric

Usage:
  python lowe_scorer.py --mode prosecution --input "path/to/papers"
  python lowe_scorer.py --mode tsr --input "path/to/papers"
  python lowe_scorer.py --mode prosecution --input "path/to/papers" --output results.xlsx

Author: David Lowe / Claude
Version: 2.0 | 2026-02-25
"""

import os, sys, json, re, time, glob, argparse
from pathlib import Path
from datetime import datetime

try:
    from openai import OpenAI
except ImportError:
    print("Installing openai package...")
    os.system(f"{sys.executable} -m pip install openai -q")
    from openai import OpenAI

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl package...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Config ──────────────────────────────────────────────────────────────
# Precedence: environment variable > config.txt value. This lets the secret
# live in an env var (or a gitignored config.txt) instead of being hard-coded.
# config.txt is OPTIONAL — env vars alone are enough.
PLACEHOLDER_KEYS = {"sk-PASTE_YOUR_KEY_HERE", "sk-YOUR_KEY_HERE", ""}
ENV_OVERRIDES = ("OPENAI_API_KEY", "BASE_URL", "MODEL", "TEMPERATURE", "MAX_TOKENS")


def load_config(config_path=None):
    if config_path is None:
        config_path = Path(__file__).parent / "config.txt"
    cfg = {}
    config_path = Path(config_path)
    if config_path.exists():
        with open(config_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    cfg[k.strip()] = v.strip()
    # Untracked sibling config.local.txt overrides the tracked config.txt,
    # so the real secret can live on disk without being committed/shared.
    local_path = config_path.with_name(f"{config_path.stem}.local{config_path.suffix}")
    if local_path.exists():
        with open(local_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    k, v = line.split("=", 1)
                    cfg[k.strip()] = v.strip()
    # Environment variables win over file values (keeps secrets out of files).
    for key in ENV_OVERRIDES:
        env_val = os.environ.get(key)
        if env_val:
            cfg[key] = env_val
    return cfg


def get_client(cfg):
    return OpenAI(
        api_key=cfg["OPENAI_API_KEY"],
        base_url=cfg.get("BASE_URL", "https://api.openai.com/v1"),
    )


# ── Rubrics ─────────────────────────────────────────────────────────────

PROSECUTION_SYSTEM = """You are a rigorous philosophy of science examiner. You evaluate theories of consciousness 
against strict structural criteria. You are not hostile — you are precise. You score ONLY what is demonstrated 
in the text, not what is implied or could be inferred. Absence of evidence is scored as absence.

You must return ONLY valid JSON, no markdown, no commentary."""

PROSECUTION_USER = """Score this theory/paper against the 8-criteria Prosecution Standard.

For each criterion, provide:
- score: integer 0-5 (0=absent, 1=gestured at, 2=partially present, 3=present but weak, 4=strong, 5=rigorous/complete)
- evidence: one sentence quoting or citing the specific text that justifies your score
- note: one sentence on what's missing or what would raise the score

THE 8 CRITERIA:

1. FORMAL AXIOMS: Does it state formal axioms (explicit foundational assumptions that could be listed and numbered)?
   0=none, 3=implicit assumptions extractable, 5=numbered formal axiom system

2. QUANTITATIVE PREDICTIONS: Does it make specific, testable numerical predictions?
   0=none, 3=qualitative predictions only, 5=specific numbers with units and significance thresholds

3. DEFEAT CONDITIONS: Does it state what would falsify or break the theory?
   0=none, 3=vague "if X were shown wrong", 5=explicit defeat conditions tied to specific experiments

4. EMPIRICAL CONFIRMATION: Does it cite empirical data supporting the theory?
   0=none, 2=anecdotal or <3σ, 3=3-5σ single dataset, 4=multiple datasets >3σ, 5=multiple independent datasets >5σ

5. HARD PROBLEM: How does it handle the hard problem of consciousness?
   0=ignores it, 1=acknowledges but punts, 2=dissolves/redefines it, 3=addresses partially, 5=provides explanatory bridge from physics to experience

6. BINDING PROBLEM: Does it explain the unity of consciousness?
   0=ignores, 3=addresses partially, 5=formal mechanism for integration

7. MORAL CONSCIOUSNESS: Does it address why we experience moral obligation, not just sensation?
   0=ignores, 3=mentions ethics tangentially, 5=formalizes moral consciousness as part of the theory

8. COMPLETE CAUSAL CHAIN: Does it provide an unbroken chain from fundamental physics to subjective experience?
   0=no, 2=partial chain with gaps, 4=nearly complete with one identified gap, 5=complete chain

Respond with EXACTLY this JSON structure:
{{
  "theory_name": "extracted or inferred name",
  "total_score": 0,
  "max_possible": 40,
  "percentage": 0.0,
  "criteria": {{
    "formal_axioms": {{"score": 0, "evidence": "", "note": ""}},
    "quantitative_predictions": {{"score": 0, "evidence": "", "note": ""}},
    "defeat_conditions": {{"score": 0, "evidence": "", "note": ""}},
    "empirical_confirmation": {{"score": 0, "evidence": "", "note": ""}},
    "hard_problem": {{"score": 0, "evidence": "", "note": ""}},
    "binding_problem": {{"score": 0, "evidence": "", "note": ""}},
    "moral_consciousness": {{"score": 0, "evidence": "", "note": ""}},
    "complete_causal_chain": {{"score": 0, "evidence": "", "note": ""}}
  }},
  "verdict": "COMPLETE or INCOMPLETE or PARTIAL",
  "one_line_summary": ""
}}

DOCUMENT TO SCORE:
FILENAME: {filename}

{content}"""


TSR_SYSTEM = """You are the Theophysics Quality Director. You evaluate papers written within the Theophysics 
framework against the TSR100 rubric. Score what is present in the text. Be precise, not generous.

You must return ONLY valid JSON, no markdown, no commentary."""

TSR_USER = """Rate this Theophysics paper using the TSR100 rubric.

[30 pts] FRAMEWORK CONTRIBUTION
  Does it advance Theophysics axioms, Master Equation variables, or Ten Laws?
  Does it derive theological commitment from formal reasoning (not retrofit)?
  30 = Major advancement of multiple axioms/conditions
  20 = Solid advancement of one area
  10 = Tangential contribution
  0  = Does not advance the framework

[25 pts] DEDUCTIVE RIGOR
  Is the argument chain math→truth, not apologetics→proof-text?
  Are falsification criteria stated or implied?
  Is there formal structure (definitions, propositions, derivations)?
  25 = Publication-grade rigor with falsification criteria
  15 = Clear logical chain, minor gaps
  8  = Informal but traceable reasoning
  0  = No discernible logical structure

[20 pts] COHERENCE INTEGRATION
  Does it connect to Master Equation / Ten Laws without contradiction?
  Does it reference or build on other framework papers?
  Does it maintain consistency with established axioms?
  20 = Deep integration, extends existing framework
  12 = References framework, no contradictions
  6  = Loosely connected
  0  = Contradicts or ignores framework

[15 pts] COMPLETENESS
  Is this a full argument or a fragment?
  Could it stand alone as a publishable piece?
  15 = Publication-ready or near-ready
  10 = Complete argument, needs polish
  5  = Partial argument, key sections missing
  0  = Fragment or outline only

[10 pts] ORIGINALITY
  Does this say something new within the framework?
  Does it bridge domains that haven't been connected before?
  10 = Novel insight or unprecedented connection
  6  = New angle on known territory
  3  = Competent restatement
  0  = Pure repetition

Respond with EXACTLY this JSON:
{{
  "filename": "{filename}",
  "total_score": 0,
  "framework_contribution": 0,
  "deductive_rigor": 0,
  "coherence_integration": 0,
  "completeness": 0,
  "originality": 0,
  "summary": "Two sentences max.",
  "strongest": "One sentence.",
  "weakest": "One sentence.",
  "tags": [],
  "publication_ready": false
}}

DOCUMENT TO SCORE:
FILENAME: {filename}

{content}"""


ADVERSARIAL_PROMPTS = [
    {
        "role": "system",
        "prompt": "You are a hostile analytic philosopher. Find every logical flaw, circular argument, and unsupported assertion."
    },
    {
        "role": "system", 
        "prompt": "You are an experimental physicist. Identify every empirical claim that lacks data, every prediction that cannot be tested, and every mechanism that violates known physics."
    },
    {
        "role": "system",
        "prompt": "You are a conservative theologian trained at Westminster Seminary. Check every scriptural claim for orthodoxy. Flag eisegesis, proof-texting, and theological novelty that contradicts historic creeds."
    },
    {
        "role": "system",
        "prompt": "You find what others miss. Identify hidden assumptions, unstated premises, circular definitions, and blind spots the author cannot see."
    }
]


# ── Core scoring functions ──────────────────────────────────────────────

def call_openai(client, cfg, system_msg, user_msg, expect_json=True):
    try:
        response = client.chat.completions.create(
            model=cfg.get("MODEL", "gpt-4o"),
            messages=[
                {"role": "system", "content": system_msg},
                {"role": "user", "content": user_msg},
            ],
            temperature=float(cfg.get("TEMPERATURE", 0.3)),
            max_tokens=int(cfg.get("MAX_TOKENS", 2000)),
        )
        text = response.choices[0].message.content.strip()
        if expect_json:
            text = re.sub(r'^```(?:json)?\s*', '', text)
            text = re.sub(r'\s*```$', '', text)
            return json.loads(text)
        return text
    except json.JSONDecodeError as e:
        print(f"  ⚠ JSON parse error: {e}")
        print(f"  Raw response: {text[:200]}...")
        return None
    except Exception as e:
        print(f"  ⚠ API error: {e}")
        return None


def truncate_content(content, max_chars=60000):
    if len(content) > max_chars:
        half = max_chars // 2
        return content[:half] + "\n\n[...TRUNCATED...]\n\n" + content[-half:]
    return content


def score_prosecution(client, cfg, filename, content):
    content = truncate_content(content)
    user_msg = PROSECUTION_USER.format(filename=filename, content=content)
    print(f"  📋 Call 1/1: Prosecution scoring...")
    result = call_openai(client, cfg, PROSECUTION_SYSTEM, user_msg)
    if result and "criteria" in result:
        total = sum(c["score"] for c in result["criteria"].values())
        result["total_score"] = total
        result["percentage"] = round(total / 40 * 100, 1)
        if total >= 35:
            result["verdict"] = "COMPLETE"
        elif total >= 20:
            result["verdict"] = "PARTIAL"
        else:
            result["verdict"] = "INCOMPLETE"
    return result


def score_tsr(client, cfg, filename, content):
    content = truncate_content(content)
    user_msg = TSR_USER.format(filename=filename, content=content)
    print(f"  📋 Call 1/5: TSR scoring...")
    result = call_openai(client, cfg, TSR_SYSTEM, user_msg)
    if result is None:
        return None

    adversarial_results = []
    for i, adv in enumerate(ADVERSARIAL_PROMPTS):
        print(f"  ⚔️  Call {i+2}/5: {adv['role']} adversarial review...")
        adv_user = f"Review this paper and identify all weaknesses from your perspective.\n\nFILENAME: {filename}\n\n{content}"
        adv_result = call_openai(client, cfg, adv['prompt'], adv_user, expect_json=False)
        adversarial_results.append({
            "reviewer": adv["prompt"][:50],
            "critique": adv_result if adv_result else "API call failed"
        })
        time.sleep(0.5)

    result["adversarial_reviews"] = adversarial_results
    return result


# ── Excel output ────────────────────────────────────────────────────────

def write_prosecution_excel(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prosecution Scores"

    headers = ["#", "Filename", "Theory Name", "Total (0-40)", "%",
               "Axioms", "Predictions", "Defeat Cond.", "Empirical",
               "Hard Problem", "Binding", "Moral Consc.", "Causal Chain",
               "Verdict", "Summary"]

    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor='1A3A5C')
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font, c.fill, c.border = hdr_font, hdr_fill, thin
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    red = PatternFill('solid', fgColor='FFE0E0')
    yellow = PatternFill('solid', fgColor='FFFDE0')
    green = PatternFill('solid', fgColor='E0FFE0')

    for idx, r in enumerate(results, 1):
        if r is None:
            continue
        row = idx + 1
        crit = r.get("criteria", {})
        scores = [
            crit.get("formal_axioms", {}).get("score", 0),
            crit.get("quantitative_predictions", {}).get("score", 0),
            crit.get("defeat_conditions", {}).get("score", 0),
            crit.get("empirical_confirmation", {}).get("score", 0),
            crit.get("hard_problem", {}).get("score", 0),
            crit.get("binding_problem", {}).get("score", 0),
            crit.get("moral_consciousness", {}).get("score", 0),
            crit.get("complete_causal_chain", {}).get("score", 0),
        ]

        data = [idx, r.get("_filename", ""), r.get("theory_name", ""),
                r.get("total_score", 0), r.get("percentage", 0)]
        data.extend(scores)
        data.extend([r.get("verdict", ""), r.get("one_line_summary", "")])

        for col, val in enumerate(data, 1):
            cell = ws.cell(row, col, val)
            cell.font = Font(name='Arial', size=9)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            if col >= 6 and col <= 13 and isinstance(val, (int, float)):
                if val <= 1: cell.fill = red
                elif val <= 3: cell.fill = yellow
                else: cell.fill = green

    widths = [5, 35, 30, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 12, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A1:O{len(results)+1}"
    wb.save(output_path)
    print(f"\n📊 Excel saved: {output_path}")


def write_tsr_excel(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TSR100 Scores"

    headers = ["#", "Filename", "Total (0-100)", "Framework (30)",
               "Rigor (25)", "Coherence (20)", "Completeness (15)",
               "Originality (10)", "Pub Ready", "Summary", "Strongest", "Weakest"]

    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor='1A3A5C')
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font, c.fill, c.border = hdr_font, hdr_fill, thin
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for idx, r in enumerate(results, 1):
        if r is None:
            continue
        row = idx + 1
        data = [
            idx, r.get("filename", ""), r.get("total_score", 0),
            r.get("framework_contribution", 0), r.get("deductive_rigor", 0),
            r.get("coherence_integration", 0), r.get("completeness", 0),
            r.get("originality", 0), str(r.get("publication_ready", False)),
            r.get("summary", ""), r.get("strongest", ""), r.get("weakest", ""),
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row, col, val)
            cell.font = Font(name='Arial', size=9)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    widths = [5, 35, 10, 10, 10, 10, 10, 10, 8, 40, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"\n📊 Excel saved: {output_path}")


# ── Main ────────────────────────────────────────────────────────────────

def collect_files(input_path):
    p = Path(input_path)
    if p.is_file():
        return [p]
    files = []
    for ext in ["*.md", "*.txt"]:
        files.extend(sorted(p.glob(ext)))
    if not files:
        for ext in ["*.md", "*.txt"]:
            files.extend(sorted(p.rglob(ext)))
    return files


def main():
    parser = argparse.ArgumentParser(description="Lowe Standard Full Scorer v2.0")
    parser.add_argument("--mode", choices=["prosecution", "tsr"], required=True)
    parser.add_argument("--input", required=True, help="File or folder to score")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx name/path, CONSTRAINED under EXPORTS\\ "
                             "(relative resolves there; outside paths are redirected "
                             "into EXPORTS\\excel\\). JSON sidecar written alongside.")
    parser.add_argument("--config", default=None, help="Config file path")
    parser.add_argument("--limit", type=int, default=0, help="Max files to process (0=all)")
    args = parser.parse_args()

    cfg = load_config(args.config)
    if cfg.get("OPENAI_API_KEY", "") in PLACEHOLDER_KEYS:
        print("❌ No API key found. Either:")
        print("   • set the OPENAI_API_KEY environment variable, or")
        print("   • copy config.template.txt -> config.txt and paste your key.")
        print("   (config.txt is gitignored — never commit or share it.)")
        sys.exit(1)

    client = get_client(cfg)
    files = collect_files(args.input)
    if args.limit > 0:
        files = files[:args.limit]

    print(f"\n{'='*60}")
    print(f"  LOWE STANDARD FULL SCORER v2.0")
    print(f"  Mode: {args.mode.upper()}")
    print(f"  Files: {len(files)}")
    print(f"  Model: {cfg.get('MODEL', 'gpt-4o')}")
    print(f"  Gateway: {cfg.get('BASE_URL', 'direct')}")
    print(f"{'='*60}\n")

    results = []
    for i, fpath in enumerate(files, 1):
        print(f"\n[{i}/{len(files)}] {fpath.name}")
        try:
            content = fpath.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            print(f"  ⚠ Read error: {e}")
            results.append(None)
            continue

        if len(content.strip()) < 50:
            print(f"  ⚠ Skipping (too short: {len(content)} chars)")
            results.append(None)
            continue

        if args.mode == "prosecution":
            result = score_prosecution(client, cfg, fpath.name, content)
        else:
            result = score_tsr(client, cfg, fpath.name, content)

        if result:
            result["_filename"] = fpath.name
            result["_path"] = str(fpath)
            results.append(result)
            if args.mode == "prosecution":
                v = result.get("verdict", "?")
                s = result.get("total_score", 0)
                print(f"  ✅ Score: {s}/40 ({result.get('percentage',0)}%) — {v}")
            else:
                s = result.get("total_score", 0)
                print(f"  ✅ Score: {s}/100")
        else:
            results.append(None)
            print(f"  ❌ Scoring failed")

        time.sleep(1)

    # Save results
    valid = [r for r in results if r is not None]
    if not valid:
        print("\n❌ No valid results to save.")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # All exports go to the station-root EXPORTS folder ONLY.
    exports = Path(__file__).resolve().parent / "EXPORTS"
    if args.output:
        # --output is CONSTRAINED under EXPORTS (no bypass / escape hatch).
        # Relative paths resolve inside EXPORTS; any path that would land
        # outside EXPORTS is redirected into EXPORTS\excel by basename.
        requested = Path(args.output)
        out = requested if requested.is_absolute() else (exports / requested)
        out = out.resolve()
        if not out.is_relative_to(exports.resolve()):
            print(f"⚠ --output '{args.output}' is outside EXPORTS; "
                  f"redirecting into EXPORTS\\excel\\{out.name}")
            out = exports / "excel" / out.name
        out.parent.mkdir(parents=True, exist_ok=True)
        xlsx_path = str(out)
        json_path = str(out.with_suffix(".json"))
    else:
        excel_dir = exports / "excel"
        json_dir = exports / "json"
        excel_dir.mkdir(parents=True, exist_ok=True)
        json_dir.mkdir(parents=True, exist_ok=True)
        base = f"{args.mode}_scores_{timestamp}"
        xlsx_path = str(excel_dir / f"{base}.xlsx")
        json_path = str(json_dir / f"{base}.json")

    if args.mode == "prosecution":
        write_prosecution_excel(valid, xlsx_path)
    else:
        write_tsr_excel(valid, xlsx_path)

    # Also save raw JSON (path computed above alongside EXPORTS layout)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(valid, f, indent=2, ensure_ascii=False)
    print(f"📄 JSON saved: {json_path}")

    # Print summary
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    if args.mode == "prosecution":
        scores = [r["total_score"] for r in valid]
        verdicts = [r.get("verdict", "?") for r in valid]
        print(f"  Papers scored: {len(valid)}")
        print(f"  Average: {sum(scores)/len(scores):.1f}/40")
        print(f"  COMPLETE: {verdicts.count('COMPLETE')}")
        print(f"  PARTIAL:  {verdicts.count('PARTIAL')}")
        print(f"  INCOMPLETE: {verdicts.count('INCOMPLETE')}")
    else:
        scores = [r["total_score"] for r in valid]
        print(f"  Papers scored: {len(valid)}")
        print(f"  Average: {sum(scores)/len(scores):.1f}/100")
        print(f"  Highest: {max(scores)}/100")
        print(f"  Lowest:  {min(scores)}/100")
        pub = sum(1 for r in valid if r.get("publication_ready"))
        print(f"  Publication ready: {pub}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
