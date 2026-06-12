from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any


Q_FIELDS = [
    ("Q1_identity", "Identity"),
    ("Q2_scope", "Scope"),
    ("Q3_mechanism", "Mechanism"),
    ("Q4_evidence", "Evidence"),
    ("Q5_falsifiability", "Falsifiability"),
    ("Q6_boundary", "Boundary"),
    ("Q7_listener_risk", "Listener Risk"),
]
WORKFLOW_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = WORKFLOW_ROOT / "paper-grade-dashboard-template.html"


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def find_snapshot_for_grade(grade_path: Path, snapshot_root: Path | None) -> Path | None:
    if not snapshot_root or not snapshot_root.exists():
        return None
    stem = grade_path.name.replace(".paper-grade.json", "")
    candidates = list(snapshot_root.rglob(f"{stem}.paper-snapshot.json"))
    return candidates[0] if candidates else None


def semantic_manifest_lookup(manifest_path: Path | None, grade_path: Path) -> dict[str, Any] | None:
    if not manifest_path or not manifest_path.exists():
        return None
    manifest = load_json(manifest_path)
    grade_key = str(grade_path.resolve()).lower()
    grade_name = grade_path.name.lower()
    for row in manifest.get("rows") or []:
        source = str(row.get("source") or "")
        if str(Path(source).resolve()).lower() == grade_key or Path(source).name.lower() == grade_name:
            return row
    return None


def semantic_row_as_snapshot(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    filename_safe = row.get("filename_safe") or ""
    parts = filename_safe.split("__")
    return {
        "semantic_address": {
            "address": row.get("address", ""),
            "filename_safe": filename_safe,
            "domain": parts[0] if parts else "",
        },
        "semantic_vector": {"vector": row.get("vector", "")},
        "semantic_hash": {"hash": row.get("hash", "")},
        "epistemic_status": {},
        "math_translation_layer": {},
        "station_marks": [],
    }


def claim_distribution(claims: list[dict[str, Any]]) -> Counter:
    return Counter(str(c.get("claim_maturity_label") or "Unlabeled") for c in claims)


def q_coverage(claims: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    total = max(len(claims), 1)
    for field, label in Q_FIELDS:
        values = Counter(str(c.get(field) or "missing") for c in claims)
        weak = values.get("missing", 0) + values.get("broad", 0) + values.get("implicit", 0)
        if field == "Q7_listener_risk":
            weak = values.get("high", 0)
        rows.append(
            {
                "field": field,
                "label": label,
                "values": values,
                "weak": weak,
                "coverage_pct": round(100 * (total - weak) / total, 1),
            }
        )
    return rows


def issue_summary(claims: list[dict[str, Any]]) -> Counter:
    issues: Counter = Counter()
    for claim in claims:
        for field in ("Q3_mechanism", "Q4_evidence", "Q5_falsifiability", "Q6_boundary"):
            if claim.get(field) == "missing":
                issues[f"weak:{field}"] += 1
        if claim.get("Q2_scope") == "broad":
            issues["weak:Q2_scope"] += 1
        if claim.get("Q1_identity") == "implicit":
            issues["weak:Q1_identity"] += 1
        if claim.get("Q7_listener_risk") == "high":
            issues["risk:Q7_listener_risk"] += 1
    return issues


def evidence_distribution(claims: list[dict[str, Any]]) -> Counter:
    out: Counter = Counter()
    for claim in claims:
        bar = str(claim.get("evidence_bar") or "").strip()
        if not bar or bar == "No explicit evidence marker in sentence.":
            out["no explicit marker"] += 1
            continue
        for item in bar.split(","):
            clean = item.strip().lower()
            if clean:
                out[clean] += 1
    return out


def esc(value: Any) -> str:
    return escape(str(value if value is not None else ""), quote=True)


def excel_safe(value: Any) -> Any:
    if value is None or isinstance(value, (int, float, bool)):
        return value
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", str(value))


def append_safe(ws: Any, row: list[Any]) -> None:
    ws.append([excel_safe(value) for value in row])


def pct(value: int, total: int) -> str:
    if total <= 0:
        return "0%"
    return f"{round(100 * value / total, 1)}%"


def pct_number(value: int, total: int) -> str:
    if total <= 0:
        return "0"
    return str(round(100 * value / total, 1))


def css_token(value: Any, default: str = "unknown") -> str:
    token = re.sub(r"[^a-z0-9]+", "-", str(value or "").lower()).strip("-")
    return token or default


def fill_slots(fragment: str, values: dict[str, Any]) -> str:
    out = fragment
    for key, value in values.items():
        out = out.replace("{{" + key + "}}", esc(value))
    return out


def replace_slot_block(template: str, slot_name: str, rendered: str) -> str:
    pattern = re.compile(
        rf"(?P<start>[ \t]*<!--\s*[^>]*TEMPLATE:SLOT:{re.escape(slot_name)}[^>]*-->\s*)"
        rf"(?P<body>.*?)"
        rf"(?P<end>[ \t]*<!--\s*END:TEMPLATE:SLOT:{re.escape(slot_name)}\s*-->)",
        re.DOTALL,
    )

    def repl(match: re.Match[str]) -> str:
        indent = re.match(r"[ \t]*", match.group("start")).group(0)
        clean = rendered.rstrip()
        if clean:
            clean = "\n".join(indent + line if line else line for line in clean.splitlines()) + "\n"
        return f"{match.group('start')}{clean}{indent}{match.group('end')}"

    return pattern.sub(repl, template, count=1)


def q_fill_class(coverage_pct: float) -> str:
    if coverage_pct >= 70:
        return "high"
    if coverage_pct >= 40:
        return "mid"
    return "low"


def maturity_class(label: Any) -> str:
    lower = str(label or "").lower()
    if "proof" in lower:
        return "mat-proof"
    if "formal" in lower:
        return "mat-formal"
    if "structural" in lower:
        return "mat-structural"
    if "metaphor" in lower:
        return "mat-metaphor"
    return "mat-unknown"


def grade_from_claims(claims: list[dict[str, Any]]) -> tuple[str, str]:
    if not claims:
        return "N/A", "No claim candidates"
    levels: list[float] = []
    for claim in claims:
        try:
            levels.append(float(claim.get("claim_maturity_level") or 0))
        except (TypeError, ValueError):
            continue
    avg = sum(levels) / len(levels) if levels else 0
    if avg >= 6:
        return "A", f"avg maturity {avg:.1f}/7"
    if avg >= 5:
        return "B", f"avg maturity {avg:.1f}/7"
    if avg >= 4:
        return "C", f"avg maturity {avg:.1f}/7"
    if avg >= 3:
        return "D", f"avg maturity {avg:.1f}/7"
    return "F", f"avg maturity {avg:.1f}/7"


def extract_kill_conditions(claims: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for claim in claims:
        text = str(claim.get("kill_conditions") or "").strip()
        if not text or text.lower() in {"missing", "none", "n/a"} or text in seen:
            continue
        seen.add(text)
        rows.append(
            {
                "KILL_ID": f"K{len(rows) + 1:02d}",
                "KILL_STATUS": "Open",
                "KILL_STATUS_CLASS": "open",
                "KILL_TITLE": text[:80],
                "KILL_DESCRIPTION": text,
                "KILL_EVIDENCE": str(claim.get("evidence_bar") or "Evidence not marked."),
            }
        )
        if len(rows) >= 6:
            break
    if rows:
        return rows
    return [
        {
            "KILL_ID": "K00",
            "KILL_STATUS": "Missing",
            "KILL_STATUS_CLASS": "missing",
            "KILL_TITLE": "No explicit kill condition extracted",
            "KILL_DESCRIPTION": "The paper-grader did not extract a deterministic falsification test for this article.",
            "KILL_EVIDENCE": "Repair target: add claim-specific failure conditions.",
        }
    ]


def fruit_rows() -> list[dict[str, str]]:
    fruits = [
        ("love", "Love", "agape", "V(r) at minimum"),
        ("joy", "Joy", "chara", "omega = omega0 resonance"),
        ("peace", "Peace", "eirene", "sum F = 0"),
        ("patience", "Patience", "hypomone", "C = dQ/dT high"),
        ("kindness", "Kindness", "chrestotes", "E_activation -> min"),
        ("goodness", "Goodness", "agathosyne", "W_out > W_in"),
        ("faithfulness", "Faithfulness", "pistis", "dL/dt = 0"),
        ("gentleness", "Gentleness", "prautes", "F = k*F_needed, k <= 1"),
        ("self-control", "Self-Control", "enkrateia", "u(t) = -K*e(t) PID"),
    ]
    return [
        {
            "FRUIT_ID": fid,
            "FRUIT_NAME": name,
            "FRUIT_GREEK": greek,
            "FRUIT_EQUATION": equation,
            "FRUIT_STATUS": "Reference",
            "FRUIT_STATUS_CLASS": "reference",
        }
        for fid, name, greek, equation in fruits
    ]


def build_template_html(grade: dict[str, Any], snapshot: dict[str, Any] | None, source_grade: Path, source_snapshot: Path | None) -> str | None:
    if not TEMPLATE_PATH.exists():
        return None

    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    metrics = grade.get("metrics") or {}
    claims = grade.get("claims") or []
    equations = grade.get("equations") or []
    sections = grade.get("sections") or []
    dist = claim_distribution(claims)
    q_rows = q_coverage(claims)
    issues = issue_summary(claims)
    evidence = evidence_distribution(claims)
    bits = snapshot_bits(snapshot)
    article_title = str(grade.get("paper_id") or source_grade.name.replace(".paper-grade.json", ""))
    article_slug = clean_slug(article_title)
    series_code = article_slug.split("-")[0].upper() if "-" in article_slug else "AXIOMS"
    generated = datetime.now().isoformat(timespec="seconds")
    article_grade, grade_confidence = grade_from_claims(claims)
    kill_rows = extract_kill_conditions(claims)

    q_fragment = "\n".join(
        fill_slots(
            """<div class="pg-7q-item" data-component="7q-item" data-name="7q-{{Q_ID}}">
  <div class="pg-7q-label">{{Q_LABEL}}</div>
  <div class="pg-7q-values">{{Q_OBSERVED_VALUES}}</div>
  <div class="pg-7q-bar"><div class="pg-7q-fill {{Q_FILL_CLASS}}" style="width:{{Q_COVERAGE_PCT}}%"></div></div>
  <div class="pg-7q-pct">{{Q_COVERAGE_PCT}}% coverage</div>
  <div class="pg-7q-weak">{{Q_WEAK_COUNT}} weak</div>
</div>""",
            {
                "Q_ID": row["field"],
                "Q_LABEL": row["label"],
                "Q_OBSERVED_VALUES": ", ".join(f"{k}: {v}" for k, v in row["values"].most_common()),
                "Q_FILL_CLASS": q_fill_class(float(row["coverage_pct"])),
                "Q_COVERAGE_PCT": row["coverage_pct"],
                "Q_WEAK_COUNT": row["weak"],
            },
        )
        for row in q_rows
    )
    maturity_fragment = "\n".join(
        fill_slots(
            """<div class="pg-mat-row" data-component="mat-row" data-name="mat-{{MATURITY_LEVEL}}">
  <span class="pg-mat-row-label">{{MATURITY_LABEL}}</span>
  <span><span class="pg-mat-row-val">{{MATURITY_COUNT}}</span><span class="pg-mat-row-pct">{{MATURITY_PCT}}%</span></span>
</div>""",
            {
                "MATURITY_LEVEL": css_token(label),
                "MATURITY_LABEL": label,
                "MATURITY_COUNT": count,
                "MATURITY_PCT": pct_number(count, len(claims)),
            },
        )
        for label, count in dist.most_common()
    ) or "<div class=\"pg-mat-row\"><span class=\"pg-mat-row-label\">No claims</span><span class=\"pg-mat-row-val\">0</span></div>"
    evidence_fragment = "\n".join(
        fill_slots(
            """<div class="pg-mat-row" data-component="ev-row" data-name="ev-{{MARKER_TYPE}}">
  <span class="pg-mat-row-label">{{MARKER_LABEL}}</span>
  <span class="pg-mat-row-val">{{MARKER_COUNT}}</span>
</div>""",
            {"MARKER_TYPE": css_token(label), "MARKER_LABEL": label, "MARKER_COUNT": count},
        )
        for label, count in evidence.most_common(12)
    ) or "<div class=\"pg-mat-row\"><span class=\"pg-mat-row-label\">No explicit marker</span><span class=\"pg-mat-row-val\">0</span></div>"
    pressure_fragment = "\n".join(
        fill_slots(
            """<div class="pg-pressure-card" data-component="pressure-card" data-name="pressure-{{Q_ID}}">
  <span class="pg-pressure-name">{{PRESSURE_LABEL}}</span>
  <span class="pg-pressure-count">{{PRESSURE_COUNT}}</span>
</div>""",
            {"Q_ID": css_token(name), "PRESSURE_LABEL": name, "PRESSURE_COUNT": count},
        )
        for name, count in issues.most_common(8)
    ) or "<div class=\"pg-pressure-card\"><span class=\"pg-pressure-name\">No deterministic weak flags</span><span class=\"pg-pressure-count\">0</span></div>"
    claim_fragment = "\n".join(
        fill_slots(
            """<tr data-component="claim-row" data-name="claim-{{CLAIM_NUMBER}}">
  <td>{{CLAIM_NUMBER}}</td>
  <td>{{CLAIM_SECTION}}</td>
  <td>{{CLAIM_TEXT}}</td>
  <td><span class="mat-badge {{MATURITY_CLASS}}">{{MATURITY_LABEL}}</span><br/><span class="ev-note">Level {{MATURITY_LEVEL_NUM}}</span></td>
  <td><span class="ev-note">{{EVIDENCE_TEXT}}</span></td>
  <td><span class="kill-note">{{KILL_CONDITION}}</span></td>
  <td><span class="ev-note">{{PROOF_BOUNDARY}}</span></td>
</tr>""",
            {
                "CLAIM_NUMBER": idx,
                "CLAIM_SECTION": claim.get("section", ""),
                "CLAIM_TEXT": claim.get("one_sentence_claim", ""),
                "MATURITY_CLASS": maturity_class(claim.get("claim_maturity_label")),
                "MATURITY_LABEL": claim.get("claim_maturity_label", ""),
                "MATURITY_LEVEL_NUM": claim.get("claim_maturity_level", ""),
                "EVIDENCE_TEXT": claim.get("evidence_bar", ""),
                "KILL_CONDITION": claim.get("kill_conditions", ""),
                "PROOF_BOUNDARY": claim.get("proof_boundary", ""),
            },
        )
        for idx, claim in enumerate(claims[:40], 1)
    ) or "<tr><td colspan=\"7\">No claim candidates detected.</td></tr>"
    law_fragment = "\n".join(
        f'<span class="pg-deriv-node unknown" style="font-size:.55rem;padding:.2rem .45rem" data-component="law-badge" data-name="law-{n}">L{n}</span>'
        for n in range(1, 11)
    )
    fruit_fragment = "\n".join(
        fill_slots(
            """<div class="pg-fruit-card" data-component="fruit-card" data-name="fruit-{{FRUIT_ID}}">
  <div class="pg-fruit-name">{{FRUIT_NAME}}</div>
  <div class="pg-fruit-greek">{{FRUIT_GREEK}}</div>
  <div class="pg-fruit-eq">{{FRUIT_EQUATION}}</div>
  <span class="pg-fruit-status {{FRUIT_STATUS_CLASS}}">{{FRUIT_STATUS}}</span>
</div>""",
            row,
        )
        for row in fruit_rows()
    )
    kill_fragment = "\n".join(
        fill_slots(
            """<div class="pg-kill-card" data-component="kill-card" data-name="kill-{{KILL_ID}}">
  <div class="pg-kill-row">
    <span class="pg-kill-code">{{KILL_ID}}</span>
    <span class="pg-kill-status {{KILL_STATUS_CLASS}}">{{KILL_STATUS}}</span>
  </div>
  <h3>{{KILL_TITLE}}</h3>
  <p>{{KILL_DESCRIPTION}}</p>
  <div class="pg-kill-foot">{{KILL_EVIDENCE}}</div>
</div>""",
            row,
        )
        for row in kill_rows
    )
    station_marks = bits.get("station_marks", [])
    station_fragment = "\n".join(
        fill_slots(
            """<tr data-component="station-row" data-name="station-{{STATION_ID}}">
  <td>{{STATION_NAME}}</td>
  <td>{{STATION_STATUS}}</td>
  <td>{{STATION_WARNINGS}}</td>
</tr>""",
            {
                "STATION_ID": s.get("station_id", ""),
                "STATION_NAME": s.get("station_id", ""),
                "STATION_STATUS": s.get("status", ""),
                "STATION_WARNINGS": ", ".join(s.get("warnings") or []),
            },
        )
        for s in station_marks
    ) or "<tr><td colspan=\"3\">No snapshot station marks loaded.</td></tr>"

    rendered = template
    for slot_name, fragment in {
        "7q-item": q_fragment,
        "maturity-row": maturity_fragment,
        "evidence-marker-row": evidence_fragment,
        "pressure-card": pressure_fragment,
        "claim-row": claim_fragment,
        "law-badge": law_fragment,
        "fruit-card": fruit_fragment,
        "kill-card": kill_fragment,
        "station-row": station_fragment,
    }.items():
        rendered = replace_slot_block(rendered, slot_name, fragment)

    scalar_values = {
        "ARTICLE_TITLE": article_title,
        "ARTICLE_SLUG": article_slug,
        "ARTICLE_GRADE": article_grade,
        "GRADE_CONFIDENCE": grade_confidence,
        "SERIES_CODE": series_code,
        "SERIES_SLUG": clean_slug(series_code),
        "PUBLISH_DATE": generated[:10],
        "MODIFY_DATE": generated[:10],
        "GENERATION_TIMESTAMP": generated,
        "WORD_COUNT": metrics.get("word_count", 0),
        "SECTION_COUNT": metrics.get("section_count", len(sections)),
        "EQUATION_COUNT": metrics.get("equation_count", len(equations)),
        "CLAIM_COUNT": metrics.get("claim_candidate_count", len(claims)),
        "EVIDENCE_MARKED": sum(count for label, count in evidence.items() if label != "no explicit marker"),
        "KILL_COUNT": len(kill_rows),
        "MATURITY_LEVEL": "",
        "MATURITY_LABEL": "",
        "MATURITY_COUNT": "",
        "MATURITY_PCT": "",
        "MARKER_TYPE": "",
        "MARKER_LABEL": "",
        "MARKER_COUNT": "",
        "Q_ID": "",
        "Q_LABEL": "",
        "Q_OBSERVED_VALUES": "",
        "Q_FILL_CLASS": "",
        "Q_COVERAGE_PCT": "",
        "Q_WEAK_COUNT": "",
        "PRESSURE_LABEL": "",
        "PRESSURE_COUNT": "",
        "SEMANTIC_ADDRESS": bits.get("address") or "not present in loaded snapshot",
        "FILENAME_SAFE": bits.get("filename_safe") or article_slug,
        "VECTOR_HASH": bits.get("vector") or "not present",
        "HASH_SHORT": str(bits.get("hash") or "not present")[:12],
        "RIGOR_VERDICT": bits.get("rigor_verdict") or "not loaded",
        "MATH_LAYER": bits.get("math_status") or "not loaded",
        "TRANSLATED_SPANS": bits.get("translated_spans", 0),
        "SNAPSHOT_STATUS": "loaded" if snapshot else "not loaded",
        "SOURCE_JSON_PATH": source_grade,
        "DERIV_TRUTH_CLASS": "unknown",
        "DERIV_OBSERVATION_CLASS": "unknown",
        "DERIV_EXISTENCE_CLASS": "unknown",
        "DERIV_GROUNDING_CLASS": "unknown",
        "DERIV_SPIRITUAL_CLASS": "unknown",
        "DERIV_MORAL_CLASS": "unknown",
        "DERIV_GRACE_CLASS": "unknown",
        "LAW_NUMBER": "",
        "LAW_COVERAGE_CLASS": "",
        "FRUIT_ID": "",
        "FRUIT_NAME": "",
        "FRUIT_GREEK": "",
        "FRUIT_EQUATION": "",
        "FRUIT_STATUS": "",
        "FRUIT_STATUS_CLASS": "",
        "KILL_ID": "",
        "KILL_STATUS": "",
        "KILL_STATUS_CLASS": "",
        "KILL_TITLE": "",
        "KILL_DESCRIPTION": "",
        "KILL_EVIDENCE": "",
        "CLAIM_NUMBER": "",
        "CLAIM_SECTION": "",
        "CLAIM_TEXT": "",
        "MATURITY_CLASS": "",
        "MATURITY_LEVEL_NUM": "",
        "EVIDENCE_TEXT": "",
        "KILL_CONDITION": "",
        "PROOF_BOUNDARY": "",
        "STATION_ID": "",
        "STATION_NAME": "",
        "STATION_STATUS": "",
        "STATION_WARNINGS": "",
        "SLOT_NAME": "",
    }
    rendered = fill_slots(rendered, scalar_values)
    rendered = re.sub(r"\{\{[A-Z0-9_]+\}\}", "", rendered)
    return rendered


def clean_slug(value: Any) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", str(value or "").lower()).strip("-")
    return slug or "paper"


def snapshot_bits(snapshot: dict[str, Any] | None) -> dict[str, Any]:
    if not snapshot:
        return {}
    addr = snapshot.get("semantic_address") or {}
    vec = snapshot.get("semantic_vector") or {}
    hsh = snapshot.get("semantic_hash") or {}
    epi = snapshot.get("epistemic_status") or {}
    math_layer = snapshot.get("math_translation_layer") or {}
    return {
        "address": addr.get("address", ""),
        "filename_safe": addr.get("filename_safe", ""),
        "domain": addr.get("domain", ""),
        "vector": vec.get("vector", ""),
        "hash": hsh.get("hash", ""),
        "rigor_verdict": epi.get("rigor_verdict", ""),
        "overall_tier": epi.get("overall_tier", ""),
        "math_status": math_layer.get("status", ""),
        "translated_spans": len(math_layer.get("translated_spans") or []),
        "station_marks": snapshot.get("station_marks") or [],
    }


def build_html(grade: dict[str, Any], snapshot: dict[str, Any] | None, source_grade: Path, source_snapshot: Path | None) -> str:
    template_html = build_template_html(grade, snapshot, source_grade, source_snapshot)
    if template_html is not None:
        return template_html

    metrics = grade.get("metrics") or {}
    claims = grade.get("claims") or []
    equations = grade.get("equations") or []
    sections = grade.get("sections") or []
    dist = claim_distribution(claims)
    q_rows = q_coverage(claims)
    issues = issue_summary(claims)
    evidence = evidence_distribution(claims)
    bits = snapshot_bits(snapshot)

    dist_rows = "\n".join(
        f"<tr><td>{esc(label)}</td><td>{count}</td><td>{pct(count, len(claims))}</td></tr>"
        for label, count in dist.most_common()
    )
    q_html = "\n".join(
        "<tr>"
        f"<td><b>{esc(row['label'])}</b><span>{esc(row['field'])}</span></td>"
        f"<td>{esc(', '.join(f'{k}: {v}' for k, v in row['values'].most_common()))}</td>"
        f"<td>{row['weak']}</td>"
        f"<td><div class='bar'><i style='width:{row['coverage_pct']}%'></i></div><em>{row['coverage_pct']}%</em></td>"
        "</tr>"
        for row in q_rows
    )
    issue_cards = "\n".join(
        f"<div class='issue'><strong>{esc(name)}</strong><span>{count}</span></div>"
        for name, count in issues.most_common(8)
    ) or "<div class='issue'><strong>No deterministic weak flags</strong><span>0</span></div>"
    evidence_rows = "\n".join(
        f"<tr><td>{esc(label)}</td><td>{count}</td></tr>"
        for label, count in evidence.most_common(12)
    )
    station_rows = "\n".join(
        f"<tr><td>{esc(s.get('station_id'))}</td><td>{esc(s.get('status'))}</td><td>{esc(', '.join(s.get('warnings') or []))}</td></tr>"
        for s in bits.get("station_marks", [])
    ) or "<tr><td colspan='3'>No snapshot station marks loaded.</td></tr>"

    claim_rows = "\n".join(
        "<tr>"
        f"<td>{idx}</td>"
        f"<td>{esc(claim.get('section'))}</td>"
        f"<td>{esc(claim.get('one_sentence_claim'))}</td>"
        f"<td><b>{esc(claim.get('claim_maturity_label'))}</b><span>Level {esc(claim.get('claim_maturity_level'))}</span></td>"
        f"<td>{esc(claim.get('evidence_bar'))}</td>"
        f"<td>{esc(claim.get('kill_conditions'))}</td>"
        f"<td>{esc(claim.get('proof_boundary'))}</td>"
        "</tr>"
        for idx, claim in enumerate(claims[:40], 1)
    )

    source_snapshot_text = str(source_snapshot) if source_snapshot else "not loaded"
    generated = datetime.now().isoformat(timespec="seconds")
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Paper Grade Dashboard - {esc(grade.get('paper_id'))}</title>
<style>
:root {{
  --ink:#e9dfc2;
  --paper:#10120f;
  --paper2:#171a15;
  --line:#4a3a16;
  --deep:#050605;
  --gold:#c79a2c;
  --green:#4fb273;
  --red:#d6674f;
  --blue:#6fb8db;
  --muted:#a89667;
  --shadow:0 20px 70px rgba(0,0,0,.45);
}}
* {{ box-sizing:border-box; }}
body {{
  margin:0;
  color:var(--ink);
  font-family: "Iowan Old Style", "Palatino Linotype", Georgia, serif;
  background:
    radial-gradient(circle at top left, rgba(199,154,44,.18), transparent 34rem),
    linear-gradient(135deg, #060707 0%, #17130b 55%, #262014 100%);
}}
main {{ max-width:1280px; margin:0 auto; padding:38px 22px 80px; }}
.hero {{
  border:1px solid rgba(199,154,44,.24);
  background:rgba(10,11,10,.9);
  border-radius:28px;
  padding:30px;
  box-shadow:var(--shadow);
  position:relative;
  overflow:hidden;
}}
.hero:after {{
  content:"";
  position:absolute;
  right:-80px; top:-110px;
  width:340px; height:340px;
  border-radius:50%;
  border:38px solid rgba(199,154,44,.12);
}}
.eyebrow {{ color:var(--gold); letter-spacing:.16em; text-transform:uppercase; font-size:.76rem; font-weight:800; }}
h1 {{ margin:.35rem 0 .45rem; font-size:clamp(2.1rem,4vw,4.8rem); line-height:.92; max-width:960px; color:#fff6dc; }}
.sub {{ color:var(--muted); max-width:940px; line-height:1.65; }}
.grid {{ display:grid; gap:16px; }}
.metrics {{ grid-template-columns:repeat(4,minmax(0,1fr)); margin:22px 0; }}
.metric, .panel, .issue {{
  background:rgba(255,255,255,.03);
  border:1px solid rgba(199,154,44,.18);
  border-radius:20px;
  box-shadow:0 10px 30px rgba(33,24,8,.08);
}}
.metric {{ padding:18px; min-height:112px; }}
.metric strong {{ display:block; font-size:2rem; line-height:1; color:#fff0c5; }}
.metric span {{ color:var(--muted); font-size:.82rem; text-transform:uppercase; letter-spacing:.09em; }}
.two {{ grid-template-columns:1.1fr .9fr; align-items:start; margin-top:16px; }}
.panel {{ padding:20px; overflow:hidden; }}
h2 {{ margin:.2rem 0 1rem; color:#fff0c5; font-size:1.35rem; }}
table {{ width:100%; border-collapse:collapse; font-size:.92rem; }}
th,td {{ border-bottom:1px solid rgba(199,154,44,.12); padding:11px 9px; vertical-align:top; text-align:left; }}
th {{ color:var(--gold); font-size:.72rem; text-transform:uppercase; letter-spacing:.1em; }}
td span {{ display:block; color:var(--muted); font-size:.78rem; margin-top:.25rem; }}
.bar {{ display:inline-block; width:120px; height:10px; background:rgba(199,154,44,.14); border-radius:999px; overflow:hidden; vertical-align:middle; margin-right:8px; }}
.bar i {{ display:block; height:100%; background:linear-gradient(90deg,var(--red),var(--gold),var(--green)); border-radius:999px; }}
em {{ color:var(--muted); font-style:normal; }}
.issues {{ grid-template-columns:repeat(4,minmax(0,1fr)); }}
.issue {{ padding:14px; display:flex; justify-content:space-between; gap:12px; align-items:center; }}
.issue strong {{ font-size:.84rem; color:var(--muted); }}
.issue span {{ font-size:1.45rem; color:var(--red); font-weight:900; }}
.address {{
  font-family:"Cascadia Mono","Consolas",monospace;
  font-size:.78rem;
  line-height:1.65;
  color:#e9dfc2;
  background:rgba(199,154,44,.06);
  border:1px dashed rgba(199,154,44,.28);
  border-radius:16px;
  padding:14px;
  overflow-wrap:anywhere;
}}
.claims td:nth-child(3) {{ min-width:320px; }}
.footer {{ margin-top:18px; color:var(--muted); font-size:.82rem; line-height:1.55; }}
@media (max-width:900px) {{ .metrics,.two,.issues {{ grid-template-columns:1fr; }} .hero {{ padding:22px; border-radius:20px; }} }}
</style>
</head>
<body>
<main>
  <section class="hero">
    <div class="eyebrow">Paper Grader Surface / Canary</div>
    <h1>{esc(grade.get('paper_id'))}</h1>
    <p class="sub">This dashboard pulls the deterministic paper-grader metrics into a review surface: claims, equations, 7Q gaps, evidence markers, maturity distribution, semantic address, and station status. It does not replace the canonical JSON/XLSX; it makes the audit readable.</p>
    <section class="grid metrics">
      <div class="metric"><strong>{esc(metrics.get('word_count', 0))}</strong><span>Words</span></div>
      <div class="metric"><strong>{esc(metrics.get('section_count', len(sections)))}</strong><span>Sections</span></div>
      <div class="metric"><strong>{esc(metrics.get('equation_count', len(equations)))}</strong><span>Equations</span></div>
      <div class="metric"><strong>{esc(metrics.get('claim_candidate_count', len(claims)))}</strong><span>Claim Candidates</span></div>
    </section>
    <div class="address">
      <b>Semantic Address:</b> {esc(bits.get('address') or 'not present in loaded snapshot')}<br>
      <b>Filename Safe:</b> {esc(bits.get('filename_safe') or 'not present')}<br>
      <b>Vector / Hash:</b> {esc(bits.get('vector'))} / {esc(bits.get('hash'))}<br>
      <b>Rigor Verdict:</b> {esc(bits.get('rigor_verdict'))} &nbsp; <b>Math Layer:</b> {esc(bits.get('math_status'))} ({esc(bits.get('translated_spans'))} translated spans)
    </div>
  </section>

  <section class="grid two">
    <div class="panel">
      <h2>7Q Coverage</h2>
      <table><thead><tr><th>Question</th><th>Observed Values</th><th>Weak</th><th>Coverage</th></tr></thead><tbody>{q_html}</tbody></table>
    </div>
    <div class="panel">
      <h2>Maturity Distribution</h2>
      <table><thead><tr><th>Level</th><th>Claims</th><th>Share</th></tr></thead><tbody>{dist_rows}</tbody></table>
      <h2 style="margin-top:1.4rem;">Evidence Markers</h2>
      <table><thead><tr><th>Marker</th><th>Claims</th></tr></thead><tbody>{evidence_rows}</tbody></table>
    </div>
  </section>

  <section class="panel" style="margin-top:16px;">
    <h2>Primary Repair Pressure</h2>
    <div class="grid issues">{issue_cards}</div>
  </section>

  <section class="panel" style="margin-top:16px;">
    <h2>Station Marks</h2>
    <table><thead><tr><th>Station</th><th>Status</th><th>Warnings</th></tr></thead><tbody>{station_rows}</tbody></table>
  </section>

  <section class="panel" style="margin-top:16px;">
    <h2>Claim Audit Preview</h2>
    <table class="claims"><thead><tr><th>#</th><th>Section</th><th>Claim</th><th>Maturity</th><th>Evidence</th><th>Kill Condition</th><th>Proof Boundary</th></tr></thead><tbody>{claim_rows}</tbody></table>
    <p class="footer">Showing first 40 claims. Full canonical claim audit remains in the XLSX/JSON.</p>
  </section>

  <p class="footer">
    Generated {esc(generated)} from {esc(source_grade)}. Snapshot: {esc(source_snapshot_text)}.
  </p>
</main>
</body>
</html>
"""


def write_review_xlsx(source_xlsx: Path, output_xlsx: Path, grade: dict[str, Any], snapshot: dict[str, Any] | None) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - environment guard
        raise SystemExit(f"openpyxl is required to write the styled workbook: {exc}") from exc

    try:
        if not source_xlsx.exists():
            raise FileNotFoundError(source_xlsx)
        wb = load_workbook(source_xlsx)
    except Exception:
        wb = Workbook()
        summary = wb.active
        summary.title = "Summary"
        metrics = grade.get("metrics") or {}
        append_safe(summary, ["Paper ID", grade.get("paper_id", "")])
        append_safe(summary, ["Source", grade.get("source_file", "")])
        append_safe(summary, ["Generated", grade.get("generated_at", "")])
        append_safe(summary, ["Words", metrics.get("word_count", 0)])
        append_safe(summary, ["Sections", metrics.get("section_count", 0)])
        append_safe(summary, ["Equations", metrics.get("equation_count", 0)])
        append_safe(summary, ["Claim Candidates", metrics.get("claim_candidate_count", 0)])
        append_safe(summary, ["Top Terms", metrics.get("top_terms", "")])

        claim_headers = [
            "paper_id", "section", "one_sentence_claim", "claim_maturity_level",
            "claim_maturity_label", "facts_snapshot", "Q1_identity", "Q2_scope",
            "Q3_mechanism", "Q4_evidence", "Q5_falsifiability", "Q6_boundary",
            "Q7_listener_risk", "forward_test", "reverse_test", "evidence_bar",
            "kill_conditions", "not_claimed", "proof_boundary", "nearby_equation",
        ]
        claims_ws = wb.create_sheet("Claim Audit")
        append_safe(claims_ws, claim_headers)
        for claim in grade.get("claims") or []:
            append_safe(claims_ws, [claim.get(h, "") for h in claim_headers])

        equations_ws = wb.create_sheet("Equations")
        append_safe(equations_ws, ["paper_id", "equation"])
        for eq in grade.get("equations") or []:
            append_safe(equations_ws, [grade.get("paper_id", ""), eq])

        sections_ws = wb.create_sheet("Sections")
        append_safe(sections_ws, ["paper_id", "section", "character_count", "preview"])
        for section in grade.get("sections") or []:
            text = str(section.get("text") or "")
            append_safe(sections_ws, [grade.get("paper_id", ""), section.get("title", ""), len(text), text[:500]])
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)

    metrics = grade.get("metrics") or {}
    claims = grade.get("claims") or []
    bits = snapshot_bits(snapshot)
    issues = issue_summary(claims)
    dist = claim_distribution(claims)

    title_fill = PatternFill("solid", fgColor="33250B")
    header_fill = PatternFill("solid", fgColor="B8871E")
    soft_fill = PatternFill("solid", fgColor="FFF7DC")
    line = Side(style="thin", color="D6BD73")
    border = Border(left=line, right=line, top=line, bottom=line)

    ws["A1"] = "Paper Grade Dashboard"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:H1")

    rows = [
        ("Paper ID", grade.get("paper_id", "")),
        ("Generated", grade.get("generated_at", "")),
        ("Words", metrics.get("word_count", 0)),
        ("Sections", metrics.get("section_count", 0)),
        ("Equations", metrics.get("equation_count", 0)),
        ("Claims", metrics.get("claim_candidate_count", 0)),
        ("Semantic Vector", bits.get("vector", "")),
        ("Semantic Hash", bits.get("hash", "")),
        ("Rigor Verdict", bits.get("rigor_verdict", "")),
        ("Math Layer", bits.get("math_status", "")),
        ("Filename Safe", bits.get("filename_safe", "")),
    ]
    start = 3
    for offset, (label, value) in enumerate(rows):
        row = start + offset
        ws.cell(row=row, column=1, value=excel_safe(label))
        ws.cell(row=row, column=2, value=excel_safe(value))
        ws.cell(row=row, column=1).font = Font(bold=True, color="33250B")
        ws.cell(row=row, column=1).fill = soft_fill
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    issue_start = 3
    ws["D2"] = "Repair Pressure"
    ws["D2"].font = Font(bold=True, color="33250B")
    for idx, (name, count) in enumerate(issues.most_common(10), issue_start):
        ws.cell(row=idx, column=4, value=excel_safe(name))
        ws.cell(row=idx, column=5, value=count)
    dist_start = 16
    ws.cell(row=dist_start, column=4, value="Maturity Distribution")
    ws.cell(row=dist_start, column=4).font = Font(bold=True, color="33250B")
    for offset, (label, count) in enumerate(dist.most_common(), 1):
        ws.cell(row=dist_start + offset, column=4, value=excel_safe(label))
        ws.cell(row=dist_start + offset, column=5, value=count)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.row == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = header_fill if sheet.title != "Dashboard" else title_fill
        for col_idx in range(1, min(sheet.max_column, 12) + 1):
            letter = get_column_letter(col_idx)
            width = 14
            if sheet.title == "Claim Audit" and col_idx in {3, 14, 15, 17, 18, 19}:
                width = 42
            elif sheet.title == "Equations" and col_idx == 2:
                width = 72
            elif sheet.title == "Sections" and col_idx == 4:
                width = 72
            elif sheet.title == "Dashboard" and col_idx in {2, 4}:
                width = 48
            sheet.column_dimensions[letter].width = width
        if sheet.title == "Claim Audit" and sheet.max_row > 1:
            level_col = 4
            rng = f"{get_column_letter(level_col)}2:{get_column_letter(level_col)}{sheet.max_row}"
            sheet.conditional_formatting.add(
                rng,
                ColorScaleRule(start_type="num", start_value=1, start_color="B3442F", mid_type="num", mid_value=4, mid_color="E5C766", end_type="num", end_value=7, end_color="267A4A"),
            )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> int:
    parser = argparse.ArgumentParser(description="Render a human review surface for deterministic paper-grade outputs.")
    parser.add_argument("paper_grade_json", type=Path)
    parser.add_argument("--snapshot", type=Path)
    parser.add_argument("--snapshot-root", type=Path, default=WORKFLOW_ROOT / "03_FINAL_READY")
    parser.add_argument("--semantic-manifest", type=Path)
    parser.add_argument("--xlsx", type=Path)
    parser.add_argument("--out-html", type=Path)
    parser.add_argument("--out-xlsx", type=Path)
    args = parser.parse_args()

    grade_path = args.paper_grade_json
    grade = load_json(grade_path)
    snapshot_path = args.snapshot or find_snapshot_for_grade(grade_path, args.snapshot_root)
    snapshot = load_json(snapshot_path) if snapshot_path and snapshot_path.exists() else None
    if snapshot is None:
        snapshot = semantic_row_as_snapshot(semantic_manifest_lookup(args.semantic_manifest, grade_path))

    out_html = args.out_html or grade_path.with_suffix("").with_suffix(".paper-grade-dashboard.html")
    out_html.parent.mkdir(parents=True, exist_ok=True)
    out_html.write_text(build_html(grade, snapshot, grade_path, snapshot_path), encoding="utf-8")

    source_xlsx = args.xlsx or grade_path.with_suffix("").with_suffix(".paper-grade.xlsx")
    out_xlsx = args.out_xlsx or grade_path.with_suffix("").with_suffix(".paper-grade.review.xlsx")
    review_error = None
    try:
        write_review_xlsx(source_xlsx, out_xlsx, grade, snapshot)
    except Exception as exc:
        # Keep the HTML dashboard useful even when an older minimal XLSX has
        # invalid worksheet XML from pre-sanitization grader output.
        review_error = repr(exc)

    print(json.dumps({
        "ok": True,
        "paper_grade_json": str(grade_path),
        "snapshot": str(snapshot_path) if snapshot_path else None,
        "dashboard_html": str(out_html),
        "review_xlsx": str(out_xlsx) if review_error is None else None,
        "review_xlsx_error": review_error,
        "claims": len(grade.get("claims") or []),
        "equations": len(grade.get("equations") or []),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
