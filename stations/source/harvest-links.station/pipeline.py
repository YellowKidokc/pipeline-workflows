"""
harvest-links workflow (post-Cannon refactor, 2026-05-02).

Reads URLs from one of:
  - Excel (.xlsx)    — input.path with input.url_column
  - CSV (.csv)       — same fields
  - Postgres         — input.postgres_query (must SELECT a url column;
                       optionally also id, treated as source_id)

For each URL:
  1. requests.get with browser-ish user agent
  2. BeautifulSoup extracts visible text (readability-style heuristic)
  3. SBERT embed via Infinity (NAS) -> Qdrant collection upsert
  4. DeBERTa classify against the apologetics labels (Postgres scalar)
  5. INSERT scalar fields into Postgres target table (no BYTEA vector)

Architecture note: vectors live in Qdrant under the collection named after
the target table. Postgres holds scalars only (label, confidence, status,
text, error, etc.). Cluster IDs come back from the cluster_runner pass.
"""
from __future__ import annotations

import csv
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

HERE = Path(__file__).resolve().parent
STATIONS = HERE.parent              # X:\Backside\stations
ROOT = HERE.parent.parent           # X:\Backside  (shared _LOGS lives here)
LOG_DIR = ROOT / "_LOGS"

# Post-reorg layout: each old numbered tool is now its own sibling station.
# Keep the old tool keys as stable references; resolve them to the new dirs.
TOOL_STATIONS = {
    "02_SBERT":    "sbert-embedder.station",
    "03_DEBERTA":  "deberta-runner.station",
    "04_HDBSCAN":  "hdbscan-cluster.station",
    "07_POSTGRES": "postgres-sync.station",
}


def _tool_dir(tool: str) -> Path:
    return STATIONS / TOOL_STATIONS[tool]


for _tool in TOOL_STATIONS:
    _p = _tool_dir(_tool)
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))


# Postgres holds scalars only — vectors live in Qdrant.
HARVEST_DDL = """
CREATE TABLE IF NOT EXISTS {table} (
    id SERIAL PRIMARY KEY,
    source_id INT,
    url TEXT UNIQUE,
    http_status INT,
    title TEXT,
    text TEXT,
    text_len INT,
    deberta_label TEXT,
    deberta_confidence FLOAT,
    cluster_id INT,
    fetched_at TIMESTAMP DEFAULT NOW(),
    error TEXT
);
CREATE INDEX IF NOT EXISTS idx_{table}_url ON {table}(url);
CREATE INDEX IF NOT EXISTS idx_{table}_cluster_null ON {table}(id) WHERE cluster_id IS NULL;
"""


def _setup_logging(name: str) -> logging.Logger:
    LOG_DIR.mkdir(exist_ok=True)
    logfile = LOG_DIR / f"workflow_{name}_{datetime.now():%Y%m%d}.log"
    logger = logging.getLogger(f"workflow.{name}")
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    fh = logging.FileHandler(logfile, encoding="utf-8")
    fh.setFormatter(fmt)
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(sh)
    return logger


def _ensure_imports(log: logging.Logger) -> bool:
    missing = []
    for mod, pkg in [("requests", "requests"), ("bs4", "beautifulsoup4"),
                     ("httpx", "httpx"), ("qdrant_client", "qdrant-client")]:
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    if missing:
        log.error(
            "Missing packages: %s\nInstall with:\n  %s -m pip install %s",
            ", ".join(missing),
            sys.executable,
            " ".join(missing),
        )
        return False
    return True


def _read_input_urls(cfg: dict, log: logging.Logger) -> list[dict]:
    inp = cfg["input"]
    typ = inp.get("type", "excel")
    if typ in ("excel", "csv"):
        path = Path(inp.get("path", ""))
        if not path or not path.exists():
            log.error("input.path does not exist: %s", path)
            return []
        url_col = inp.get("url_column", "url")
        if typ == "excel":
            try:
                from openpyxl import load_workbook  # type: ignore
            except ImportError:
                log.error("openpyxl not installed. pip install openpyxl")
                return []
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            ws = wb.active
            header = [c.value for c in next(ws.iter_rows(max_row=1))]
            if url_col not in header:
                log.error("column %r not found in %s. headers=%s", url_col, path, header)
                return []
            url_idx = header.index(url_col)
            id_idx = header.index("id") if "id" in header else None
            rows: list[dict] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[url_idx] is None:
                    continue
                rows.append({
                    "url": str(row[url_idx]),
                    "source_id": (None if id_idx is None else row[id_idx]),
                })
            return rows
        else:  # csv
            with open(path, "r", encoding="utf-8", newline="") as f:
                rdr = csv.DictReader(f)
                if url_col not in rdr.fieldnames:
                    log.error("column %r not found in CSV. headers=%s", url_col, rdr.fieldnames)
                    return []
                return [
                    {"url": r[url_col], "source_id": r.get("id")}
                    for r in rdr if r.get(url_col)
                ]
    if typ == "postgres":
        from db_utils import Database

        with Database(application_name="harvest_input") as db:
            rows = db.query(inp["postgres_query"])
        out: list[dict] = []
        for r in rows:
            url = r.get("url")
            if not url:
                continue
            out.append({"url": url, "source_id": r.get("id")})
        return out
    log.error("unknown input.type: %s", typ)
    return []


def _extract_text(html: str, strategy: str = "body") -> tuple[str, str]:
    """Return (title, text). Best-effort extraction."""
    from bs4 import BeautifulSoup  # type: ignore

    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    title = (soup.title.get_text(strip=True) if soup.title else "") or ""
    if strategy == "readability_then_body":
        try:
            from readability import Document  # type: ignore

            doc = Document(html)
            content_html = doc.summary()
            content_soup = BeautifulSoup(content_html, "html.parser")
            for t in content_soup(["script", "style", "noscript"]):
                t.decompose()
            text = content_soup.get_text(separator=" ", strip=True)
            if text:
                return title, text
        except Exception:
            pass
    body = soup.body or soup
    text = body.get_text(separator=" ", strip=True)
    return title, text


def _fetch_and_parse(url: str, ua: str, timeout: int, strategy: str, max_chars: int) -> dict:
    import requests  # type: ignore

    headers = {"User-Agent": ua}
    resp = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
    out = {"http_status": resp.status_code}
    if resp.status_code != 200:
        out["title"] = ""
        out["text"] = ""
        return out
    title, text = _extract_text(resp.text, strategy=strategy)
    if max_chars > 0:
        text = text[:max_chars]
    out["title"] = title[:500] if title else ""
    out["text"] = text
    return out


def main() -> int:
    cfg = json.loads((HERE / "config.json").read_text(encoding="utf-8"))
    log = _setup_logging(cfg.get("name", "harvest-links"))
    log.info("=== START harvest-links ===")

    if not _ensure_imports(log):
        return 1

    urls = _read_input_urls(cfg, log)
    log.info("input urls: %d", len(urls))
    if not urls:
        return 1

    fetch_cfg = cfg.get("fetch", {})
    ua = fetch_cfg.get("user_agent", "Mozilla/5.0 BrainHarvester/1.0")
    timeout = int(fetch_cfg.get("timeout_sec", 20))
    strategy = fetch_cfg.get("extract_strategy", "readability_then_body")
    max_chars = int(fetch_cfg.get("max_chars", 20000))
    skip_non_200 = bool(fetch_cfg.get("skip_if_status_not_200", True))

    target = cfg["target"]
    table = target["table"]
    collection = target.get("qdrant_collection") or table

    # SBERT runner now wraps Infinity HTTP. DeBERTa stays a local model.
    import sbert_runner
    import deberta_runner
    from db_utils import Database
    from qdrant_client.http.models import PointStruct

    sb_cfg = json.loads((_tool_dir("02_SBERT") / "config.json").read_text(encoding="utf-8"))
    db_cfg = json.loads((_tool_dir("03_DEBERTA") / "config.json").read_text(encoding="utf-8"))
    labels = db_cfg["labels"]

    em = sbert_runner.InfinityClient(
        base_url=sb_cfg["infinity_url"],
        model=sb_cfg.get("model_settings", {}).get("model_name", "sentence-transformers/all-MiniLM-L6-v2"),
        http_batch_size=int(sb_cfg.get("model_settings", {}).get("http_batch_size", 32)),
    )
    log.info("Infinity ready: model=%s dim=%d", em.model, em.dim)

    qd = sbert_runner._qdrant_client(sb_cfg["qdrant_url"])
    sbert_runner.ensure_collection(qd, collection, em.dim, sb_cfg.get("vector_distance", "cosine"))
    log.info("Qdrant collection ready: %s", collection)

    clf = deberta_runner.Classifier(
        model_name=db_cfg["model_settings"].get("model_name"),
        device=db_cfg["model_settings"].get("device", "auto"),
        cache_dir=db_cfg.get("model_cache_dir"),
        hypothesis_template=db_cfg["model_settings"].get("hypothesis_template", "This text is about {}."),
    )
    log.info("DeBERTa loaded device=%s labels=%d", clf.device, len(labels))

    summary_path = Path(cfg.get("summary_csv", str(LOG_DIR / "harvest_summary_YYYYMMDD.csv"))
                        .replace("YYYYMMDD", datetime.now().strftime("%Y%m%d")))
    summary_path.parent.mkdir(parents=True, exist_ok=True)

    inserted = 0
    failed = 0

    with Database(application_name="harvest_links") as db, \
         open(summary_path, "a", newline="", encoding="utf-8") as csv_f:
        if target.get("ensure_table", True):
            with db._cursor() as cur:
                cur.execute(HARVEST_DDL.format(table=table))
            log.info("ensured table %s", table)

        existing_urls = {r["url"] for r in db.query(f"SELECT url FROM {table}")}
        log.info("existing urls in target: %d", len(existing_urls))

        w = csv.DictWriter(csv_f, fieldnames=[
            "source_id", "url", "http_status", "title", "top_label", "top_score", "error",
        ])
        if csv_f.tell() == 0:
            w.writeheader()

        max_text_chars = int(db_cfg["model_settings"].get("max_text_chars", 2000))

        for i, item in enumerate(urls, 1):
            url = item["url"]
            sid = item.get("source_id")
            if url in existing_urls:
                continue

            row = {
                "source_id": sid, "url": url, "http_status": None,
                "title": "", "text": "",
                "vector": None, "label": "", "confidence": None,
                "error": None,
            }
            try:
                fetched = _fetch_and_parse(url, ua, timeout, strategy, max_chars)
                row.update(fetched)
                if skip_non_200 and row["http_status"] != 200:
                    row["error"] = f"non-200 status {row['http_status']}"
            except Exception as e:
                row["error"] = f"fetch_error: {e!r}"

            if not row["error"] and row["text"]:
                try:
                    row["vector"] = em.embed([row["text"]])[0]
                except Exception as e:
                    row["error"] = f"embed_error: {e!r}"

            if not row["error"] and row["text"]:
                try:
                    res = clf.classify(
                        row["text"][:max_text_chars] if max_text_chars > 0 else row["text"],
                        labels,
                    )
                    row["label"] = res["label"]
                    row["confidence"] = float(res["score"])
                except Exception as e:
                    row["error"] = (row["error"] or "") + f"; classify_error: {e!r}"

            try:
                rowcount = db.execute(
                    f"INSERT INTO {table} "
                    f"(source_id, url, http_status, title, text, text_len, "
                    f" deberta_label, deberta_confidence, error) "
                    f"VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (url) DO NOTHING "
                    f"RETURNING id",
                    (
                        row["source_id"], row["url"], row["http_status"],
                        row["title"], row["text"], len(row["text"] or ""),
                        row["label"], row["confidence"], row["error"],
                    ),
                )
                # We need the new row's id to use as Qdrant point id. RETURNING
                # only fires on the actual insert path; on conflict-do-nothing
                # the id will already exist — fetch it.
                inserted_id = None
                fetched = db.query(f"SELECT id FROM {table} WHERE url = %s", (row["url"],))
                if fetched:
                    inserted_id = fetched[0]["id"]

                if row["vector"] is not None and inserted_id is not None:
                    payload = {
                        "source_table": table,
                        "source_id": inserted_id,
                        "url": row["url"],
                        "title": row["title"][:300] if row["title"] else "",
                        "label": row["label"],
                        "confidence": row["confidence"],
                        "embedded_at": datetime.now().isoformat(timespec="seconds"),
                    }
                    qd.upsert(
                        collection_name=collection,
                        points=[PointStruct(
                            id=int(inserted_id),
                            vector=row["vector"].tolist(),
                            payload=payload,
                        )],
                        wait=True,
                    )

                if row["error"]:
                    failed += 1
                else:
                    inserted += 1
                existing_urls.add(url)
            except Exception as e:
                log.exception("DB/Qdrant write failed for %s: %s", url, e)
                failed += 1

            w.writerow({
                "source_id": row["source_id"],
                "url": row["url"],
                "http_status": row["http_status"],
                "title": row["title"][:120] if row["title"] else "",
                "top_label": row["label"],
                "top_score": row["confidence"],
                "error": row["error"],
            })
            csv_f.flush()

            if i % 25 == 0 or i == len(urls):
                log.info("[%d/%d] inserted=%d failed=%d", i, len(urls), inserted, failed)

    em.close()
    log.info("harvest done. inserted=%d failed=%d", inserted, failed)

    if cfg.get("cluster_after", True):
        log.info("--- cluster pass ---")
        try:
            import cluster_runner

            cl_cfg = json.loads((_tool_dir("04_HDBSCAN") / "config.json").read_text(encoding="utf-8"))
            cl_cfg.setdefault("source", {})["type"] = "qdrant"
            cl_cfg["source"]["collection"] = collection
            cl_cfg["source"]["postgres_table"] = table
            cluster_runner.run(cl_cfg)
        except Exception as e:
            log.warning("cluster pass skipped or failed: %s", e)
            log.warning("note: 04_HDBSCAN/cluster_runner.py expects vectors from Postgres BYTEA. "
                        "It needs the same Infinity+Qdrant refactor before cluster_after can succeed.")

    log.info("=== END harvest-links ===")
    return 0


if __name__ == "__main__":
    sys.exit(main())
