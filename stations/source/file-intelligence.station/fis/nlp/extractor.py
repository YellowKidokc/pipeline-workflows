"""Text extraction from various file formats."""

from pathlib import Path

from fis.log import get_logger

log = get_logger("extractor")


def extract_text(file_path: str) -> str:
    """Extract text content from a file based on its extension."""
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext in (".txt", ".md", ".markdown"):
        return _read_text(path)
    elif ext == ".pdf":
        return _read_pdf(path)
    elif ext in (".docx", ".doc"):
        return _read_docx(path)
    elif ext in (".xlsx", ".xls"):
        return _read_excel(path)
    elif ext in (".mp4", ".mp3", ".wav", ".m4a", ".webm", ".mkv", ".avi"):
        return _transcribe_audio(path)
    elif ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"):
        return _ocr_image(path)
    else:
        # Try as plain text
        try:
            return _read_text(path)
        except UnicodeDecodeError:
            return ""


def _read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="latin-1")


def _read_pdf(path: Path) -> str:
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(str(path))
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        return text
    except ImportError:
        log.warning("PyMuPDF not installed — pip install PyMuPDF")
        return ""
    except Exception as e:
        log.error("PDF extraction failed for %s: %s", path, e)
        return ""


def _read_docx(path: Path) -> str:
    try:
        from docx import Document

        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs)
    except ImportError:
        log.warning("python-docx not installed — pip install python-docx")
        return ""
    except Exception as e:
        log.error("DOCX extraction failed for %s: %s", path, e)
        return ""


def _read_excel(path: Path) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        text_parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    text_parts.append(" ".join(cells))
        wb.close()
        return "\n".join(text_parts)
    except Exception as e:
        log.error("Excel extraction failed for %s: %s", path, e)
        return ""


def _transcribe_audio(path: Path) -> str:
    try:
        from faster_whisper import WhisperModel

        from fis.db.connection import get_config

        config = get_config()
        model_size = config.get("whisper", "model_size", fallback="small")
        compute_type = config.get("whisper", "compute_type", fallback="int8")
        device = config.get("whisper", "device", fallback="cpu")

        model = WhisperModel(model_size, device=device, compute_type=compute_type)
        segments, _ = model.transcribe(str(path))
        return " ".join(seg.text for seg in segments)
    except ImportError:
        log.warning("faster-whisper not installed — pip install faster-whisper")
        return ""
    except Exception as e:
        log.error("Audio transcription failed for %s: %s", path, e)
        return ""


def _ocr_image(path: Path) -> str:
    # OCR is optional — requires tesseract
    try:
        import pytesseract
        from PIL import Image

        img = Image.open(str(path))
        return pytesseract.image_to_string(img)
    except ImportError:
        return ""
