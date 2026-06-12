"""Export kickout files to Excel for review."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from fis.db.connection import get_connection


def export_kickouts(output_path: str = "kickouts.xlsx"):
    """Export all pending and kickout files to an Excel review sheet."""
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT f.file_id, f.sequence_id, f.original_name, f.proposed_name,
                       f.domain, f.subject_codes, f.slug, f.confidence, f.status,
                       f.file_path, f.created_at,
                       array_agg(DISTINCT t.tag) FILTER (WHERE t.tag IS NOT NULL) AS tags
                FROM files f
                LEFT JOIN file_tags t ON f.file_id = t.file_id
                WHERE f.status IN ('pending', 'kickout')
                GROUP BY f.file_id
                ORDER BY f.confidence ASC, f.created_at DESC
            """)
            rows = cur.fetchall()
    finally:
        conn.close()

    from fis.log import get_logger
    log = get_logger("export")

    if not rows:
        log.info("No pending or kickout files to export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "FIS Review Queue"

    # Headers
    headers = [
        "File ID", "Seq ID", "Original Name", "Proposed Name",
        "Domain", "Subjects", "Slug", "Confidence", "Status",
        "Tags", "File Path", "Created",
        "YOUR Domain", "YOUR Subjects", "YOUR Slug", "APPROVED?"
    ]
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    correction_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Highlight correction columns
    for col in range(13, 17):
        ws.cell(row=1, column=col).fill = correction_fill

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["file_id"])
        ws.cell(row=row_idx, column=2, value=row["sequence_id"])
        ws.cell(row=row_idx, column=3, value=row["original_name"])
        ws.cell(row=row_idx, column=4, value=row["proposed_name"])
        ws.cell(row=row_idx, column=5, value=row["domain"])
        ws.cell(row=row_idx, column=6, value=", ".join(row["subject_codes"] or []))
        ws.cell(row=row_idx, column=7, value=row["slug"])
        ws.cell(row=row_idx, column=8, value=round(row["confidence"] or 0, 1))
        ws.cell(row=row_idx, column=9, value=row["status"])
        ws.cell(row=row_idx, column=10, value=", ".join(row["tags"] or []))
        ws.cell(row=row_idx, column=11, value=row["file_path"])
        ws.cell(row=row_idx, column=12, value=str(row["created_at"]))

        # Correction columns — green highlight
        for col in range(13, 17):
            ws.cell(row=row_idx, column=col).fill = correction_fill

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(output_path)
    log.info("Exported %d files to %s", len(rows), output_path)


def import_corrections(excel_path: str):
    """Import corrections from the review Excel back into Postgres."""
    from openpyxl import load_workbook

    from fis.db.models import insert_correction

    wb = load_workbook(excel_path)
    ws = wb.active

    corrections = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        file_id = row[0].value
        if not file_id:
            continue

        new_domain = row[12].value  # Column M
        new_subjects = row[13].value  # Column N
        new_slug = row[14].value  # Column O
        approved = row[15].value  # Column P

        if not approved:
            continue

        old = {
            "domain": row[4].value,
            "subjects": row[5].value.split(", ") if row[5].value else [],
            "slug": row[6].value,
        }
        new = {
            "domain": new_domain or old["domain"],
            "subjects": new_subjects.split(", ") if new_subjects else old["subjects"],
            "slug": new_slug or old["slug"],
        }

        insert_correction(file_id, old, new)
        corrections += 1

    from fis.log import get_logger
    get_logger("export").info("Imported %d corrections.", corrections)
    wb.close()


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == "import":
        import_corrections(sys.argv[2] if len(sys.argv) > 2 else "kickouts.xlsx")
    else:
        export_kickouts()
