#!/usr/bin/env python3
"""
Build a human review packet for every math expression found in a series.

This station does not translate math. It inventories math and prepares rows for
the hand-authored MTL Excel workbook.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover - optional comparison only
    openpyxl = None


BLOCK_TAGS = re.compile(
    r"</?(?:p|div|section|article|header|footer|main|br|hr|li|tr|td|th|table|h[1-6]|figure|figcaption)[^>]*>",
    re.I,
)
TAG_RE = re.compile(r"<[^>]+>")
SCRIPT_STYLE_RE = re.compile(r"<(script|style|nav|footer|header)\b.*?</\1>", re.I | re.S)
DETAILS_RE = re.compile(r"<details\b.*?</details>", re.I | re.S)

MATH_PATTERNS = [
    ("mathjax_display", re.compile(r"\\\[(.+?)\\\]", re.S)),
    ("mathjax_inline", re.compile(r"\\\((.+?)\\\)", re.S)),
    ("dollar_display", re.compile(r"\$\$(.+?)\$\$", re.S)),
    ("script_mathtex", re.compile(r"<script[^>]+type=[\"']math/tex[^\"']*[\"'][^>]*>(.+?)</script>", re.I | re.S)),
    ("data_tex", re.compile(r"\bdata-(?:tex|latex|math)=[\"']([^\"']+)[\"']", re.I | re.S)),
]
CALLOUT_RE = re.compile(
    r"<details\b(?=[^>]*\bdata-eq-id=[\"']([^\"']+)[\"'])([^>]*)>(.*?)</details>",
    re.I | re.S,
)
TITLE_RE = re.compile(r"<title[^>]*>(.*?)</title>", re.I | re.S)
H1_RE = re.compile(r"<h1[^>]*>(.*?)</h1>", re.I | re.S)


@dataclass
class MathItem:
    article_order: int
    article_title: str
    source_file: str
    source_relpath: str
    equation_order: int
    data_eq_id: str
    suggested_workbook_id: str
    extraction_method: str
    latex_or_math: str
    context_before: str
    context_after: str
    existing_callout_text: str
    workbook_match: str
    workbook_status: str
    needs_review: bool
    notes: str


def text_from_html(fragment: str) -> str:
    fragment = SCRIPT_STYLE_RE.sub(" ", fragment)
    fragment = DETAILS_RE.sub(" ", fragment)
    fragment = BLOCK_TAGS.sub("\n", fragment)
    fragment = TAG_RE.sub(" ", fragment)
    text = html.unescape(fragment)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n\s+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def compact(text: str, max_len: int = 700) -> str:
    text = re.sub(r"\s+", " ", html.unescape(text or "")).strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 1].rstrip() + "..."


def normalize_math(text: str) -> str:
    text = html.unescape(text or "")
    text = re.sub(r"\s+", "", text)
    return text.strip("$")


def article_title(raw: str, fallback: str) -> str:
    for pattern in (TITLE_RE, H1_RE):
        match = pattern.search(raw)
        if match:
            title = compact(text_from_html(match.group(1)), 180)
            if title:
                return title
    return fallback


def find_explicit_math(raw: str) -> list[dict[str, Any]]:
    matches: list[dict[str, Any]] = []
    for method, pattern in MATH_PATTERNS:
        for match in pattern.finditer(raw):
            value = html.unescape(match.group(1)).strip()
            if value:
                matches.append({
                    "method": method,
                    "start": match.start(),
                    "end": match.end(),
                    "math": compact(value, 2000),
                })
    matches.sort(key=lambda item: item["start"])
    return matches


def nearest_text_before(raw: str, pos: int) -> str:
    window = raw[max(0, pos - 2600):pos]
    text = text_from_html(window)
    chunks = [chunk.strip() for chunk in re.split(r"\n+", text) if chunk.strip()]
    if not chunks:
        return ""
    useful = chunks[-4:]
    return compact(" / ".join(useful), 700)


def nearest_text_after(raw: str, pos: int) -> str:
    window = raw[pos:pos + 1600]
    text = text_from_html(window)
    chunks = [chunk.strip() for chunk in re.split(r"\n+", text) if chunk.strip()]
    if not chunks:
        return ""
    return compact(" / ".join(chunks[:3]), 500)


def looks_like_math(text: str) -> bool:
    signals = ("\\", "=", "≥", "≤", "≠", "→", "∇", "Δ", "χ", "ψ", "Φ", "Σ", "R", "z =", "p <", " / ")
    return any(signal in text for signal in signals)


def choose_math_for_callout(raw: str, callout_start: int, explicit_matches: list[dict[str, Any]]) -> tuple[str, str]:
    previous = [item for item in explicit_matches if item["end"] <= callout_start]
    if previous and callout_start - previous[-1]["end"] < 2200:
        item = previous[-1]
        return item["math"], "data-eq-id_nearest_" + item["method"]

    context = nearest_text_before(raw, callout_start)
    if looks_like_math(context):
        return context, "data-eq-id_visible_context"
    return "", "data-eq-id_no_math_found"


def suggested_id(data_eq_id: str, source_stem: str, equation_order: int) -> str:
    if data_eq_id:
        return data_eq_id
    return f"{source_stem}-EQ-{equation_order:03d}"


def load_workbook_index(workbook_path: Path | None) -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    if not workbook_path or not workbook_path.exists() or openpyxl is None:
        return {}, {}
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    by_id: dict[str, dict[str, str]] = {}
    by_math: dict[str, str] = {}
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))}
        eq_id = str(row.get("id") or "").strip()
        latex = str(row.get("latex") or row.get("equation") or "").strip()
        if not eq_id:
            continue
        status = "needs_review" if str(row.get("needsReview") or "").lower() in {"true", "1", "yes"} else "complete"
        for col in ("easy", "audioSafe"):
            if not str(row.get(col) or "").strip():
                status = "needs_review"
        by_id[eq_id] = {"id": eq_id, "latex": latex, "status": status}
        if latex:
            by_math[normalize_math(latex)] = eq_id
    return by_id, by_math


def scan_article(path: Path, relpath: str, order: int, workbook_by_id: dict[str, dict[str, str]], workbook_by_math: dict[str, str]) -> list[MathItem]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    title = article_title(raw, path.stem)
    explicit = find_explicit_math(raw)
    items: list[MathItem] = []
    used_math_keys: set[str] = set()

    callouts = list(CALLOUT_RE.finditer(raw))
    for callout_index, match in enumerate(callouts, start=1):
        data_eq_id = html.unescape(match.group(1)).strip()
        math_text, method = choose_math_for_callout(raw, match.start(), explicit)
        if math_text:
            used_math_keys.add(normalize_math(math_text))
        existing = text_from_html(match.group(3))
        workbook_match = ""
        workbook_status = ""
        if data_eq_id in workbook_by_id:
            workbook_match = data_eq_id
            workbook_status = workbook_by_id[data_eq_id]["status"]
        elif math_text and normalize_math(math_text) in workbook_by_math:
            workbook_match = workbook_by_math[normalize_math(math_text)]
            workbook_status = workbook_by_id.get(workbook_match, {}).get("status", "")

        notes: list[str] = []
        if data_eq_id and not data_eq_id.startswith("eq_"):
            notes.append("Article-scoped id; align workbook id or alias before JSON lookup.")
        if not math_text:
            notes.append("No nearby explicit math found; review article context manually.")
        if not workbook_match:
            notes.append("Not matched to workbook by exact id or exact latex.")

        items.append(MathItem(
            article_order=order,
            article_title=title,
            source_file=str(path),
            source_relpath=relpath,
            equation_order=len(items) + 1,
            data_eq_id=data_eq_id,
            suggested_workbook_id=suggested_id(data_eq_id, path.stem, callout_index),
            extraction_method=method,
            latex_or_math=math_text,
            context_before=nearest_text_before(raw, match.start()),
            context_after=nearest_text_after(raw, match.end()),
            existing_callout_text=existing,
            workbook_match=workbook_match,
            workbook_status=workbook_status,
            needs_review=not bool(workbook_match and workbook_status == "complete"),
            notes=" ".join(notes),
        ))

    for explicit_index, item in enumerate(explicit, start=1):
        key = normalize_math(item["math"])
        if key in used_math_keys:
            continue
        workbook_match = workbook_by_math.get(key, "")
        workbook_status = workbook_by_id.get(workbook_match, {}).get("status", "") if workbook_match else ""
        items.append(MathItem(
            article_order=order,
            article_title=title,
            source_file=str(path),
            source_relpath=relpath,
            equation_order=len(items) + 1,
            data_eq_id="",
            suggested_workbook_id=suggested_id("", path.stem, len(items) + 1),
            extraction_method=item["method"],
            latex_or_math=item["math"],
            context_before=nearest_text_before(raw, item["start"]),
            context_after=nearest_text_after(raw, item["end"]),
            existing_callout_text="",
            workbook_match=workbook_match,
            workbook_status=workbook_status,
            needs_review=not bool(workbook_match and workbook_status == "complete"),
            notes="" if workbook_match else "No data-eq-id; add callout id or workbook alias before production wiring.",
        ))

    return items


def sorted_html_files(root: Path) -> list[Path]:
    files = [path for path in root.rglob("*.html") if "_backups" not in path.parts and "archive" not in path.parts]
    return sorted(files, key=lambda p: str(p).lower())


def write_csv(path: Path, rows: list[MathItem]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(asdict(rows[0]).keys()) if rows else ["article_order"])
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def write_excel_intake_csv(path: Path, rows: list[MathItem], series_code: str) -> None:
    fieldnames = ["id", "series", "sourceFile", "latex", "symbols", "easy", "standard", "academic", "audioSafe", "needsReview", "notes"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({
                "id": row.suggested_workbook_id,
                "series": series_code,
                "sourceFile": row.source_relpath,
                "latex": row.latex_or_math,
                "symbols": "",
                "easy": "",
                "standard": "",
                "academic": "",
                "audioSafe": "",
                "needsReview": "TRUE",
                "notes": (row.notes + " Fill easy + audioSafe only; standard/academic optional.").strip(),
            })


def write_packet_html(path: Path, rows: list[MathItem], series_root: Path, series_code: str) -> None:
    by_article: dict[tuple[int, str, str], list[MathItem]] = {}
    for row in rows:
        key = (row.article_order, row.article_title, row.source_relpath)
        by_article.setdefault(key, []).append(row)

    complete = sum(1 for row in rows if row.workbook_status == "complete")
    needs = sum(1 for row in rows if row.needs_review)
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    sections: list[str] = []
    for (order, title, relpath), article_rows in by_article.items():
        eq_blocks: list[str] = []
        for row in article_rows:
            status = row.workbook_status or "not in workbook"
            eq_blocks.append(f"""
      <article class="eq-card">
        <div class="eq-meta">Equation {row.equation_order:03d} | HTML id: {html.escape(row.data_eq_id or '(none)')} | Suggested workbook id: {html.escape(row.suggested_workbook_id)} | {html.escape(status)}</div>
        <pre>{html.escape(row.latex_or_math or '(no math captured)')}</pre>
        <dl>
          <dt>Before</dt><dd>{html.escape(row.context_before)}</dd>
          <dt>After</dt><dd>{html.escape(row.context_after)}</dd>
          <dt>Existing Callout</dt><dd>{html.escape(row.existing_callout_text or '(none)')}</dd>
          <dt>Notes</dt><dd>{html.escape(row.notes or '(none)')}</dd>
        </dl>
      </article>""")
        sections.append(f"""
    <section class="page">
      <h2>Page {order:03d}: {html.escape(title)}</h2>
      <p class="path">{html.escape(relpath)}</p>
      <p class="count">{len(article_rows)} math item(s)</p>
      {''.join(eq_blocks)}
    </section>""")

    document = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>{html.escape(series_code)} Math Review Packet</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 0; color: #111; background: #f7f7f3; }}
    header {{ padding: 32px 44px; background: #111; color: #f8f0d0; }}
    h1 {{ margin: 0 0 8px; font-size: 28px; }}
    .summary {{ display: flex; gap: 24px; flex-wrap: wrap; font-size: 14px; }}
    .page {{ page-break-before: always; padding: 32px 44px; }}
    .page:first-of-type {{ page-break-before: auto; }}
    h2 {{ margin: 0 0 4px; font-size: 22px; }}
    .path, .count, .eq-meta {{ color: #666; font-size: 12px; }}
    .eq-card {{ background: #fff; border: 1px solid #d7d0b8; border-left: 4px solid #b8922f; padding: 14px 16px; margin: 16px 0; border-radius: 4px; }}
    pre {{ white-space: pre-wrap; overflow-wrap: anywhere; background: #111; color: #f8f0d0; padding: 12px; border-radius: 4px; font-size: 13px; }}
    dl {{ display: grid; grid-template-columns: 110px 1fr; gap: 8px 12px; font-size: 13px; }}
    dt {{ font-weight: 700; color: #5c4212; }}
    dd {{ margin: 0; }}
    @media print {{ body {{ background: #fff; }} header {{ color: #111; background: #fff; border-bottom: 2px solid #111; }} .eq-card {{ break-inside: avoid; }} }}
  </style>
</head>
<body>
  <header>
    <h1>{html.escape(series_code)} Math Review Packet</h1>
    <div class="summary">
      <div>Generated: {html.escape(generated)}</div>
      <div>Series root: {html.escape(str(series_root))}</div>
      <div>Total math items: {len(rows)}</div>
      <div>Workbook complete matches: {complete}</div>
      <div>Needs review: {needs}</div>
    </div>
  </header>
  {''.join(sections)}
</body>
</html>
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(document, encoding="utf-8")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--series-root", required=True, help="Root folder containing series HTML files")
    parser.add_argument("--out", required=True, help="Output folder or parent export folder")
    parser.add_argument("--series-code", default="MDA", help="Series code for Excel intake rows")
    parser.add_argument("--workbook", default="", help="Optional MTL workbook for exact id/latex status comparison")
    args = parser.parse_args(argv)

    series_root = Path(args.series_root)
    if not series_root.exists():
        raise SystemExit(f"Series root not found: {series_root}")

    workbook_path = Path(args.workbook) if args.workbook else None
    workbook_by_id, workbook_by_math = load_workbook_index(workbook_path)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_root = Path(args.out)
    out_dir = out_root / f"{stamp}_{args.series_code.lower()}_math_review_packet"
    out_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[MathItem] = []
    files = sorted_html_files(series_root)
    for order, path in enumerate(files, start=1):
        relpath = str(path.relative_to(series_root))
        all_rows.extend(scan_article(path, relpath, order, workbook_by_id, workbook_by_math))

    write_packet_html(out_dir / "series_math_review_packet.html", all_rows, series_root, args.series_code)
    write_csv(out_dir / "series_math_inventory.csv", all_rows)
    write_excel_intake_csv(out_dir / "excel_intake_rows.csv", all_rows, args.series_code)
    manifest = {
        "seriesRoot": str(series_root),
        "seriesCode": args.series_code,
        "workbook": str(workbook_path) if workbook_path else "",
        "articleCount": len(files),
        "mathItemCount": len(all_rows),
        "needsReviewCount": sum(1 for row in all_rows if row.needs_review),
        "outputs": {
            "packetHtml": str(out_dir / "series_math_review_packet.html"),
            "inventoryCsv": str(out_dir / "series_math_inventory.csv"),
            "excelIntakeCsv": str(out_dir / "excel_intake_rows.csv"),
        },
    }
    (out_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps(manifest, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
