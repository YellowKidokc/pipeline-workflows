from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def normalize_latex(value: str) -> str:
    text = (value or "").strip()
    text = re.sub(r"^\$+", "", text)
    text = re.sub(r"\$+$", "", text)
    text = re.sub(r"^\\\[", "", text)
    text = re.sub(r"\\\]$", "", text)
    text = re.sub(r"^\\\(", "", text)
    text = re.sub(r"\\\)$", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    while "\\\\" in text:
        text = text.replace("\\\\", "\\")
    text = text.replace("\\displaystyle", "")
    text = text.replace("\\left", "").replace("\\right", "")
    text = re.sub(r"\\(?:,|;|!|:)", "", text)
    text = re.sub(r"\\mathrm\{([^{}]+)\}", r"\\text{\1}", text)
    text = text.replace("·", r"\cdot").replace("×", r"\times")
    text = text.replace("χ", r"\chi").replace("δ", r"\delta").replace("β", r"\beta")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def regex_for_equation(equation: str) -> str:
    normalized = normalize_latex(equation)
    return "^" + re.escape(normalized) + "$"


def slug(value: str, fallback: str) -> str:
    out = re.sub(r"[^a-zA-Z0-9]+", "-", value).strip("-").lower()
    return out[:70] or fallback


def convert(xlsx_path: Path, sheet_name: str | None = None) -> dict:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    index = {header.lower(): idx for idx, header in enumerate(headers)}
    required = ["equation", "short", "medium", "audio"]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Found: {headers}")

    equations = []
    seen: set[str] = set()
    for ordinal, row in enumerate(rows, start=2):
        raw = row[index["equation"]] if index["equation"] < len(row) else None
        equation = normalize_latex(str(raw or ""))
        if not equation or equation in seen:
            continue
        seen.add(equation)
        short = str(row[index["short"]] or "").strip()
        medium = str(row[index["medium"]] or "").strip()
        audio = str(row[index["audio"]] or "").strip()
        plain_english = ""
        structural_parts = []
        audio_safe = ""
        if "plain_english" in index and index["plain_english"] < len(row):
            plain_english = str(row[index["plain_english"]] or "").strip()
        if "audio_safe" in index and index["audio_safe"] < len(row):
            audio_safe = str(row[index["audio_safe"]] or "").strip()
        if "structural_parts_json" in index and index["structural_parts_json"] < len(row):
            raw_parts = str(row[index["structural_parts_json"]] or "").strip()
            if raw_parts:
                try:
                    structural_parts = json.loads(raw_parts)
                except json.JSONDecodeError:
                    structural_parts = []
        equation_id = f"xlsx-{ordinal:04d}-{slug(equation, 'equation')}"
        entry = {
            "equationId": equation_id,
            "title": short or f"Equation {ordinal}",
            "patterns": [regex_for_equation(equation)],
            "narrative": audio or medium or short,
            "summary": medium or short or audio,
            "source": {
                "type": "xlsx",
                "file": str(xlsx_path),
                "sheet": ws.title,
                "row": ordinal,
            },
            "rawLatex": equation,
        }
        if plain_english:
            entry["plainEnglish"] = plain_english
        if structural_parts:
            entry["structuralParts"] = structural_parts
        if audio_safe:
            entry["audioSafe"] = audio_safe
        equations.append(entry)

    return {
        "schema": "theophysics.math_dictionary_supplement.v1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_xlsx": str(xlsx_path),
        "sheet": ws.title,
        "equations": equations,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert math translation workbook to JSON dictionary supplement.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()
    data = convert(args.xlsx, args.sheet)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"out": str(args.out), "equations": len(data["equations"])}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
