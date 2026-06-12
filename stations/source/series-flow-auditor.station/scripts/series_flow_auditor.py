from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import dataclass, asdict
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any


FUNCTION_ORDER = [
    "FRAMING",
    "DEFINITION",
    "DOMAIN_OVERVIEW",
    "NARRATIVE_CASE",
    "DATA_EVIDENCE",
    "METHOD",
    "STATISTICAL_SYNTHESIS",
    "CONTROL_CASE",
    "OBJECTION_RESPONSE",
    "PREDICTION",
    "RECOVERY_PATH",
    "APPENDIX",
    "ARCHIVE",
]

CRITICAL_CONCEPTS = [
    "moral decline",
    "coherence",
    "domain",
    "evidence layer",
    "model",
    "prediction",
    "exception",
    "statistical synthesis",
    "nine domains",
    "measurable decline",
    "control case",
    "collapse",
    "theology",
    "physics analogy",
]

SEVERITY_TERMS = [
    "proof",
    "proves",
    "prediction",
    "predicts",
    "collapse",
    "certain",
    "inevitable",
    "civilizational",
    "theology",
    "physics",
    "measurable",
]

FUNCTION_KEYWORDS = {
    "FRAMING": ["introduction", "series", "why this matters", "problem", "question", "stakes", "overview"],
    "DEFINITION": ["definition", "define", "means", "what counts as", "what is", "terminology"],
    "DOMAIN_OVERVIEW": ["domain", "nine domains", "categories", "area", "dimension", "institution"],
    "NARRATIVE_CASE": ["story", "case", "example", "family", "city", "person", "scene", "witness"],
    "DATA_EVIDENCE": ["data", "evidence", "study", "survey", "rate", "trend", "chart", "measured"],
    "METHOD": ["method", "methodology", "score", "metric", "model", "how we measure", "protocol"],
    "STATISTICAL_SYNTHESIS": ["statistical", "synthesis", "correlation", "aggregate", "regression", "meta"],
    "CONTROL_CASE": ["control", "exception", "counterexample", "baseline", "comparison case"],
    "OBJECTION_RESPONSE": ["objection", "response", "criticism", "counterargument", "skeptic", "does not prove"],
    "PREDICTION": ["prediction", "predicts", "forecast", "if this continues", "will happen", "testable"],
    "RECOVERY_PATH": ["recovery", "restore", "repair", "solution", "path forward", "repentance", "rebuild"],
    "APPENDIX": ["appendix", "notes", "technical note", "reference", "supplement"],
    "ARCHIVE": ["archive", "legacy", "old version", "deprecated"],
}

INTRO_PATTERNS = [
    r"\bwhat (?:is|are|counts as)\b",
    r"\bdefine[sd]?\b",
    r"\bmeans\b",
    r"\bwe call\b",
    r"\bthis model\b",
    r"\bthis method\b",
]


@dataclass
class Article:
    ordinal: int
    title: str
    path: str
    text: str
    summary: str = ""
    article_number: str = ""


@dataclass
class ArticleAudit:
    ordinal: int
    title: str
    path: str
    function: str
    expected_function: str
    confidence: float
    function_scores: dict[str, int]
    required_prior_concepts: list[str]
    introduced_concepts: list[str]
    missing_prior_concepts: list[str]
    severity_flags: list[str]
    depends_on: list[int]
    enables: list[int]
    suggested_placement: str
    editorial_action: str
    placement_verdict: str
    notes: str


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig", errors="replace")


def clean_html(raw: str) -> tuple[str, str]:
    title = ""
    title_match = re.search(r"<title[^>]*>(.*?)</title>", raw, flags=re.I | re.S)
    if title_match:
        title = unescape(strip_tags(title_match.group(1))).strip()
    raw = re.sub(r"<(script|style)[^>]*>.*?</\1>", " ", raw, flags=re.I | re.S)
    raw = re.sub(r"<!--.*?-->", " ", raw, flags=re.S)
    text = strip_tags(raw)
    return title, normalize_space(unescape(text))


def strip_tags(raw: str) -> str:
    return re.sub(r"<[^>]+>", " ", raw)


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def title_from_markdown(text: str, fallback: str) -> str:
    for line in text.splitlines()[:40]:
        if line.lstrip().startswith("#"):
            return line.lstrip("# ").strip() or fallback
    return fallback


def natural_key(path: Path) -> tuple[int, str]:
    match = re.search(r"(?:mda|gtq|article)?[-_\s]*(\d{1,3})", path.stem, flags=re.I)
    number = int(match.group(1)) if match else 9999
    return number, path.name.lower()


def load_articles_from_dir(root: Path) -> list[Article]:
    paths = [
        path
        for path in root.iterdir()
        if path.is_file() and path.suffix.lower() in {".html", ".htm", ".md", ".txt", ".json"}
    ]
    paths.sort(key=natural_key)
    return [load_article(path, index + 1) for index, path in enumerate(paths)]


def load_articles_from_manifest(manifest_path: Path) -> list[Article]:
    payload = json.loads(read_text(manifest_path))
    articles = []
    base = manifest_path.parent
    for index, item in enumerate(payload, start=1):
        path = Path(item["path"])
        if not path.is_absolute():
            path = base / path
        article_number = item["article_number"] if "article_number" in item and item["article_number"] is not None else index
        article = load_article(path, int(article_number))
        article.title = item.get("title") or article.title
        article.article_number = str(article_number)
        articles.append(article)
    articles.sort(key=lambda item: item.ordinal)
    return articles


def load_article(path: Path, ordinal: int) -> Article:
    raw = read_text(path)
    suffix = path.suffix.lower()
    title = path.stem.replace("-", " ").replace("_", " ").strip()
    summary = ""
    text = raw
    if suffix in {".html", ".htm"}:
        html_title, text = clean_html(raw)
        title = html_title or title
    elif suffix == ".md":
        title = title_from_markdown(raw, title)
        text = normalize_space(re.sub(r"^---.*?---", " ", raw, flags=re.S))
    elif suffix == ".json":
        data = json.loads(raw)
        title = data.get("title") or title
        parts: list[str] = []
        for key in ("subtitle", "summary", "body", "text"):
            value = data.get(key)
            if isinstance(value, str):
                parts.append(value)
        for key in ("claims", "evidence", "domain_tags", "internal_links"):
            value = data.get(key)
            if value:
                parts.append(json.dumps(value, ensure_ascii=False))
        summary = data.get("summary", "") if isinstance(data.get("summary"), str) else ""
        text = normalize_space(" ".join(parts) or raw)
    else:
        text = normalize_space(raw)
    return Article(ordinal=ordinal, title=title, path=str(path), text=text, summary=summary, article_number=str(ordinal))


def count_keywords(text: str, keywords: list[str]) -> int:
    low = text.lower()
    return sum(low.count(keyword.lower()) for keyword in keywords)


def classify_function(article: Article) -> tuple[str, dict[str, int]]:
    low = f"{article.title} {article.summary} {article.text[:6000]}".lower()
    scores = {label: count_keywords(low, words) for label, words in FUNCTION_KEYWORDS.items()}
    if re.search(r"\bmda[-_\s]*0*0\b|\bseries map\b|\bstart here\b", low):
        scores["FRAMING"] += 4
    if re.search(r"\b\d+(\.\d+)?\s*%|\b\d{4}\b|\bsigma\b|\bcorrelation\b", low):
        scores["DATA_EVIDENCE"] += 2
    if re.search(r"\bdoes not prove\b|\blimitation\b|\bobjection\b", low):
        scores["OBJECTION_RESPONSE"] += 3
    best = max(FUNCTION_ORDER, key=lambda label: (scores.get(label, 0), -FUNCTION_ORDER.index(label)))
    if scores[best] == 0:
        best = "NARRATIVE_CASE"
    return best, scores


def classification_confidence(scores: dict[str, int], label: str) -> float:
    total = sum(max(0, value) for value in scores.values())
    if total <= 0:
        return 0.35
    best = max(0, scores.get(label, 0))
    confidence = 0.35 + min(0.6, best / max(total, 1))
    return round(confidence, 2)


def concepts_used(article: Article) -> list[str]:
    low = f"{article.title} {article.text}".lower()
    return [concept for concept in CRITICAL_CONCEPTS if concept in low]


def concepts_introduced(article: Article, used: list[str], function: str) -> list[str]:
    low = f"{article.title} {article.text[:2500]}".lower()
    intro_signal = function in {"FRAMING", "DEFINITION", "DOMAIN_OVERVIEW", "METHOD"} or any(
        re.search(pattern, low) for pattern in INTRO_PATTERNS
    )
    if not intro_signal:
        return []
    introduced = []
    for concept in used:
        idx = low.find(concept)
        window = low[max(0, idx - 120) : idx + len(concept) + 220] if idx >= 0 else low[:400]
        if function in {"DEFINITION", "DOMAIN_OVERVIEW", "METHOD"} or any(re.search(pattern, window) for pattern in INTRO_PATTERNS):
            introduced.append(concept)
    return introduced


def severity_flags(article: Article) -> list[str]:
    low = f"{article.title} {article.text}".lower()
    flags = []
    for term in SEVERITY_TERMS:
        if term in low:
            flags.append(term)
    return sorted(set(flags))


def audit_articles(articles: list[Article]) -> list[ArticleAudit]:
    introduced_so_far: set[str] = set()
    audits: list[ArticleAudit] = []
    for article in articles:
        function, scores = classify_function(article)
        confidence = classification_confidence(scores, function)
        used = concepts_used(article)
        introduced = concepts_introduced(article, used, function)
        missing = [concept for concept in used if concept not in introduced_so_far and concept not in introduced]
        severity = severity_flags(article)
        verdict = placement_verdict(article, function, missing, severity)
        notes = placement_notes(function, missing, severity)
        expected = expected_function_for_position(article.ordinal, len(articles))
        action = editorial_action(function, verdict, missing, severity)
        audits.append(
            ArticleAudit(
                ordinal=article.ordinal,
                title=article.title,
                path=article.path,
                function=function,
                expected_function=expected,
                confidence=confidence,
                function_scores=scores,
                required_prior_concepts=used,
                introduced_concepts=introduced,
                missing_prior_concepts=missing,
                severity_flags=severity,
                depends_on=[],
                enables=[],
                suggested_placement="",
                editorial_action=action,
                placement_verdict=verdict,
                notes=notes,
            )
        )
        introduced_so_far.update(introduced)
    infer_dependencies(audits)
    for audit in audits:
        audit.suggested_placement = suggested_position(audit)
    return audits


def expected_function_for_position(ordinal: int, total: int) -> str:
    if ordinal <= 1:
        return "FRAMING"
    if ordinal <= 3:
        return "DEFINITION / DOMAIN_OVERVIEW"
    if total and ordinal >= max(total - 1, 1):
        return "OBJECTION_RESPONSE / PREDICTION / RECOVERY_PATH / APPENDIX"
    middle = ordinal / max(total, 1)
    if middle < 0.45:
        return "NARRATIVE_CASE / DATA_EVIDENCE"
    if middle < 0.7:
        return "METHOD / STATISTICAL_SYNTHESIS / CONTROL_CASE"
    return "OBJECTION_RESPONSE / PREDICTION / RECOVERY_PATH"


def editorial_action(function: str, verdict: str, missing: list[str], severity: list[str]) -> str:
    if function in {"APPENDIX", "ARCHIVE"}:
        return "Appendix"
    if "too early" in verdict or missing:
        return "Move"
    if severity and function in {"FRAMING", "DEFINITION"}:
        return "Move"
    if function in {"FRAMING", "DEFINITION", "DOMAIN_OVERVIEW"}:
        return "Keep"
    return "Keep"


def placement_verdict(article: Article, function: str, missing: list[str], severity: list[str]) -> str:
    if article.ordinal == 1 and function not in {"FRAMING", "DEFINITION", "DOMAIN_OVERVIEW"}:
        return "too early / weak entry"
    if missing and severity:
        return "too early or missing bridge"
    if missing:
        return "needs prior concept bridge"
    if severity and article.ordinal <= 3 and function not in {"FRAMING", "DEFINITION", "METHOD"}:
        return "escalates early"
    return "right place or low-risk"


def placement_notes(function: str, missing: list[str], severity: list[str]) -> str:
    parts = [f"Classified as {function}."]
    if missing:
        parts.append("Uses before introduction: " + ", ".join(missing) + ".")
    if severity:
        parts.append("Severity language present: " + ", ".join(severity) + ".")
    return " ".join(parts)


def infer_dependencies(audits: list[ArticleAudit]) -> None:
    intro_by_concept: dict[str, int] = {}
    for audit in audits:
        for concept in audit.introduced_concepts:
            intro_by_concept.setdefault(concept, audit.ordinal)
    for audit in audits:
        deps = sorted({intro_by_concept[c] for c in audit.required_prior_concepts if c in intro_by_concept and intro_by_concept[c] < audit.ordinal})
        audit.depends_on = deps
    enables: dict[int, list[int]] = {audit.ordinal: [] for audit in audits}
    for audit in audits:
        for dep in audit.depends_on:
            enables.setdefault(dep, []).append(audit.ordinal)
    for audit in audits:
        audit.enables = sorted(enables.get(audit.ordinal, []))


def score_audit(audits: list[ArticleAudit]) -> dict[str, int]:
    if not audits:
        return {}
    first_ok = audits[0].function in {"FRAMING", "DEFINITION", "DOMAIN_OVERVIEW"}
    missing_total = sum(len(a.missing_prior_concepts) for a in audits)
    severe_early = sum(1 for a in audits[:4] if a.severity_flags and a.placement_verdict != "right place or low-risk")
    order_jumps = count_order_jumps(audits)
    repeated = repeated_function_penalty(audits)
    method_seen_before_synthesis = method_precedes_synthesis(audits)
    scores = {
        "entry_clarity": clamp(85 if first_ok else 45),
        "definition_order": clamp(100 - missing_total * 8),
        "claim_progression": clamp(90 - severe_early * 15 - order_jumps * 4),
        "evidence_progression": clamp(85 - order_jumps * 5),
        "reader_burden": clamp(100 - missing_total * 10),
        "narrative_flow": clamp(90 - order_jumps * 5 - repeated * 4),
        "argument_dependency": clamp(95 - missing_total * 8),
        "redundancy_control": clamp(90 - repeated * 8),
        "method_visibility": clamp(85 if method_seen_before_synthesis else 55),
        "self_contained_coherence": clamp(92 - missing_total * 7 - severe_early * 10),
    }
    scores["overall_sequence_score"] = round(sum(scores.values()) / len(scores))
    return scores


def clamp(value: int) -> int:
    return max(0, min(100, int(value)))


def count_order_jumps(audits: list[ArticleAudit]) -> int:
    positions = [FUNCTION_ORDER.index(a.function) for a in audits if a.function in FUNCTION_ORDER]
    return sum(1 for left, right in zip(positions, positions[1:]) if right + 2 < left)


def repeated_function_penalty(audits: list[ArticleAudit]) -> int:
    penalty = 0
    run_label = None
    run_count = 0
    for audit in audits:
        if audit.function == run_label:
            run_count += 1
        else:
            penalty += max(0, run_count - 2)
            run_label = audit.function
            run_count = 1
    penalty += max(0, run_count - 2)
    return penalty


def method_precedes_synthesis(audits: list[ArticleAudit]) -> bool:
    method_idx = next((i for i, a in enumerate(audits) if a.function == "METHOD"), None)
    synth_idx = next((i for i, a in enumerate(audits) if a.function == "STATISTICAL_SYNTHESIS"), None)
    if synth_idx is None:
        return method_idx is not None
    return method_idx is not None and method_idx < synth_idx


def recommended_order(audits: list[ArticleAudit]) -> list[ArticleAudit]:
    return sorted(audits, key=lambda a: (FUNCTION_ORDER.index(a.function), a.ordinal))


def missing_bridges(audits: list[ArticleAudit]) -> list[str]:
    concepts = Counter(concept for audit in audits for concept in audit.missing_prior_concepts)
    bridges = []
    for concept, _count in concepts.most_common():
        title = {
            "moral decline": "What Counts as Moral Decline?",
            "coherence": "What Coherence Means in This Series",
            "domain": "The Domains Being Measured",
            "evidence layer": "What Counts as Evidence Here?",
            "model": "How the Model Works",
            "prediction": "What This Series Can and Cannot Predict",
            "exception": "Exceptions, Controls, and Counterexamples",
            "statistical synthesis": "How the Statistical Synthesis Works",
            "nine domains": "The Nine Domains",
            "measurable decline": "What Counts as Measurable Decline?",
            "control case": "Control Cases and False Positives",
            "collapse": "Collapse Language and Its Limits",
            "theology": "Where Theology Enters and Where It Does Not",
            "physics analogy": "Physics Analogies: Structural Use and Limits",
        }.get(concept, f"Bridge for {concept}")
        bridges.append(title)
    return bridges


def verdict(scores: dict[str, int]) -> str:
    overall = scores.get("overall_sequence_score", 0)
    if overall >= 85:
        return "Strong order"
    if overall >= 70:
        return "Mostly strong with fixes"
    if overall >= 50:
        return "Structurally unstable"
    return "Needs full reordering"


def render_markdown(audits: list[ArticleAudit], scores: dict[str, int]) -> str:
    rec = recommended_order(audits)
    bridges = missing_bridges(audits)
    lines = [
        "# Series Flow Audit",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "## A. Executive Verdict",
        "",
        f"Verdict: **{verdict(scores)}**",
        "",
        f"Overall sequence score: **{scores.get('overall_sequence_score', 0)}/100**",
        "",
        "## B. Current Sequence Map",
        "",
        "| # | Title | Current Function | Required Prior Concepts | Placement Verdict | Notes |",
        "|---:|---|---|---|---|---|",
    ]
    for audit in audits:
        lines.append(
            f"| {audit.ordinal} | {escape_md(audit.title)} | `{audit.function}` | {escape_md(', '.join(audit.required_prior_concepts) or '-')} | {escape_md(audit.placement_verdict)} | {escape_md(audit.notes)} |"
        )
    lines.extend(
        [
            "",
            "## B2. Editorial Action Table",
            "",
            "| Current Order # | Title | Detected Function | Expected Function | Concepts Used Before Defined | Severity Flags | Dependencies | Suggested Placement | Confidence | Keep / Move / Merge / Appendix |",
            "|---:|---|---|---|---|---|---|---|---:|---|",
        ]
    )
    for audit in audits:
        lines.append(
            f"| {audit.ordinal} | {escape_md(audit.title)} | `{audit.function}` | {escape_md(audit.expected_function)} | {escape_md(', '.join(audit.missing_prior_concepts) or '-')} | {escape_md(', '.join(audit.severity_flags) or '-')} | {escape_md(', '.join(str(d) for d in audit.depends_on) or '-')} | {escape_md(audit.suggested_placement)} | {audit.confidence:.2f} | {audit.editorial_action} |"
        )
    lines.extend(["", "## C. Dependency Graph", ""])
    for audit in audits:
        deps = ", ".join(f"Article {d}" for d in audit.depends_on) or "none detected"
        enables = ", ".join(f"Article {d}" for d in audit.enables) or "none detected"
        lines.extend([f"### Article {audit.ordinal}: {audit.title}", "", f"Depends on: {deps}", "", f"Enables: {enables}", ""])
    lines.extend(["## D. Misordered Articles", "", "| Article | Problem | Why It Is Misordered | Better Placement |", "|---|---|---|---|"])
    for audit in audits:
        if audit.placement_verdict != "right place or low-risk":
            better = suggested_position(audit)
            lines.append(f"| {escape_md(audit.title)} | {escape_md(audit.placement_verdict)} | {escape_md(audit.notes)} | {escape_md(better)} |")
    lines.extend(["", "## E. Missing Bridge Articles or Sections", ""])
    if bridges:
        lines.extend(f"- {bridge}" for bridge in bridges)
    else:
        lines.append("- No high-priority bridge article detected by deterministic pass.")
    lines.extend(["", "## F. Recommended New Order", ""])
    for index, audit in enumerate(rec, start=1):
        lines.append(f"{index}. {audit.title} - {audit.function}: {reason_for_position(audit)}")
    lines.extend(["", "## G. Merge / Appendix / Cut List", "", "### Merge", ""])
    lines.extend(merge_recommendations(audits) or ["- None detected by deterministic pass."])
    lines.extend(["", "### Move to appendix", ""])
    lines.extend(appendix_recommendations(audits) or ["- None detected by deterministic pass."])
    lines.extend(["", "### Cut or hold", "", "- None detected by deterministic pass."])
    lines.extend(["", "## H. Reader Confusion Points", ""])
    confusion = reader_confusion_points(audits)
    lines.extend(f"{i}. {item}" for i, item in enumerate(confusion[:10], start=1))
    if not confusion:
        lines.append("1. No top confusion point detected by deterministic pass.")
    lines.extend(["", "## I. Strongest Possible Story Arc", ""])
    lines.extend(f"{i}. {step}" for i, step in enumerate(story_arc(rec), start=1))
    lines.extend(["", "## J. Final Structural Rule", "", "Never let a severe claim appear before the series has introduced the definition, domain, evidence layer, and method needed to bear its weight.", ""])
    lines.extend(["## Scoring", "", "| Metric | Score |", "|---|---:|"])
    for key, value in scores.items():
        lines.append(f"| {key.replace('_', ' ').title()} | {value} |")
    return "\n".join(lines)


def escape_md(text: str) -> str:
    return str(text).replace("|", "\\|").replace("\n", " ")


def suggested_position(audit: ArticleAudit) -> str:
    position = FUNCTION_ORDER.index(audit.function) + 1
    if audit.missing_prior_concepts:
        return "After bridge article(s): " + ", ".join(missing_bridges([audit]))
    return f"Near other {audit.function} articles, around skeleton position {position}."


def reason_for_position(audit: ArticleAudit) -> str:
    if audit.function in {"FRAMING", "DEFINITION", "DOMAIN_OVERVIEW", "METHOD"}:
        return "establishes reader prerequisites"
    if audit.function in {"DATA_EVIDENCE", "STATISTICAL_SYNTHESIS", "CONTROL_CASE"}:
        return "loads evidence after the frame is established"
    if audit.function in {"OBJECTION_RESPONSE", "PREDICTION", "RECOVERY_PATH"}:
        return "belongs after the core case can carry it"
    return "fits this stage of the default series skeleton"


def merge_recommendations(audits: list[ArticleAudit]) -> list[str]:
    items = []
    for left, right in zip(audits, audits[1:]):
        if left.function == right.function and left.function in {"FRAMING", "DEFINITION", "DOMAIN_OVERVIEW"}:
            items.append(f"- Consider merging Article {left.ordinal} and Article {right.ordinal} if they repeat the same {left.function} job.")
    return items


def appendix_recommendations(audits: list[ArticleAudit]) -> list[str]:
    return [f"- Article {a.ordinal}: {a.title}" for a in audits if a.function in {"APPENDIX", "ARCHIVE"}]


def reader_confusion_points(audits: list[ArticleAudit]) -> list[str]:
    points = []
    for audit in audits:
        if audit.missing_prior_concepts:
            points.append(f"Article {audit.ordinal} uses {', '.join(audit.missing_prior_concepts)} before the reader has a detected introduction.")
        if audit.severity_flags and audit.ordinal <= 3:
            points.append(f"Article {audit.ordinal} uses severity language early: {', '.join(audit.severity_flags)}.")
    return points


def story_arc(audits: list[ArticleAudit]) -> list[str]:
    seen = []
    for label in FUNCTION_ORDER:
        matches = [a for a in audits if a.function == label]
        if matches:
            seen.append(f"{label}: {matches[0].title}")
    return seen[:15]


def write_csv(path: Path, audits: list[ArticleAudit]) -> None:
    rows = [flatten_audit(audit) for audit in audits]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()) if rows else ["status"])
        writer.writeheader()
        if rows:
            writer.writerows(rows)
        else:
            writer.writerow({"status": "no rows"})


def flatten_audit(audit: ArticleAudit) -> dict[str, Any]:
    data = asdict(audit)
    for key in ("function_scores", "required_prior_concepts", "introduced_concepts", "missing_prior_concepts", "severity_flags", "depends_on", "enables"):
        data[key] = json.dumps(data[key], ensure_ascii=False)
    return data


def write_workbook(path: Path, audits: list[ArticleAudit], scores: dict[str, int]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Verdict"
    ws.append(["metric", "value"])
    ws.append(["verdict", verdict(scores)])
    for key, value in scores.items():
        ws.append([key, value])
    rows = [flatten_audit(audit) for audit in audits]
    sheet = wb.create_sheet("Articles")
    if rows:
        sheet.append(list(rows[0].keys()))
        for row in rows:
            sheet.append([row[key] for key in rows[0].keys()])
    rec = wb.create_sheet("Recommended Order")
    rec.append(["new_order", "old_order", "title", "function", "reason"])
    for index, audit in enumerate(recommended_order(audits), start=1):
        rec.append([index, audit.ordinal, audit.title, audit.function, reason_for_position(audit)])
    for sheet_obj in wb.worksheets:
        for cell in sheet_obj[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
        for row in sheet_obj.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet_obj.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit an ordered article series for narrative and claim dependency flow.")
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--series-dir", type=Path)
    source.add_argument("--manifest", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()

    articles = load_articles_from_manifest(args.manifest) if args.manifest else load_articles_from_dir(args.series_dir)
    if not articles:
        raise SystemExit("No supported article files found.")

    audits = audit_articles(articles)
    scores = score_audit(audits)
    args.out.mkdir(parents=True, exist_ok=True)
    md = render_markdown(audits, scores)
    (args.out / "series-flow-audit.md").write_text(md, encoding="utf-8")
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "verdict": verdict(scores),
        "scores": scores,
        "articles": [asdict(audit) for audit in audits],
        "recommended_order": [audit.ordinal for audit in recommended_order(audits)],
        "missing_bridges": missing_bridges(audits),
    }
    (args.out / "series-flow-audit.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    write_csv(args.out / "series-flow-audit.csv", audits)
    write_workbook(args.out / "series-flow-audit.xlsx", audits, scores)
    print(f"Series flow audit written: {args.out}")
    print(f"Articles: {len(audits)}")
    print(f"Verdict: {payload['verdict']}")
    print(f"Overall score: {scores.get('overall_sequence_score', 0)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
