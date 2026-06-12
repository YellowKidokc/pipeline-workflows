#!/usr/bin/env python3
"""
fruits_coherence_engine.py

Best-current Fruits evaluator for David's Theophysics / MDA paper grader.

Purpose
-------
This is NOT just a Christian virtue word counter.

It runs a multi-layer Fruits analysis:

1. WORD TRACE
   Every trigger word/phrase is recorded with:
   file, paragraph, sentence, token/span, category, fruit, direction, weight, context.

2. SENTENCE SCORE
   Trigger scores aggregate into sentence-level fruit / anti-fruit pressure.

3. PARAGRAPH SCORE
   Sentence scores aggregate into paragraph scores.
   Paragraphs are also assigned a dominant role:
   HOOK, DEFINE, EXPLAIN, PROVE, EXAMPLE, OBJECTION, ANSWER,
   TRANSITION, CONCLUSION, APPLICATION, or MIXED.

4. DOMAIN POLARITY
   The script detects domain-language:
   medicine, engineering, logic, information, psychology, programming,
   AI safety, economics, thermodynamics, law, social systems, etc.
   This lets it score "type-safe" as constructive in programming,
   "homeostatic" as constructive in biology/medicine,
   "misaligned" as destructive in AI safety, etc.

5. STRUCTURAL INVARIANT CHECK
   Fruits are treated as invariants, not only words:
   Love         = relational cohesion / positive-sum bonding
   Joy          = resonance / positive feedback / flourishing
   Peace        = internal consistency / stability
   Patience     = iterative stability / convergence over time
   Kindness     = low-friction repair / non-harmful outward action
   Goodness     = generative surplus / non-parasitic output
   Faithfulness = identity preservation / invariant fidelity
   Gentleness   = power under restraint / non-damaging correction
   Self-Control = boundary integrity / bounded scope

6. OUTPUTS
   - JSON report
   - Markdown report
   - CSV trace files
   - optional XLSX workbook with all trace layers

Usage
-----
Basic:
    python fruits_coherence_engine.py paper.md

Folder:
    python fruits_coherence_engine.py ./articles --outdir fruits_reports

With your enhanced lexicon spreadsheet:
    python fruits_coherence_engine.py paper.md --lexicon paper_grader_lexicons_master_enhanced.xlsx --xlsx

Strict trace with all matched words:
    python fruits_coherence_engine.py paper.md --xlsx --context-window 70

Notes
-----
- This is a detector, not a truth oracle.
- Lexicon hits are signals, not final proof.
- The strongest output is when lexical signals, paragraph role,
  domain polarity, and structural invariants agree.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import math
import re
import statistics
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


# =============================================================================
# Core configuration
# =============================================================================

FRUITS = [
    "love", "joy", "peace", "patience", "kindness",
    "goodness", "faithfulness", "gentleness", "self_control",
]

DEFAULT_WEIGHTS = {
    "fruit": 1.00,
    "anti_fruit": -1.50,
    "domain_constructive": 0.70,
    "domain_destructive": -0.90,
    "grounding": 1.40,
    "evidence": 1.30,
    "falsifier": 1.20,
    "definition": 0.65,
    "hedge": 0.20,
    "absolute": -0.25,
    "propaganda": -2.00,
    "contradiction": -0.60,
    "jargon": -0.25,
}

NEGATORS = {
    "not", "no", "never", "without", "lack", "lacks", "lacking",
    "absence", "absent", "against", "anti", "non", "cannot", "can't",
}

OBJECTION_CUES = {
    "objection", "critic", "critics", "skeptic", "skeptics", "counterargument",
    "someone might argue", "one might object", "this would fail if",
    "the strongest objection", "a critic will say", "could be wrong",
}

ROLE_KEYWORDS = {
    "HOOK": [
        "what if", "the problem", "nobody", "strange", "paradox", "broken",
        "why", "the question", "imagine", "it starts", "the trap",
    ],
    "DEFINE": [
        "means", "defined", "definition", "is called", "refers to",
        "by .* we mean", "term", "symbol", "where", "denote",
    ],
    "EXPLAIN": [
        "in plain english", "this means", "in other words", "the idea is",
        "think of", "basically", "put simply", "what this says",
    ],
    "PROVE": [
        "because", "therefore", "evidence", "data", "shows", "demonstrates",
        "follows", "derivation", "proof", "verified", "measured", "observed",
    ],
    "EXAMPLE": [
        "for example", "for instance", "case study", "consider", "such as",
        "imagine", "like when", "example",
    ],
    "OBJECTION": [
        "objection", "critic", "skeptic", "however", "but", "counterargument",
        "one might object", "the strongest objection", "it could be argued",
    ],
    "ANSWER": [
        "response", "answer", "the reply", "this fails because", "but this misses",
        "what it does not address", "the correction is",
    ],
    "TRANSITION": [
        "now", "next", "this leads", "therefore the next", "from here",
        "which brings us", "the next question",
    ],
    "CONCLUSION": [
        "therefore", "so", "conclusion", "in summary", "the result",
        "the takeaway", "this is why", "the point is",
    ],
    "APPLICATION": [
        "what changes", "this means for", "application", "so what", "therefore we should",
        "if this is true", "practically",
    ],
}


# =============================================================================
# Built-in lexicon seeds
# =============================================================================

BUILTIN_FRUIT_TERMS = {
    "love": [
        "love", "agape", "charity", "compassion", "mercy", "forgiveness",
        "covenant", "communion", "hospitality", "solidarity", "benevolence",
        "cooperation", "reciprocity", "mutual", "positive-sum", "bond", "cohesion",
        "care", "sacrifice", "relationship", "relational", "symbiosis",
    ],
    "joy": [
        "joy", "rejoice", "delight", "gladness", "gratitude", "awe", "wonder",
        "resonance", "flourishing", "thrive", "positive feedback", "virtuous cycle",
        "hopeful energy", "celebration", "fulfillment",
    ],
    "peace": [
        "peace", "shalom", "harmony", "stable", "stability", "tranquility",
        "homeostasis", "internal consistency", "coherence", "non-conflict",
        "reconciliation", "wholeness", "integration", "balance", "rest",
    ],
    "patience": [
        "patience", "long-suffering", "forbearance", "endurance", "iterative",
        "over time", "convergence", "gradual", "steady", "wait", "delay",
        "slow", "season", "process", "development", "asymptotic", "persistence",
    ],
    "kindness": [
        "kindness", "kind", "gentle action", "benevolent", "helpful", "generous",
        "non-harm", "low-friction", "repair", "healing", "service", "neighbor",
        "tender", "considerate", "courtesy",
    ],
    "goodness": [
        "goodness", "good", "righteous", "generative", "surplus", "fruitful",
        "value creation", "beneficial", "constructive", "productive", "build",
        "create", "life-giving", "edifying", "moral excellence",
    ],
    "faithfulness": [
        "faithfulness", "faithful", "fidelity", "loyal", "loyalty", "consistent",
        "invariant", "preserve", "identity preservation", "steadfast", "reliable",
        "covenant-faithfulness", "integrity", "constancy", "adherence",
    ],
    "gentleness": [
        "gentleness", "gentle", "meekness", "humility", "power under restraint",
        "restraint", "careful correction", "nonviolent", "soft answer", "mild",
        "humble", "tender correction",
    ],
    "self_control": [
        "self-control", "self control", "temperance", "discipline", "boundary",
        "bounded", "restraint", "scope", "limit", "moderation", "regulated",
        "constraint", "containment", "impulse control", "guardrail",
    ],
}

BUILTIN_ANTI_FRUIT_TERMS = {
    "love": [
        "hatred", "malice", "cruelty", "contempt", "selfish ambition", "envy",
        "betrayal", "exploitation", "parasitic", "isolation", "fragmentation",
        "zero-sum", "domination", "capture", "abuse",
    ],
    "joy": [
        "despair", "hopeless", "misery", "nihilism", "deadness", "apathy",
        "burnout", "cynicism", "dread", "resentment", "negative feedback",
    ],
    "peace": [
        "discord", "faction", "factions", "dissension", "anxiety", "oscillation",
        "instability", "contradiction", "incoherence", "conflict", "war",
        "fragmented", "dysregulated",
    ],
    "patience": [
        "impatience", "impulsive", "rush", "force", "premature", "instant",
        "overreaction", "short-term", "panic", "shock", "volatility",
    ],
    "kindness": [
        "harshness", "cruel", "callous", "needless damage", "humiliation",
        "mockery", "bullying", "indifference", "hard-hearted", "unmerciful",
    ],
    "goodness": [
        "corruption", "parasitic", "extractive", "destructive", "wasteful",
        "degenerate", "moral hazard", "exploitation", "poison", "toxic",
    ],
    "faithfulness": [
        "unfaithful", "betrayal", "drift", "inconsistent", "compromise",
        "corrupt", "abandon", "distort", "false witness", "unreliable",
    ],
    "gentleness": [
        "violence", "domination", "overcorrection", "brutality", "crushing",
        "coercion", "arrogance", "pride", "reckless force",
    ],
    "self_control": [
        "unbounded", "runaway", "addiction", "debauchery", "drunkenness",
        "orgies", "sexual immorality", "leakage", "scope creep", "impulse",
        "unrestrained", "appetite", "boundary failure",
    ],
}

GROUNDING_TERMS = [
    "source", "citation", "data", "evidence", "measured", "observed", "verified",
    "replicated", "primary source", "dataset", "table", "method", "experiment",
    "derivation", "proof", "theorem", "lean", "falsifiable", "kill condition",
]

FALSIFIER_TERMS = [
    "falsify", "falsifiable", "kill condition", "would fail if", "fails if",
    "defeat condition", "counterexample", "what would break", "disconfirm",
    "rebut", "weaken", "collapse if", "null result",
]

PROPAGANDA_TERMS = [
    "everyone knows", "no one can deny", "obviously", "unquestionable",
    "enemies", "traitors", "destroy them", "only an idiot", "wake up",
]

JARGON_TERMS = [
    "paradigm", "synergy", "holistic", "robust framework", "novel architecture",
    "leverage", "disruptive", "revolutionary", "unprecedented",
]

ABSOLUTE_TERMS = [
    "proves", "proved", "proof", "undeniable", "impossible", "always", "never",
    "must", "cannot", "exact", "identical", "certain", "settled", "irrefutable",
    "perfect", "every", "all", "none",
]

HEDGE_TERMS = [
    "may", "might", "could", "suggests", "appears", "possibly", "likely",
    "seems", "candidate", "proposed", "within this framework", "model",
]


DOMAIN_POLARITY = {
    "medicine": {
        "constructive": ["healthy", "viable", "functional", "healing", "recovering", "homeostatic"],
        "destructive": ["sick", "nonviable", "dysfunctional", "degenerative", "disease", "pathological"],
        "invariant": "organism maintains viable function under stress",
    },
    "engineering": {
        "constructive": ["stable", "structural integrity", "load-bearing", "reliable", "fault-tolerant"],
        "destructive": ["failed", "collapsed", "fractured", "fatigue", "buckling", "brittle"],
        "invariant": "structure survives load and perturbation",
    },
    "logic": {
        "constructive": ["valid", "sound", "consistent", "non-contradiction", "entails"],
        "destructive": ["invalid", "unsound", "inconsistent", "contradiction", "non sequitur"],
        "invariant": "conclusion follows without contradiction",
    },
    "information": {
        "constructive": ["signal", "compressed", "coherent", "lossless", "recoverable", "fidelity"],
        "destructive": ["noise", "corrupted", "incoherent", "lossy", "unrecoverable", "distortion"],
        "invariant": "usable information is preserved under transmission",
    },
    "psychology": {
        "constructive": ["integrated", "resilient", "flourishing", "regulated", "secure attachment"],
        "destructive": ["dissociated", "fragile", "suffering", "dysregulated", "fragmented"],
        "invariant": "person maintains integrated agency under stress",
    },
    "game_theory": {
        "constructive": ["cooperation", "trust", "reciprocity", "positive-sum", "iterated trust"],
        "destructive": ["defection", "betrayal", "exploitation", "zero-sum", "free rider"],
        "invariant": "strategy sustains cooperative equilibrium",
    },
    "ai_safety": {
        "constructive": ["aligned", "corrective", "bounded", "safe", "interpretable", "calibrated"],
        "destructive": ["misaligned", "drifting", "unbounded", "unsafe", "opaque", "uncalibrated"],
        "invariant": "system remains aligned under optimization pressure",
    },
    "thermodynamics": {
        "constructive": ["ordered", "negentropic", "homeostatic", "free energy", "low entropy"],
        "destructive": ["disordered", "entropic", "equilibrium", "heat death", "dissipation"],
        "invariant": "system maintains usable order by paying energy cost",
    },
    "programming": {
        "constructive": ["functional", "type-safe", "terminating", "verified", "idempotent", "deterministic"],
        "destructive": ["buggy", "type-error", "non-terminating", "undefined", "exception", "race condition"],
        "invariant": "program executes correctly under constraints",
    },
    "law": {
        "constructive": ["lawful", "just", "rights-preserving", "due process", "accountable"],
        "destructive": ["unlawful", "unjust", "violating", "arbitrary", "corrupt"],
        "invariant": "rights and duties are preserved by enforceable order",
    },
    "social": {
        "constructive": ["trust", "cohesion", "cooperation", "shared norms", "social capital"],
        "destructive": ["betrayal", "fragmentation", "defection", "distrust", "collapse"],
        "invariant": "relationships sustain cooperation across time",
    },
    "economics": {
        "constructive": ["solvent", "productive", "growth", "savings", "value creation", "liquid"],
        "destructive": ["insolvent", "destructive", "collapse", "overleveraged", "moral hazard"],
        "invariant": "resources remain allocated toward sustainable value",
    },
    "aesthetics": {
        "constructive": ["harmonious", "coherent", "beautiful", "proportionate", "elegant"],
        "destructive": ["discordant", "incoherent", "ugly", "garish", "chaotic"],
        "invariant": "form preserves perceivable order and fittingness",
    },
}


# =============================================================================
# Data classes
# =============================================================================

@dataclass
class MatchTrace:
    file: str
    paragraph_id: int
    sentence_id: int
    start: int
    end: int
    matched_text: str
    category: str
    fruit: str
    domain: str
    direction: int
    base_weight: float
    adjusted_weight: float
    negated: bool
    role: str
    context: str

@dataclass
class SentenceScore:
    file: str
    paragraph_id: int
    sentence_id: int
    text: str
    role: str
    domain: str
    raw_score: float
    normalized_score: float
    fruit_scores: Dict[str, float]
    trigger_count: int
    constructive_count: int
    destructive_count: int

@dataclass
class ParagraphScore:
    file: str
    paragraph_id: int
    role: str
    role_confidence: float
    domain: str
    word_count: int
    sentence_count: int
    raw_score: float
    normalized_score: float
    fruit_scores: Dict[str, float]
    dominant_fruit: str
    anti_fruit_pressure: str
    invariant_notes: List[str]
    warnings: List[str]

@dataclass
class PaperScore:
    file: str
    word_count: int
    sentence_count: int
    paragraph_count: int
    raw_score: float
    normalized_score: float
    grade: str
    fruit_scores: Dict[str, float]
    domain_scores: Dict[str, float]
    top_positive_fruits: List[Tuple[str, float]]
    top_negative_fruits: List[Tuple[str, float]]
    paragraph_role_distribution: Dict[str, int]
    summary: str
    warnings: List[str]


# =============================================================================
# Utility functions
# =============================================================================

def clean_html(raw: str) -> str:
    raw = re.sub(r"(?is)<script.*?>.*?</script>", " ", raw)
    raw = re.sub(r"(?is)<style.*?>.*?</style>", " ", raw)
    raw = re.sub(r"(?is)<[^>]+>", " ", raw)
    return html.unescape(raw)

def read_text(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    if path.suffix.lower() in {".html", ".htm"}:
        raw = clean_html(raw)
    return raw.replace("\r\n", "\n").replace("\r", "\n")

def iter_input_files(path: Path) -> List[Path]:
    if path.is_file():
        return [path]
    allowed = {".md", ".txt", ".html", ".htm"}
    return sorted([p for p in path.rglob("*") if p.is_file() and p.suffix.lower() in allowed])

def split_paragraphs(text: str) -> List[str]:
    # Keep block-ish markdown paragraphs, but merge ultra-short lines.
    paras = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    out = []
    buf = ""
    for p in paras:
        if len(p) < 80 and not re.match(r"^#{1,6}\s+", p):
            buf = (buf + " " + p).strip()
        else:
            if buf:
                out.append(buf)
                buf = ""
            out.append(p)
    if buf:
        out.append(buf)
    return out

def split_sentences(paragraph: str) -> List[str]:
    p = re.sub(r"\s+", " ", paragraph.strip())
    if not p:
        return []
    parts = re.split(r"(?<=[.!?])\s+(?=[A-Z0-9\"'“‘])", p)
    if len(parts) == 1 and len(p) > 250:
        # fallback split on semicolons for long technical prose
        parts = re.split(r"(?<=;)\s+", p)
    return [s.strip() for s in parts if s.strip()]

def word_count(text: str) -> int:
    return len(re.findall(r"[A-Za-z0-9_𝔽φχζΩσρψαβγλμνκ]+", text))

def normalize_score(raw: float, count: int) -> float:
    # Smooth to 0-100; 50 = neutral
    if count <= 0:
        return 50.0
    scaled = raw / math.sqrt(max(count, 1))
    return round(max(0.0, min(100.0, 50.0 + 12.0 * scaled)), 2)

def grade_from_score(score: float) -> str:
    if score >= 90: return "A"
    if score >= 84: return "A-"
    if score >= 78: return "B+"
    if score >= 72: return "B"
    if score >= 66: return "B-"
    if score >= 60: return "C+"
    if score >= 54: return "C"
    if score >= 48: return "C-"
    if score >= 42: return "D"
    return "F"

def context_window(text: str, start: int, end: int, window: int) -> str:
    a = max(0, start - window)
    b = min(len(text), end + window)
    return re.sub(r"\s+", " ", text[a:b]).strip()

def term_to_pattern(term: str) -> str:
    term = term.strip()
    if not term:
        return r"a^"  # never
    # exact phrase with word-ish boundaries
    escaped = re.escape(term)
    if re.search(r"\s", term):
        return r"(?<!\w)" + escaped + r"(?!\w)"
    return r"\b" + escaped + r"\b"

def is_negated(sentence: str, match_start: int, window_tokens: int = 4) -> bool:
    before = sentence[:match_start].lower()
    tokens = re.findall(r"[a-z']+", before)[-window_tokens:]
    return any(t in NEGATORS for t in tokens)

def detect_role(sentence_or_paragraph: str) -> Tuple[str, float, Dict[str, int]]:
    text = sentence_or_paragraph.lower()
    counts: Dict[str, int] = {}
    for role, keys in ROLE_KEYWORDS.items():
        c = 0
        for k in keys:
            try:
                c += len(re.findall(k, text))
            except re.error:
                c += text.count(k)
        counts[role] = c
    if not any(counts.values()):
        return "EXPLAIN", 0.35, counts
    sorted_roles = sorted(counts.items(), key=lambda kv: kv[1], reverse=True)
    top, top_count = sorted_roles[0]
    second = sorted_roles[1][1] if len(sorted_roles) > 1 else 0
    confidence = (top_count + 1) / (top_count + second + 2)
    if second > 0 and top_count == second:
        return "MIXED", round(confidence, 2), counts
    return top, round(confidence, 2), counts

def detect_domain(text: str) -> Tuple[str, Dict[str, int]]:
    lower = text.lower()
    counts: Dict[str, int] = {}
    for domain, spec in DOMAIN_POLARITY.items():
        terms = spec["constructive"] + spec["destructive"]
        counts[domain] = sum(len(re.findall(term_to_pattern(t.lower()), lower)) for t in terms)
    best = max(counts.items(), key=lambda kv: kv[1])
    return (best[0] if best[1] > 0 else "general"), counts


# =============================================================================
# Lexicon loading
# =============================================================================

def add_term(lex: Dict[str, Dict[str, set]], category: str, key: str, value: Any) -> None:
    if value is None:
        return
    fruit = str(key).strip().lower().replace("-", "_").replace(" ", "_")
    term = str(value).strip()
    if not term or term.lower() in {"none", "nan", "n/a"}:
        return
    if fruit not in FRUITS:
        # common normalizations
        if fruit == "self-control":
            fruit = "self_control"
        elif fruit == "self control":
            fruit = "self_control"
        else:
            return
    lex.setdefault(category, defaultdict(set))
    lex[category][fruit].add(term.lower())

def load_builtin_lexicon() -> Dict[str, Dict[str, set]]:
    lex: Dict[str, Dict[str, set]] = {
        "fruit": defaultdict(set),
        "anti_fruit": defaultdict(set),
    }
    for fruit, terms in BUILTIN_FRUIT_TERMS.items():
        for t in terms:
            add_term(lex, "fruit", fruit, t)
    for fruit, terms in BUILTIN_ANTI_FRUIT_TERMS.items():
        for t in terms:
            add_term(lex, "anti_fruit", fruit, t)
    return lex

def load_excel_lexicon(path: Optional[Path]) -> Dict[str, Dict[str, set]]:
    lex = load_builtin_lexicon()
    if not path:
        return lex
    try:
        from openpyxl import load_workbook
    except Exception:
        print("WARNING: openpyxl not installed; using built-in lexicon only.")
        return lex

    if not path.exists():
        print(f"WARNING: lexicon file not found: {path}; using built-in lexicon only.")
        return lex

    wb = load_workbook(path, read_only=True, data_only=True)

    # Standard extracted sheets: source_file, collection, key, subkey, item_index, value
    for sheet, category in [
        ("FRUITS_LEX", "fruit"),
        ("FRUITS", "fruit"),
        ("ANTI_FRUITS", "anti_fruit"),
    ]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
        try:
            key_i = header.index("key")
            val_i = header.index("value")
        except ValueError:
            continue
        for row in rows[1:]:
            if len(row) <= max(key_i, val_i):
                continue
            add_term(lex, category, row[key_i], row[val_i])

    # Semantic architecture sheet: add adjacent cluster and anti-fruit opposed terms.
    if "Fruit Semantic Architecture" in wb.sheetnames:
        ws = wb["Fruit Semantic Architecture"]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows:
            if not row or len(row) < 6:
                continue
            fruit = row[1]
            adjacent = row[4]
            anti = row[5]
            if fruit is None or str(fruit).strip().lower() == "canonical (english)":
                continue
            fruit_key = str(fruit).strip().lower().replace("-", "_").replace(" ", "_")
            if fruit_key not in FRUITS:
                continue
            for item in str(adjacent or "").split(","):
                add_term(lex, "fruit", fruit_key, item.strip())
            for item in str(anti or "").split(","):
                add_term(lex, "anti_fruit", fruit_key, item.strip())

    return lex


# =============================================================================
# Scoring
# =============================================================================

def compile_patterns(lex: Dict[str, Dict[str, set]]) -> List[Tuple[str, str, str, re.Pattern]]:
    compiled = []
    for category, by_fruit in lex.items():
        for fruit, terms in by_fruit.items():
            for term in sorted(terms, key=len, reverse=True):
                if not term:
                    continue
                try:
                    compiled.append((category, fruit, term, re.compile(term_to_pattern(term), re.I)))
                except re.error:
                    continue

    # Non-fruit quality patterns
    for t in GROUNDING_TERMS:
        compiled.append(("grounding", "truth", t, re.compile(term_to_pattern(t), re.I)))
    for t in FALSIFIER_TERMS:
        compiled.append(("falsifier", "self_control", t, re.compile(term_to_pattern(t), re.I)))
    for t in PROPAGANDA_TERMS:
        compiled.append(("propaganda", "truth", t, re.compile(term_to_pattern(t), re.I)))
    for t in JARGON_TERMS:
        compiled.append(("jargon", "truth", t, re.compile(term_to_pattern(t), re.I)))
    for t in ABSOLUTE_TERMS:
        compiled.append(("absolute", "self_control", t, re.compile(term_to_pattern(t), re.I)))
    for t in HEDGE_TERMS:
        compiled.append(("hedge", "humility" if "humility" in FRUITS else "gentleness", t, re.compile(term_to_pattern(t), re.I)))
    return compiled

def domain_matches(sentence: str) -> List[Tuple[str, str, str, int, float]]:
    lower = sentence.lower()
    out = []
    for domain, spec in DOMAIN_POLARITY.items():
        for t in spec["constructive"]:
            for m in re.finditer(term_to_pattern(t.lower()), lower, flags=re.I):
                out.append((domain, t, "domain_constructive", 1, DEFAULT_WEIGHTS["domain_constructive"]))
        for t in spec["destructive"]:
            for m in re.finditer(term_to_pattern(t.lower()), lower, flags=re.I):
                out.append((domain, t, "domain_destructive", -1, DEFAULT_WEIGHTS["domain_destructive"]))
    return out

def structural_invariant_notes(paragraph: str, role: str, domain: str, fruit_scores: Dict[str, float]) -> Tuple[List[str], List[str]]:
    lower = paragraph.lower()
    notes = []
    warnings = []

    # Direct invariant cues beyond word hits.
    if any(x in lower for x in ["falsifiable", "kill condition", "defeat condition", "what would break"]):
        notes.append("Self-Control invariant: paper names boundaries/failure conditions.")
    if any(x in lower for x in ["source", "citation", "data", "observed", "measured", "verified"]):
        notes.append("Truth invariant: claim is connected to evidence/source language.")
    if any(x in lower for x in ["revise", "update", "wrong", "limitation", "open problem"]):
        notes.append("Humility/Gentleness invariant: paper leaves room for correction.")
    if any(x in lower for x in ["however", "objection", "critic", "counterargument"]):
        notes.append("Patience/Truth invariant: paper engages objection rather than skipping it.")
    if any(x in lower for x in ["relationship", "community", "mutual", "cooperation", "covenant"]):
        notes.append("Love invariant: relationship/cooperation structure is active.")
    if any(x in lower for x in ["consistent", "coherent", "contradiction", "internal"]):
        notes.append("Peace invariant: internal coherence/contradiction pressure is active.")

    # Warnings.
    if role == "MIXED":
        warnings.append("Paragraph role conflict: consider splitting into one dominant job.")
    if fruit_scores.get("self_control", 0) < -1.5:
        warnings.append("Self-Control pressure negative: possible overreach, unbounded claim, or scope leakage.")
    if fruit_scores.get("truth", 0) < -1.5:
        warnings.append("Truth pressure negative: possible propaganda, distortion, or unsupported certainty.")
    if domain == "general" and len(paragraph) > 500:
        warnings.append("Long paragraph without clear domain signal.")
    if any(x in lower for x in ["proves", "undeniable", "irrefutable", "perfect"]) and not any(x in lower for x in ["evidence", "proof", "derive", "verified", "formal"]):
        warnings.append("Absolute claim without nearby proof/evidence signal.")

    return notes, warnings

def score_files(files: List[Path], lex: Dict[str, Dict[str, set]], context_window_chars: int) -> Dict[str, Any]:
    patterns = compile_patterns(lex)

    all_traces: List[MatchTrace] = []
    all_sentences: List[SentenceScore] = []
    all_paragraphs: List[ParagraphScore] = []
    paper_scores: List[PaperScore] = []

    for path in files:
        text = read_text(path)
        paras = split_paragraphs(text)

        paper_fruit_scores = defaultdict(float)
        paper_domain_scores = defaultdict(float)
        paper_raw = 0.0
        sentence_count = 0
        role_counter = Counter()
        warnings = []

        for p_idx, para in enumerate(paras, start=1):
            p_role, p_conf, role_counts = detect_role(para)
            p_domain, p_domain_counts = detect_domain(para)
            role_counter[p_role] += 1
            sentences = split_sentences(para)

            p_raw = 0.0
            p_fruit_scores = defaultdict(float)
            p_constructive = 0
            p_destructive = 0

            for s_idx, sent in enumerate(sentences, start=1):
                sentence_count += 1
                s_role, _, _ = detect_role(sent)
                s_domain, _ = detect_domain(sent)
                s_raw = 0.0
                s_fruit_scores = defaultdict(float)
                triggers = 0
                constructive = 0
                destructive = 0

                # Fruit / anti-fruit / truth patterns.
                occupied_spans = []
                for category, fruit, term, pattern in patterns:
                    for m in pattern.finditer(sent):
                        # Avoid duplicate overlaps for same term category.
                        span = (m.start(), m.end(), category, fruit)
                        if span in occupied_spans:
                            continue
                        occupied_spans.append(span)

                        neg = is_negated(sent, m.start())
                        direction = 1 if DEFAULT_WEIGHTS.get(category, 0) >= 0 else -1
                        base = DEFAULT_WEIGHTS.get(category, 0.0)

                        # Negation logic: "not false" should not count as false;
                        # "not love" should create anti-pressure. Use attenuation/inversion.
                        adjusted = base
                        if neg:
                            adjusted = -0.65 * base

                        # If paragraph is an objection/critique paragraph, negative words are often mentioned to audit them.
                        # Attenuate unless the paragraph itself endorses the destructive term.
                        if p_role in {"OBJECTION", "ANSWER"} and category in {"anti_fruit", "propaganda", "contradiction"}:
                            adjusted *= 0.45

                        s_raw += adjusted
                        s_fruit_scores[fruit] += adjusted
                        triggers += 1
                        if adjusted >= 0:
                            constructive += 1
                        else:
                            destructive += 1

                        all_traces.append(MatchTrace(
                            file=str(path),
                            paragraph_id=p_idx,
                            sentence_id=s_idx,
                            start=m.start(),
                            end=m.end(),
                            matched_text=m.group(0),
                            category=category,
                            fruit=fruit,
                            domain=s_domain,
                            direction=1 if adjusted >= 0 else -1,
                            base_weight=round(base, 4),
                            adjusted_weight=round(adjusted, 4),
                            negated=neg,
                            role=s_role,
                            context=context_window(sent, m.start(), m.end(), context_window_chars),
                        ))

                # Domain polarity matches as separate traces, mapped to closest fruit invariants.
                for domain, term, category, direction, weight in domain_matches(sent):
                    # Map domains into likely fruit invariant.
                    fruit = {
                        "logic": "peace",
                        "information": "faithfulness",
                        "programming": "self_control",
                        "ai_safety": "self_control",
                        "thermodynamics": "patience",
                        "psychology": "peace",
                        "game_theory": "love",
                        "social": "love",
                        "medicine": "goodness",
                        "engineering": "faithfulness",
                        "economics": "goodness",
                        "law": "truth",
                        "aesthetics": "joy",
                    }.get(domain, "peace")
                    adjusted = weight
                    s_raw += adjusted
                    s_fruit_scores[fruit] += adjusted
                    paper_domain_scores[domain] += adjusted
                    triggers += 1
                    if adjusted >= 0:
                        constructive += 1
                    else:
                        destructive += 1
                    # approximate position
                    lower = sent.lower()
                    pos = lower.find(term.lower())
                    if pos < 0:
                        pos = 0
                    all_traces.append(MatchTrace(
                        file=str(path),
                        paragraph_id=p_idx,
                        sentence_id=s_idx,
                        start=pos,
                        end=pos + len(term),
                        matched_text=term,
                        category=category,
                        fruit=fruit,
                        domain=domain,
                        direction=direction,
                        base_weight=round(weight, 4),
                        adjusted_weight=round(adjusted, 4),
                        negated=False,
                        role=s_role,
                        context=context_window(sent, pos, pos + len(term), context_window_chars),
                    ))

                s_norm = normalize_score(s_raw, max(1, word_count(sent)))
                all_sentences.append(SentenceScore(
                    file=str(path),
                    paragraph_id=p_idx,
                    sentence_id=s_idx,
                    text=sent,
                    role=s_role,
                    domain=s_domain,
                    raw_score=round(s_raw, 4),
                    normalized_score=s_norm,
                    fruit_scores={f: round(s_fruit_scores.get(f, 0.0), 4) for f in FRUITS},
                    trigger_count=triggers,
                    constructive_count=constructive,
                    destructive_count=destructive,
                ))

                p_raw += s_raw
                p_constructive += constructive
                p_destructive += destructive
                for f, v in s_fruit_scores.items():
                    p_fruit_scores[f] += v

            p_norm = normalize_score(p_raw, max(1, word_count(para)))
            dominant = max(FRUITS, key=lambda f: p_fruit_scores.get(f, 0.0))
            anti = min(FRUITS, key=lambda f: p_fruit_scores.get(f, 0.0))
            notes, p_warnings = structural_invariant_notes(
                para, p_role, p_domain, {f: p_fruit_scores.get(f, 0.0) for f in FRUITS}
            )
            warnings.extend([f"P{p_idx}: {w}" for w in p_warnings])

            all_paragraphs.append(ParagraphScore(
                file=str(path),
                paragraph_id=p_idx,
                role=p_role,
                role_confidence=p_conf,
                domain=p_domain,
                word_count=word_count(para),
                sentence_count=len(sentences),
                raw_score=round(p_raw, 4),
                normalized_score=p_norm,
                fruit_scores={f: round(p_fruit_scores.get(f, 0.0), 4) for f in FRUITS},
                dominant_fruit=dominant,
                anti_fruit_pressure=anti if p_fruit_scores.get(anti, 0.0) < 0 else "",
                invariant_notes=notes,
                warnings=p_warnings,
            ))

            paper_raw += p_raw
            for f, v in p_fruit_scores.items():
                paper_fruit_scores[f] += v

        wc = word_count(text)
        norm = normalize_score(paper_raw, max(1, wc))
        grade = grade_from_score(norm)
        fruit_scores = {f: round(paper_fruit_scores.get(f, 0.0), 4) for f in FRUITS}
        top_pos = sorted(fruit_scores.items(), key=lambda kv: kv[1], reverse=True)[:3]
        top_neg = sorted(fruit_scores.items(), key=lambda kv: kv[1])[:3]

        if norm >= 70:
            summary = "Strong constructive/coherence signal. Review trace to confirm invariants are real, not merely verbal."
        elif norm >= 55:
            summary = "Mixed but usable. Strong enough for review; likely needs targeted structure/claim cleanup."
        else:
            summary = "Weak or destructive signal. Requires revision or structural audit before publication."

        paper_scores.append(PaperScore(
            file=str(path),
            word_count=wc,
            sentence_count=sentence_count,
            paragraph_count=len(paras),
            raw_score=round(paper_raw, 4),
            normalized_score=norm,
            grade=grade,
            fruit_scores=fruit_scores,
            domain_scores={k: round(v, 4) for k, v in sorted(paper_domain_scores.items()) if abs(v) > 0},
            top_positive_fruits=top_pos,
            top_negative_fruits=top_neg,
            paragraph_role_distribution=dict(role_counter),
            summary=summary,
            warnings=warnings[:50],
        ))

    return {
        "schema": "fruits_coherence_engine_v1",
        "note": "Lexicon and domain-polarity detector. Signals are auditable; they are not final truth judgments.",
        "papers": [asdict(p) for p in paper_scores],
        "paragraphs": [asdict(p) for p in all_paragraphs],
        "sentences": [asdict(s) for s in all_sentences],
        "trace": [asdict(t) for t in all_traces],
    }


# =============================================================================
# Output
# =============================================================================

def write_json(path: Path, report: Dict[str, Any]) -> None:
    path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

def write_csvs(outdir: Path, report: Dict[str, Any]) -> None:
    tables = {
        "paper_scores.csv": report["papers"],
        "paragraph_scores.csv": report["paragraphs"],
        "sentence_scores.csv": report["sentences"],
        "word_trace.csv": report["trace"],
    }
    for fname, rows in tables.items():
        fpath = outdir / fname
        if not rows:
            fpath.write_text("", encoding="utf-8")
            continue
        # flatten nested dicts/lists for CSV
        flat_rows = []
        for row in rows:
            flat = {}
            for k, v in row.items():
                if isinstance(v, (dict, list)):
                    flat[k] = json.dumps(v, ensure_ascii=False)
                else:
                    flat[k] = v
            flat_rows.append(flat)
        with fpath.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(flat_rows[0].keys()))
            writer.writeheader()
            writer.writerows(flat_rows)

def write_xlsx(path: Path, report: Dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        print("WARNING: openpyxl not installed; skipping XLSX output.")
        return

    wb = Workbook()
    # remove default
    ws = wb.active
    ws.title = "paper_scores"

    sheet_map = [
        ("paper_scores", report["papers"]),
        ("paragraph_scores", report["paragraphs"]),
        ("sentence_scores", report["sentences"]),
        ("word_trace", report["trace"]),
    ]

    for idx, (name, rows) in enumerate(sheet_map):
        ws = wb["paper_scores"] if idx == 0 else wb.create_sheet(name)
        if not rows:
            ws.append(["NO DATA"])
            continue
        flat_rows = []
        for row in rows:
            flat = {}
            for k, v in row.items():
                if isinstance(v, (dict, list)):
                    flat[k] = json.dumps(v, ensure_ascii=False)
                else:
                    flat[k] = v
            flat_rows.append(flat)
        headers = list(flat_rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True)
        for row in flat_rows:
            ws.append([row.get(h, "") for h in headers])
        # freeze header and set widths
        ws.freeze_panes = "A2"
        for col_idx, header in enumerate(headers, start=1):
            max_len = min(60, max([len(str(header))] + [len(str(r.get(header, ""))) for r in flat_rows[:200]]))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)

    wb.save(path)

def render_markdown(report: Dict[str, Any]) -> str:
    lines = [
        "# Fruits Coherence Engine Report",
        "",
        report.get("note", ""),
        "",
        "## Paper Summary",
        "",
    ]
    for p in report["papers"]:
        lines += [
            f"### {Path(p['file']).name}",
            f"- Grade: **{p['grade']}**",
            f"- Normalized score: **{p['normalized_score']}**",
            f"- Raw score: `{p['raw_score']}`",
            f"- Words: {p['word_count']}",
            f"- Paragraphs: {p['paragraph_count']}",
            f"- Top positive fruits: {', '.join(f'{k}={v}' for k,v in p['top_positive_fruits'])}",
            f"- Top negative fruits: {', '.join(f'{k}={v}' for k,v in p['top_negative_fruits'])}",
            f"- Summary: {p['summary']}",
            "",
        ]
        if p["warnings"]:
            lines.append("#### Warnings")
            for w in p["warnings"][:15]:
                lines.append(f"- {w}")
            lines.append("")

    lines += [
        "## Top Paragraphs by Constructive Score",
        "",
    ]
    top_paras = sorted(report["paragraphs"], key=lambda r: r["normalized_score"], reverse=True)[:15]
    for r in top_paras:
        lines.append(
            f"- `{Path(r['file']).name}` P{r['paragraph_id']} "
            f"score={r['normalized_score']} role={r['role']} domain={r['domain']} "
            f"dominant={r['dominant_fruit']}"
        )

    lines += [
        "",
        "## Lowest Paragraphs / Revision Targets",
        "",
    ]
    low_paras = sorted(report["paragraphs"], key=lambda r: r["normalized_score"])[:15]
    for r in low_paras:
        lines.append(
            f"- `{Path(r['file']).name}` P{r['paragraph_id']} "
            f"score={r['normalized_score']} role={r['role']} domain={r['domain']} "
            f"anti={r['anti_fruit_pressure'] or 'none'} warnings={'; '.join(r['warnings'][:2])}"
        )

    lines += [
        "",
        "## Method Note",
        "",
        "This report is a traceable detector. It does not prove a paper is morally true. "
        "The important feature is the audit trail: every score can be traced back to a word, "
        "sentence, paragraph role, domain polarity, and structural invariant warning.",
    ]
    return "\n".join(lines)

def write_markdown(path: Path, report: Dict[str, Any]) -> None:
    path.write_text(render_markdown(report), encoding="utf-8")


# =============================================================================
# CLI
# =============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(description="Domain-aware Fruits of the Spirit coherence evaluator.")
    parser.add_argument("input", type=Path, help="Markdown/text/html file or folder.")
    parser.add_argument("--lexicon", type=Path, default=None, help="Optional enhanced lexicon XLSX.")
    parser.add_argument("--outdir", type=Path, default=Path("fruits_engine_reports"), help="Output directory.")
    parser.add_argument("--xlsx", action="store_true", help="Write Excel workbook with trace sheets.")
    parser.add_argument("--context-window", type=int, default=80, help="Characters around each match in word trace.")
    args = parser.parse_args()

    args.outdir.mkdir(parents=True, exist_ok=True)
    files = iter_input_files(args.input)
    if not files:
        raise SystemExit(f"No input files found: {args.input}")

    lex = load_excel_lexicon(args.lexicon)
    report = score_files(files, lex, args.context_window)

    json_path = args.outdir / "fruits_coherence_report.json"
    md_path = args.outdir / "fruits_coherence_report.md"

    write_json(json_path, report)
    write_markdown(md_path, report)
    write_csvs(args.outdir, report)

    if args.xlsx:
        write_xlsx(args.outdir / "fruits_coherence_report.xlsx", report)

    print(f"Files analyzed: {len(files)}")
    print(f"Word trace rows: {len(report['trace'])}")
    print(f"Wrote: {json_path}")
    print(f"Wrote: {md_path}")
    print(f"Wrote CSVs in: {args.outdir}")
    if args.xlsx:
        print(f"Wrote: {args.outdir / 'fruits_coherence_report.xlsx'}")

    for p in report["papers"]:
        print(f"{Path(p['file']).name}: {p['grade']} ({p['normalized_score']}) — {p['summary']}")

    return 0

if __name__ == "__main__":
    raise SystemExit(main())
