from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import math
import os
import re
import sys
from datetime import datetime
from html import unescape
from pathlib import Path


HERE = Path(__file__).resolve().parent
CONFIG_PATH = Path(os.environ.get("FRUITS_BRIDGE_CONFIG", HERE / "fruits_of_spirit_config.json"))
CONFIG = json.loads(CONFIG_PATH.read_text(encoding="utf-8")) if CONFIG_PATH.exists() else {}
TRUTH_ENGINE = Path(os.environ.get("TRUTH_ENGINE_ROOT", CONFIG.get("truth_engine_root", r"\\dlowenas\github\Truth Engine (1)")))
FRUITS_PIPELINE = TRUTH_ENGINE / "scripts" / "fruits_pipeline"
DEFAULT_OUTPUT = Path(os.environ.get("FRUITS_OUTPUT", CONFIG.get("default_output", HERE / "EXPORTS" / "reports" / "fruits_of_spirit")))
SEMANTIC_DIMS = int(CONFIG.get("semantic_vector_dimensions", 2048))
TEXT_EXTS = {".md", ".txt", ".html", ".htm"}
CANONICAL_LEXICON_PATH = Path(os.environ.get(
    "PAPER_GRADER_LEXICON_XLSX",
    CONFIG.get(
        "paper_grader_lexicon_xlsx",
        r"\\dlowenas\HPWorkstation\Desktop\paper_grader_lexicons_master_enhanced.xlsx",
    ),
))

FRUIT_ANCHORS = {
    "love": "Sustained care for others at personal cost, seeking the good of the other without exploitation.",
    "joy": "Stable hope and gratitude under pressure, not dependent on immediate circumstances.",
    "peace": "Coherent internal order, reconciliation, non-anxious stability, and resolved conflict.",
    "patience": "Long-horizon endurance, willingness to wait, and tolerance of slow repair or delayed reward.",
    "kindness": "Concrete benevolence, warmth, generosity, and attention to the needs of another.",
    "goodness": "Integrity that produces positive externalities, honors constraints, and refuses hidden harm.",
    "faithfulness": "Reliable covenantal consistency, promise-keeping, trustworthiness, and temporal fidelity.",
    "gentleness": "Non-coercive strength, careful force, precision with care, and respect for agency.",
    "self_control": "Disciplined restraint, impulse governance, and voluntary subordination of appetite to principle.",
}

ANTI_ANCHORS = {
    "coercion": "Pressure that overrides agency through threat, intimidation, manipulation, or forced compliance.",
    "domination": "Control-seeking posture that centralizes authority and treats others as instruments.",
    "deception": "Strategic concealment, misdirection, fraud, or manipulation of appearances.",
    "fear_shame": "Fear, shame, panic, urgency, or identity threat used to move the reader.",
    "fragmentation": "Breakdown of coherence, contradiction, despair, alienation, or loss of stable meaning.",
    "exploitation": "Extraction of benefit from others while externalizing harm or cost.",
    "certainty_inflation": "Unwarranted overclaiming, refusal of uncertainty, or exaggerated confidence.",
    "tribal_binding": "Identity fusion, in-group flattery, out-group contempt, or loyalty pressure.",
}

FRUIT_TERMS = {
    "love", "joy", "peace", "patience", "kindness", "goodness", "faithfulness",
    "gentleness", "self-control", "truth", "grace", "mercy", "hope", "unity",
    "coherence", "humility", "charity", "forgiveness", "compassion", "tenderness",
    "devotion", "steadfast", "endurance", "reconciliation", "restoration", "healing",
    "freedom", "gratitude", "sacrifice", "servant", "generous", "meekness", "purity",
    "trust", "integrity", "honest", "honesty", "transparency", "accountable",
    "accountability", "stewardship", "responsible", "responsibility", "wisdom",
    "discernment", "courage", "brave", "protect", "nurture", "empower", "collaborate",
    "cooperation", "dignity", "respect", "liberty", "justice", "righteous",
    "faithful", "diligent", "perseverance", "wholeness", "harmony", "fruitful",
    "productive", "flourish", "thrive", "prosper", "blessing", "encourage",
    "uplift", "edify", "build", "create", "cultivate",
}
ANTI_FRUIT_TERMS = {
    "chaos", "coercion", "domination", "deceit", "fragmentation", "fear", "pride",
    "violence", "contradiction", "confusion", "envy", "wrath", "malice", "vanity",
    "inversion", "predation", "hostility", "corruption", "control", "surveillance",
    "manipulate", "manipulation", "exploit", "oppress", "oppression", "tyranny",
    "despotism", "usurp", "enslave", "slavery", "censor", "censorship", "propaganda",
    "deception", "deceive", "fraud", "covert", "secret", "hidden", "suppress",
    "sabotage", "destroy", "destruction", "toxic", "weaponize", "kill", "harm",
    "abuse", "predatory", "parasitic", "centralize", "monopoly", "totalitarian",
    "authoritarian", "dystopian",
}
GROUNDING_TERMS = {
    "definition", "evidence", "equation", "axiom", "theorem", "proof", "because",
    "therefore", "falsification", "prediction", "experiment", "scripture",
    "reference", "claim", "derivation", "data", "study", "research", "survey",
    "statistics", "statistically", "measured", "measurement", "documented",
    "documentation", "audit", "published", "peer-reviewed", "replicate",
    "correlation", "causation", "according", "cited", "citation", "source",
    "paper", "journal", "specifically", "precisely", "verified", "confirmed",
    "historical", "factual", "empirical", "observable",
}
CONTRADICTION_TERMS = {
    "however", "although", "yet", "nevertheless", "contradiction", "versus",
    "against", "opposed", "conflict", "opposing", "opposition", "contrary",
    "incompatible", "but", "despite", "notwithstanding", "paradox", "inconsistent",
    "inconsistency", "discrepancy", "allegedly", "supposedly", "claimed",
    "purported", "questionable", "impossible", "implausible", "unlikely", "absurd",
    "illogical", "irrational", "hypocrisy", "hypocritical",
}

LEXICON_STATUS = {
    "source": "embedded fallback",
    "path": "",
    "loaded": False,
    "error": "",
}


def _normalize_term(value) -> str:
    term = str(value or "").strip().lower()
    term = re.sub(r"\s+", " ", term)
    return term


def _sheet_terms(wb, sheet_name: str, column_name: str) -> set[str]:
    if sheet_name not in wb.sheetnames:
        return set()
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(cell or "").strip() for cell in next(rows)]
    except StopIteration:
        return set()
    if column_name not in header:
        return set()
    index = header.index(column_name)
    terms: set[str] = set()
    for row in rows:
        if index >= len(row):
            continue
        term = _normalize_term(row[index])
        if term:
            terms.add(term)
    return terms


def _semantic_bucket_terms(wb, buckets: set[str] | None = None, danger_levels: set[str] | None = None) -> set[str]:
    if "SEMANTIC_BUCKETS" not in wb.sheetnames:
        return set()
    ws = wb["SEMANTIC_BUCKETS"]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(cell or "").strip() for cell in next(rows)]
    except StopIteration:
        return set()
    required = {"term", "bucket", "danger_level"}
    if not required.issubset(set(header)):
        return set()
    term_i = header.index("term")
    bucket_i = header.index("bucket")
    danger_i = header.index("danger_level")
    terms: set[str] = set()
    for row in rows:
        term = _normalize_term(row[term_i] if term_i < len(row) else "")
        bucket = _normalize_term(row[bucket_i] if bucket_i < len(row) else "").upper()
        danger = _normalize_term(row[danger_i] if danger_i < len(row) else "")
        if not term:
            continue
        if buckets and bucket not in buckets:
            continue
        if danger_levels and danger not in danger_levels:
            continue
        terms.add(term)
    return terms


def load_canonical_lexicons(path: Path) -> None:
    global FRUIT_TERMS, ANTI_FRUIT_TERMS, GROUNDING_TERMS, CONTRADICTION_TERMS
    LEXICON_STATUS.update({"path": str(path)})
    if not path.exists():
        LEXICON_STATUS.update({"error": f"Workbook not found: {path}"})
        return
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        fruit = (
            _sheet_terms(wb, "FRUITS", "value")
            | _sheet_terms(wb, "FRUITS_LEX", "value")
        )
        anti = _sheet_terms(wb, "ANTI_FRUITS", "value")
        grounding = (
            _sheet_terms(wb, "EVIDENCE_TERMS", "term")
            | _sheet_terms(wb, "FALSIFY_TERMS", "term")
            | _sheet_terms(wb, "DEPENDENCY_TERMS", "term")
            | _sheet_terms(wb, "BRIDGE_TERMS", "term")
            | _semantic_bucket_terms(wb, buckets={"EVIDENCE_STATUS", "CLAIM_STRENGTH"})
        )
        contradiction = (
            _sheet_terms(wb, "NEGATION_TERMS", "term")
            | _sheet_terms(wb, "HEDGE_TERMS", "term")
            | _sheet_terms(wb, "ABSOLUTE_TERMS", "term")
            | _semantic_bucket_terms(
                wb,
                buckets={"DANGER_LANGUAGE", "FORMAL_PROOF", "THEOLOGICAL_STATUS", "REVIEWER_RISK"},
            )
            | _semantic_bucket_terms(wb, danger_levels={"high", "critical"})
        )
        if fruit:
            FRUIT_TERMS = fruit
        if anti:
            ANTI_FRUIT_TERMS = anti
        if grounding:
            GROUNDING_TERMS = grounding
        if contradiction:
            CONTRADICTION_TERMS = contradiction
        LEXICON_STATUS.update({
            "source": "canonical workbook",
            "loaded": True,
            "error": "",
            "fruit_terms": len(FRUIT_TERMS),
            "anti_fruit_terms": len(ANTI_FRUIT_TERMS),
            "grounding_terms": len(GROUNDING_TERMS),
            "contradiction_terms": len(CONTRADICTION_TERMS),
        })
    except Exception as exc:
        LEXICON_STATUS.update({"error": repr(exc)})


load_canonical_lexicons(CANONICAL_LEXICON_PATH)


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load {name} from {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def setup_truth_engine(export_excel: bool):
    if not TRUTH_ENGINE.exists():
        raise FileNotFoundError(f"Truth Engine folder not found: {TRUTH_ENGINE}")
    sys.path.insert(0, str(TRUTH_ENGINE / "src"))
    sys.path.insert(0, str(TRUTH_ENGINE / "modules"))
    sys.path.insert(0, str(FRUITS_PIPELINE))

    try:
        score = load_module("fruits_score", FRUITS_PIPELINE / "2_score.py")
    except ModuleNotFoundError as exc:
        if exc.name != "PySide6":
            raise
        christ = load_module("christ_vector_module", TRUTH_ENGINE / "modules" / "christ_vector_module.py")
        score = LocalScoreModule(christ.compute_christ_vector_alignment)
    enrich = load_module("fruits_enrich", FRUITS_PIPELINE / "3_enrich.py")
    export = None
    if export_excel:
        try:
            export = load_module("fruits_export", FRUITS_PIPELINE / "4_export_excel.py")
            export.TEMPLATE = Path(CONFIG.get("fruits_template", TRUTH_ENGINE / "Fruits Template.xlsx"))
        except ModuleNotFoundError as exc:
            if exc.name != "openpyxl":
                raise
            print("openpyxl is not installed; skipping Fruits Template Excel exports.", flush=True)
    return score, enrich, export


class LocalScoreModule:
    def __init__(self, christ_vector_func):
        self._christ_vector_func = christ_vector_func

    def compute_scores(self, text: str, fname: str = "") -> dict:
        sentences, agg = self._score_document(text)
        tokens = agg.get("tokens", 1)
        truth = agg.get("truth", 0)
        fruit = agg.get("fruit", 0)
        grounding = agg.get("grounding", 0)
        coherence = agg.get("coherence", 0)
        anti_fruit = agg.get("anti_fruit", 0)
        contradiction = agg.get("contradiction", 0)
        signal = fruit + grounding + coherence
        noise = anti_fruit + contradiction
        snr = signal / (noise + 1e-6)
        assertion_ratio = truth / (grounding + 1e-6)
        contradiction_pressure = contradiction / (truth + 1e-6)
        coherence_density = coherence / (tokens + 1)
        propaganda = anti_fruit * 0.5 + (1 - grounding) * 0.3 + contradiction * 0.2
        christ_vector = self._christ_vector_func(text)

        def top_sentence(key: str) -> tuple[str, float]:
            ranked = sorted(sentences, key=lambda item: item.get(key, 0), reverse=True)
            if ranked:
                return ranked[0]["text"][:200], round(ranked[0].get(key, 0), 4)
            return "", 0.0

        return {
            "agg": agg,
            "sentences": sentences,
            "tokens": tokens,
            "snr": round(snr, 4),
            "assertion_ratio": round(assertion_ratio, 4),
            "contradiction_pressure": round(contradiction_pressure, 4),
            "coh_density": round(coherence_density, 6),
            "propaganda": round(propaganda, 4),
            "christ_vector": christ_vector,
            "top_fruit": top_sentence("fruit"),
            "top_anti": top_sentence("anti_fruit"),
            "top_ground": top_sentence("grounding"),
            "top_contra": top_sentence("contradiction"),
            "source_name": fname,
        }

    @staticmethod
    def _tokenize(text: str) -> list[str]:
        return re.findall(r"[A-Za-z][A-Za-z\-']+", text.lower())

    @staticmethod
    def _clamp(value: float, low: float = 0.0, high: float = 1.0) -> float:
        return max(low, min(high, value))

    @classmethod
    def _term_hits(cls, text: str, tokens: list[str], terms: set[str]) -> int:
        token_set = {term for term in terms if " " not in term}
        hits = sum(1 for token in tokens if token in token_set)
        for phrase in (term for term in terms if " " in term):
            hits += len(re.findall(r"\b" + re.escape(phrase) + r"\b", text, re.I))
        return hits

    @staticmethod
    def _split_sentences(text: str) -> list[str]:
        return [s.strip() for s in re.split(r'(?<=[.!?])\s+(?=[A-Z0-9"\'])', text.strip()) if s.strip()]

    @classmethod
    def _lexical_overlap(cls, left: str, right: str) -> float:
        left_set = set(cls._tokenize(left))
        right_set = set(cls._tokenize(right))
        if not left_set or not right_set:
            return 0.0
        return len(left_set & right_set) / len(left_set | right_set)

    @classmethod
    def _score_text(cls, text: str) -> dict:
        tokens = cls._tokenize(text)
        token_count = max(1, len(tokens))
        sentences = cls._split_sentences(text)
        fruit = cls._clamp(cls._term_hits(text, tokens, FRUIT_TERMS) / max(3.0, token_count / 35.0))
        anti = cls._clamp(cls._term_hits(text, tokens, ANTI_FRUIT_TERMS) / max(3.0, token_count / 35.0))
        ground = cls._clamp(cls._term_hits(text, tokens, GROUNDING_TERMS) / max(3.0, token_count / 40.0))
        contra = cls._clamp(cls._term_hits(text, tokens, CONTRADICTION_TERMS) / max(2.0, token_count / 50.0))
        if len(sentences) >= 2:
            overlaps = [cls._lexical_overlap(sentences[i], sentences[i + 1]) for i in range(len(sentences) - 1)]
            continuity = sum(overlaps) / len(overlaps)
        else:
            continuity = 0.45
        avg_len = token_count / max(1, len(sentences))
        balance = 1.0 - min(abs(avg_len - 22.0) / 30.0, 1.0)
        coherence = cls._clamp(continuity * 0.6 + balance * 0.4)
        integration = cls._clamp((fruit + ground + coherence) / 3.0)
        truth = cls._clamp(coherence * 0.34 + ground * 0.23 + fruit * 0.23 + integration * 0.20 - anti * 0.14 - contra * 0.12)
        return {
            "truth": truth,
            "coherence": coherence,
            "fruit": fruit,
            "anti_fruit": anti,
            "grounding": ground,
            "contradiction": contra,
            "integration": integration,
            "tokens": token_count,
            "sentences": len(sentences),
        }

    @classmethod
    def _score_document(cls, text: str) -> tuple[list[dict], dict]:
        paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
        all_sentences = []
        for paragraph in paragraphs:
            for sentence in cls._split_sentences(paragraph):
                score = cls._score_text(sentence)
                score["text"] = sentence
                all_sentences.append(score)
        if not all_sentences:
            return [], cls._score_text(text)
        keys = ["truth", "coherence", "fruit", "anti_fruit", "grounding", "contradiction", "integration"]
        agg = {key: sum(sentence[key] for sentence in all_sentences) / len(all_sentences) for key in keys}
        agg["sentences"] = len(all_sentences)
        agg["tokens"] = sum(sentence["tokens"] for sentence in all_sentences)
        return all_sentences, agg


def strip_html(raw: str) -> str:
    raw = re.sub(r"(?is)<script.*?>.*?</script>", " ", raw)
    raw = re.sub(r"(?is)<style.*?>.*?</style>", " ", raw)
    raw = re.sub(r"(?is)<title.*?>(.*?)</title>", r"\n# \1\n", raw)
    raw = re.sub(r"(?is)</(h[1-6]|p|div|section|article|li|tr)>", "\n", raw)
    raw = re.sub(
        r"(?is)<h([1-6])[^>]*>(.*?)</h\1>",
        lambda m: "\n" + ("#" * int(m.group(1))) + " " + re.sub(r"<[^>]+>", "", m.group(2)) + "\n",
        raw,
    )
    raw = re.sub(r"(?s)<[^>]+>", " ", raw)
    return unescape(raw)


def read_text(path: Path) -> str:
    text = path.read_text(encoding="utf-8", errors="ignore")
    if text.startswith("---"):
        parts = text.split("---", 2)
        if len(parts) >= 3:
            text = parts[2]
    if path.suffix.lower() in {".html", ".htm"}:
        text = strip_html(text)
    text = re.sub(r"\[\[.*?\]\]", "", text)
    text = re.sub(r"\$\$.*?\$\$", " [BLOCK_EQUATION] ", text, flags=re.DOTALL)
    text = re.sub(r"\$[^$]+\$", " [INLINE_MATH] ", text)
    text = re.sub(r"^#{1,6}\s+", "", text, flags=re.MULTILINE)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def find_inputs(input_path: Path, pattern: str, recursive: bool = True) -> list[Path]:
    if input_path.is_file():
        return [input_path] if input_path.suffix.lower() in TEXT_EXTS else []
    iterator = input_path.rglob(pattern) if recursive else input_path.glob(pattern)
    files = [p for p in iterator if p.is_file() and p.suffix.lower() in TEXT_EXTS]
    return sorted(files)


def chunks_for_semantics(text: str, max_chunks: int = 120) -> list[str]:
    sentences = LocalScoreModule._split_sentences(text)
    if not sentences:
        return [text[:1000]] if text else []
    if len(sentences) <= max_chunks:
        return sentences
    step = max(1, len(sentences) // max_chunks)
    return sentences[::step][:max_chunks]


def hashed_vector(text: str, dims: int = SEMANTIC_DIMS) -> list[float]:
    tokens = LocalScoreModule._tokenize(text)
    vector = [0.0] * dims
    grams = []
    for n in (1, 2, 3):
        grams.extend(" ".join(tokens[i:i + n]) for i in range(0, max(0, len(tokens) - n + 1)))
    for gram in grams:
        digest = hashlib.blake2b(gram.encode("utf-8"), digest_size=8).digest()
        idx = int.from_bytes(digest[:4], "little") % dims
        sign = 1.0 if digest[4] % 2 == 0 else -1.0
        weight = 1.0 + min(len(gram.split()) - 1, 2) * 0.35
        vector[idx] += sign * weight
    norm = math.sqrt(sum(value * value for value in vector)) or 1.0
    return [value / norm for value in vector]


def cosine(left: list[float], right: list[float]) -> float:
    return sum(a * b for a, b in zip(left, right))


class SemanticAnchorScorer:
    def __init__(self):
        self.backend = "hashed-ngram"
        self.model = None
        if CONFIG.get("semantic_anchor_backend", "auto") != "hashed":
            try:
                from sentence_transformers import SentenceTransformer
                self.model = SentenceTransformer(os.environ.get("FRUITS_EMBEDDING_MODEL", "sentence-transformers/all-MiniLM-L6-v2"))
                self.backend = "sentence-transformers"
            except Exception:
                self.model = None
        self.fruit_vectors = self._encode_map(FRUIT_ANCHORS)
        self.anti_vectors = self._encode_map(ANTI_ANCHORS)

    def _encode(self, texts: list[str]) -> list[list[float]]:
        if self.model is not None:
            vectors = self.model.encode(texts, normalize_embeddings=True)
            return [list(map(float, vector)) for vector in vectors]
        return [hashed_vector(text) for text in texts]

    def _encode_map(self, anchors: dict[str, str]) -> dict[str, list[float]]:
        names = list(anchors.keys())
        vectors = self._encode([anchors[name] for name in names])
        return dict(zip(names, vectors))

    def score(self, text: str) -> dict:
        chunks = chunks_for_semantics(text)
        if not chunks:
            return {
                "semantic_backend": self.backend,
                "semantic_ontology": "explicit coherence ontology",
                "semantic_fruit_alignment": 0.0,
                "semantic_anti_alignment": 0.0,
                "semantic_net_alignment": 0.0,
                "semantic_dominant_anchor": "",
                "semantic_dominant_anti_anchor": "",
            }
        chunk_vectors = self._encode(chunks)
        fruit_scores = {name: 0.0 for name in self.fruit_vectors}
        anti_scores = {name: 0.0 for name in self.anti_vectors}
        for vector in chunk_vectors:
            for name, anchor_vector in self.fruit_vectors.items():
                fruit_scores[name] += max(0.0, cosine(vector, anchor_vector))
            for name, anchor_vector in self.anti_vectors.items():
                anti_scores[name] += max(0.0, cosine(vector, anchor_vector))
        divisor = len(chunk_vectors)
        fruit_scores = {name: value / divisor for name, value in fruit_scores.items()}
        anti_scores = {name: value / divisor for name, value in anti_scores.items()}
        fruit_alignment = sum(fruit_scores.values()) / max(1, len(fruit_scores))
        anti_alignment = sum(anti_scores.values()) / max(1, len(anti_scores))
        net_alignment = fruit_alignment - anti_alignment
        return {
            "semantic_backend": self.backend,
            "semantic_ontology": "explicit coherence ontology",
            "semantic_fruit_alignment": round(fruit_alignment, 4),
            "semantic_anti_alignment": round(anti_alignment, 4),
            "semantic_net_alignment": round(net_alignment, 4),
            "semantic_dominant_anchor": max(fruit_scores.items(), key=lambda item: item[1])[0],
            "semantic_dominant_anti_anchor": max(anti_scores.items(), key=lambda item: item[1])[0],
            "semantic_fruit_anchor_scores": {name: round(value, 4) for name, value in fruit_scores.items()},
            "semantic_anti_anchor_scores": {name: round(value, 4) for name, value in anti_scores.items()},
        }


def flatten_scores(source: Path, scores: dict, meta: dict, excel_path: Path | None) -> dict:
    agg = scores.get("agg", {})
    christ = scores.get("christ_vector", {})
    semantic = scores.get("semantic_anchor", {})
    return {
        "source_file": str(source),
        "source_name": source.name,
        "tokens": scores.get("tokens", 0),
        "sentences": agg.get("sentences", 0),
        "truth": round(float(agg.get("truth", 0)), 4),
        "coherence": round(float(agg.get("coherence", 0)), 4),
        "fruit": round(float(agg.get("fruit", 0)), 4),
        "anti_fruit": round(float(agg.get("anti_fruit", 0)), 4),
        "grounding": round(float(agg.get("grounding", 0)), 4),
        "contradiction": round(float(agg.get("contradiction", 0)), 4),
        "propaganda": scores.get("propaganda", 0),
        "snr": scores.get("snr", 0),
        "assertion_ratio": scores.get("assertion_ratio", 0),
        "contradiction_pressure": scores.get("contradiction_pressure", 0),
        "coherence_density": scores.get("coh_density", 0),
        "cronkite_delta": meta.get("cronkite_delta", 0),
        "sigma_state": meta.get("sigma_state", 0),
        "archetype": meta.get("archetype", ""),
        "role": meta.get("role", ""),
        "kill_conditions": meta.get("kill_conditions", ""),
        "christ_vector_alignment": round(float(meta.get("cv_alignment", 0)), 4),
        "christ_vector_label": meta.get("cv_label", ""),
        "dominant_fruit": meta.get("dominant_fruit", christ.get("dominant_fruit", "")),
        "weakest_fruit": meta.get("weakest_fruit", christ.get("weakest_fruit", "")),
        "semantic_backend": semantic.get("semantic_backend", ""),
        "semantic_ontology": semantic.get("semantic_ontology", ""),
        "semantic_fruit_alignment": semantic.get("semantic_fruit_alignment", 0),
        "semantic_anti_alignment": semantic.get("semantic_anti_alignment", 0),
        "semantic_net_alignment": semantic.get("semantic_net_alignment", 0),
        "semantic_dominant_anchor": semantic.get("semantic_dominant_anchor", ""),
        "semantic_dominant_anti_anchor": semantic.get("semantic_dominant_anti_anchor", ""),
        "excel_path": str(excel_path) if excel_path else "",
    }


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Run Truth Engine Fruits of the Spirit scoring from Paper Grader.")
    parser.add_argument("--input", required=True, help="Single file or folder to score")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output folder")
    parser.add_argument("--pattern", default="*.md", help="Recursive glob when input is a folder")
    parser.add_argument("--no-recursive", action="store_true", help="Only scan the input folder itself")
    parser.add_argument("--limit", type=int, default=0, help="Optional max files")
    parser.add_argument("--no-excel", action="store_true", help="Skip per-paper Fruits Template Excel exports")
    args = parser.parse_args()

    score_mod, enrich_mod, export_mod = setup_truth_engine(export_excel=not args.no_excel)
    semantic_scorer = SemanticAnchorScorer()
    input_path = Path(args.input)
    files = find_inputs(input_path, args.pattern, recursive=not args.no_recursive)
    if args.limit > 0:
        files = files[: args.limit]
    if not files:
        print(f"No scoreable files found under {input_path}")
        return 2

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.output) / f"fruits_run_{stamp}"
    excel_dir = out_dir / "excel"
    rows: list[dict] = []
    errors: list[dict] = []

    for index, path in enumerate(files, 1):
        print(f"[{index}/{len(files)}] Fruits scoring: {path.name}", flush=True)
        try:
            text = read_text(path)
            scores = score_mod.compute_scores(text, fname=path.name)
            scores["semantic_anchor"] = semantic_scorer.score(text)
            meta = enrich_mod.enrich_metrics(scores)
            excel_path = None
            if not args.no_excel and export_mod is not None:
                excel_path = excel_dir / f"Fruits_{path.stem}_{stamp}.xlsx"
                export_mod.write_to_template(scores, meta, excel_path, source_name=path.name)
            rows.append(flatten_scores(path, scores, meta, excel_path))
        except Exception as exc:
            errors.append({"source_file": str(path), "error": repr(exc)})
            print(f"  ERROR: {exc}", flush=True)

    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "fruits_scores.json").write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")
    (out_dir / "fruits_errors.json").write_text(json.dumps(errors, indent=2, ensure_ascii=False), encoding="utf-8")
    write_csv(out_dir / "fruits_scores.csv", rows)
    summary = {
        "run_id": stamp,
        "input": str(input_path),
        "pattern": args.pattern,
        "files_seen": len(files),
        "scored": len(rows),
        "errors": len(errors),
        "semantic_backend": semantic_scorer.backend,
        "semantic_claim": "Measures alignment to an explicit coherence ontology, not proof of spiritual truth.",
        "lexicon": LEXICON_STATUS,
        "output_dir": str(out_dir),
    }
    (out_dir / "fruits_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2), flush=True)
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())



