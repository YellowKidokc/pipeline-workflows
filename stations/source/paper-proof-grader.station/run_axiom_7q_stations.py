from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path


HERE = Path(__file__).resolve().parent
OUTPUT = HERE / "EXPORTS" / "reports"
REFERENCE = HERE / "REFERENCE"
CANONICAL_CHAIN_SOURCE = REFERENCE / "canonical_chain_nodes.psv"
AXIOM_SEQUENCE_SOURCE = REFERENCE / "axiom_sequence_sources"
STATION_ROOT = OUTPUT / "station-runs" / datetime.now().strftime("axiom-7q-%Y%m%d_%H%M%S")


AXIOM_RULES = [
    ("truth_ground", "Truth / Logos ground", ["truth", "logos", "coherence", "person"]),
    ("information_substrate", "Information substrate", ["information", "algorithmic", "mutual information", "substrate"]),
    ("observer_actualization", "Observer / measurement", ["observer", "measurement", "observe", "qrng"]),
    ("grace_repair", "Grace / negentropy / repair", ["grace", "negentropy", "repair", "restoration"]),
    ("entropy_thermo", "Entropy / thermodynamic constraint", ["entropy", "thermodynamic", "shannon", "noise"]),
    ("falsifiability", "Falsification / kill condition", ["falsification", "falsifiable", "fails", "failure", "kill"]),
    ("master_equation", "Master Equation / chi field", ["master equation", "χ", "chi", "field"]),
    ("moral_conservation", "Moral conservation", ["moral", "justice", "fairness", "duty", "sacrifice"]),
    ("experiment_protocol", "Experimental protocol", ["experiment", "protocol", "randomization", "pre-commitment", "hash"]),
    ("model_coupling", "Model coupling / susceptibility", ["coupling", "susceptibility", "model", "slope"]),
]


AXIOM_SOURCE_FILES = [
    ("axioms-layer-0-core.html", "Layer 0 + Tier 1"),
    ("axioms-layer-2-derived.html", "Layer 2"),
    ("axioms-layer-3-extended.html", "Layer 3"),
    ("axioms-closure.html", "Closure"),
    ("fp-005.html", "FP-005"),
    ("fp-005-enhanced.html", "FP-005 Enhanced"),
]


FALLBACK_CHAIN_HINTS = {
    "truth_ground": ["FINAL-1", "A1.1"],
    "information_substrate": ["A2.1", "BC7", "A1.3"],
    "observer_actualization": ["A5.1", "BC1"],
    "grace_repair": ["BC2", "D9.1"],
    "entropy_thermo": ["T3.1"],
    "falsifiability": ["META-1", "META-3"],
    "master_equation": ["E19.1"],
    "moral_conservation": ["FINAL-3"],
    "experiment_protocol": [],
    "model_coupling": ["BC8"],
}


AXIOM_TERM_STOPWORDS = {
    "about", "above", "after", "again", "against", "another", "because", "being", "between",
    "chain", "claim", "condition", "conditions", "could", "derive", "derived", "does",
    "everything", "from", "into", "itself", "meaning", "model", "must", "other", "position",
    "required", "requires", "same", "should", "something", "source", "system", "than",
    "their", "there", "therefore", "thing", "things", "through", "with", "within", "without",
}


OPENAI_SYSTEM_PROMPT = """You are the Theophysics Axiom + 7Q verifier.
Work inside the framework, but do not rubber-stamp claims. Test each claim with:
- 7Q Forward: identity, scope, mechanism, evidence, falsifiability, boundary, listener risk.
- 7Q Reverse: what would have to fail for this claim to collapse.
- Axiom fit: which candidate axiom/concept node actually bears weight.

For axiom_ids, use only ids supplied in deterministic_axiom_hits. If the
current station does not supply the axiom you want, put that idea in
suggested_registry_terms instead of inventing a new id.

Return strict JSON only. Keep the answer compact and operational."""


OPENAI_JSON_INSTRUCTIONS = {
    "candidate": "accept | repair | reject | unsure",
    "confidence": "number from 0 to 1",
    "axiom_ids": ["candidate node ids from deterministic_axiom_hits only"],
    "suggested_registry_terms": ["missing canonical registry terms or node ideas"],
    "required_evidence": ["specific missing evidence or checks"],
    "failure_conditions": ["specific kill conditions or breakpoints"],
    "rationale": "short reason, no padding",
}


def strip_html(value: str) -> str:
    value = re.sub(r"<script\b.*?</script>", " ", value, flags=re.I | re.S)
    value = re.sub(r"<style\b.*?</style>", " ", value, flags=re.I | re.S)
    value = re.sub(r"<[^>]+>", " ", value)
    return html.unescape(" ".join(value.split()))


def chain_position(label: str) -> int | None:
    match = re.search(r"Chain Position\s+(\d+)/188", label, flags=re.I)
    return int(match.group(1)) if match else None


def sequence_sort_key(item: dict) -> float:
    position = item.get("chain_position")
    if position is not None:
        return float(position)
    if item.get("display_id", "").startswith("P0.1"):
        return -2.0
    if item.get("display_id", "").startswith("P0.2"):
        return -1.0
    return float(item.get("source_order", 9999)) + (item.get("card_order", 0) / 1000)


def registry_node_id(display_id: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", display_id.lower()).strip("_")


def registry_terms(display_id: str, name: str, statement: str) -> list[str]:
    text = f"{display_id} {name} {statement}".lower()
    words = [
        w.strip("'")
        for w in re.findall(r"[a-z][a-z0-9_'-]{5,}", text)
        if w.strip("'") not in AXIOM_TERM_STOPWORDS
    ]
    name_phrase = name.lower()
    statement_phrase = statement.lower()
    phrases = [name_phrase]
    if statement_phrase and len(statement_phrase) <= 90:
        phrases.append(statement_phrase)
    return sorted(set([w.strip("'") for w in words] + [p for p in phrases if p]))


def load_canonical_chain_registry() -> list[dict]:
    if not CANONICAL_CHAIN_SOURCE.exists():
        return []
    registry = []
    with CANONICAL_CHAIN_SOURCE.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter="|")
        for row in reader:
            try:
                position = int(row.get("position") or 0)
            except ValueError:
                position = None
            display_id = row.get("display_id") or row.get("node_id") or ""
            label = row.get("name") or display_id
            statement = row.get("formal_statement") or ""
            family = row.get("family") or ""
            node_type = row.get("node_type") or ""
            level = row.get("level") or ""
            terms = registry_terms(display_id, label, " ".join([statement, family, node_type, level]))
            terms.extend([term.strip().lower() for term in family.split(",") if term.strip()])
            registry.append({
                "node_id": row.get("node_id") or registry_node_id(display_id),
                "display_id": display_id,
                "label": label,
                "statement": statement,
                "raw_id": display_id,
                "chain_position": position,
                "sequence": position,
                "source_layer": level,
                "node_type": node_type,
                "family": family,
                "depends_on": row.get("depends_on") or "",
                "kill_condition": row.get("kill_condition") or "",
                "source_file": str(CANONICAL_CHAIN_SOURCE),
                "source_anchor": row.get("node_id") or display_id,
                "source_order": -1,
                "card_order": position or 0,
                "terms": sorted(set(terms)),
                "source": "canonical",
            })
    return sorted(registry, key=lambda item: item.get("sequence") if item.get("sequence") is not None else 99999)


def load_axiom_registry() -> list[dict]:
    registry = load_canonical_chain_registry()
    seen = set()
    canonical_display_ids = set()
    for item in registry:
        seen.add((item.get("display_id"), item.get("label"), item.get("chain_position"), item.get("source_layer")))
        canonical_display_ids.add(item.get("display_id"))
    for source_order, (filename, layer) in enumerate(AXIOM_SOURCE_FILES):
        path = AXIOM_SEQUENCE_SOURCE / filename
        if not path.exists():
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        cards = re.findall(r'<div class="axiom-card[^"]*" id="([^"]+)">(.*?)(?=<div class="axiom-card|\n\s*<h2|\n\s*</main>|\n\s*</section>)', text, flags=re.S)
        for card_order, (anchor, block) in enumerate(cards):
            id_match = re.search(r'<span class="axiom-id">(.*?)</span>', block, flags=re.S)
            name_match = re.search(r'<(?:div|h3) class="axiom-name">(.*?)</(?:div|h3)>', block, flags=re.S)
            statement_match = re.search(r'<div class="axiom-statement">(.*?)</div>', block, flags=re.S)
            if not id_match or not name_match:
                continue
            raw_id = strip_html(id_match.group(1))
            display_id = raw_id.split(" — ", 1)[0].split(" - ", 1)[0].strip()
            name = strip_html(name_match.group(1))
            statement = strip_html(statement_match.group(1)) if statement_match else ""
            if display_id.upper() in {"THEOREM"}:
                display_id = anchor.upper().replace("-", "_")
            if display_id in canonical_display_ids:
                continue
            key = (display_id, name, chain_position(raw_id), layer)
            if key in seen:
                continue
            seen.add(key)
            registry.append({
                "node_id": registry_node_id(display_id),
                "display_id": display_id,
                "label": name,
                "statement": statement,
                "raw_id": raw_id,
                "chain_position": chain_position(raw_id),
                "sequence": None,
                "source_layer": layer,
                "source_file": filename,
                "source_anchor": anchor,
                "source_order": source_order,
                "card_order": card_order,
                "terms": registry_terms(display_id, name, statement),
                "source": "registry",
            })
    registry.sort(key=lambda item: (
        0 if item.get("source") == "canonical" else 1,
        item.get("sequence") if item.get("source") == "canonical" else sequence_sort_key(item),
    ))
    for idx, item in enumerate(registry, 1):
        if item.get("source") != "canonical":
            item["sequence"] = 1000 + idx
    return registry


AXIOM_REGISTRY = load_axiom_registry()
AXIOM_BY_DISPLAY_ID = {item["display_id"]: item for item in AXIOM_REGISTRY}


def read_claims(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def axiom_hits(text: str) -> list[dict]:
    lower = text.lower()
    hits = []
    seen = set()
    for item in AXIOM_REGISTRY:
        matched = [t for t in item["terms"] if t and t in lower]
        if matched:
            hit = dict(item)
            hit["matched_terms"] = matched[:8]
            hits.append(hit)
            seen.add(item["display_id"])
    for node_id, label, terms in AXIOM_RULES:
        matched = [t for t in terms if t.lower() in lower]
        if matched:
            for display_id in FALLBACK_CHAIN_HINTS.get(node_id, []):
                item = AXIOM_BY_DISPLAY_ID.get(display_id)
                if item and item["display_id"] not in seen:
                    hit = dict(item)
                    hit["matched_terms"] = matched
                    hit["source"] = "registry_hint"
                    hits.append(hit)
                    seen.add(item["display_id"])
            hits.append({
                "node_id": node_id,
                "display_id": node_id,
                "label": label,
                "statement": "",
                "raw_id": node_id,
                "chain_position": None,
                "sequence": 10000 + len(hits),
                "source_layer": "Fallback",
                "source_file": "",
                "source_anchor": "",
                "terms": terms,
                "source": "fallback",
                "matched_terms": matched,
            })
    return sorted(hits, key=lambda h: (h.get("sequence") or 99999, h.get("label", "")))


def forward_score(row: dict) -> dict:
    q_cols = ["Q1_identity", "Q2_scope", "Q3_mechanism", "Q4_evidence", "Q5_falsifiability", "Q6_boundary", "Q7_listener_risk"]
    raw = {q: row.get(q, "") for q in q_cols}
    score = 0
    score += 1 if raw["Q1_identity"] == "clear" else 0
    score += 1 if raw["Q2_scope"] == "bounded" else 0
    score += 1 if raw["Q3_mechanism"] == "present" else 0
    score += 1 if raw["Q4_evidence"] == "present" else 0
    score += 1 if raw["Q5_falsifiability"] == "present" else 0
    score += 1 if raw["Q6_boundary"] == "present" else 0
    score += 1 if raw["Q7_listener_risk"] == "normal" else 0
    return {"score": score, "max_score": 7, "raw": raw}


def reverse_verdict(row: dict) -> dict:
    weaknesses = []
    if row.get("Q4_evidence") != "present":
        weaknesses.append("missing_evidence")
    if row.get("Q5_falsifiability") != "present":
        weaknesses.append("missing_kill_condition")
    if row.get("Q3_mechanism") != "present":
        weaknesses.append("missing_mechanism")
    if row.get("Q2_scope") != "bounded":
        weaknesses.append("overbroad_scope")
    if row.get("Q6_boundary") != "present":
        weaknesses.append("missing_boundary")
    if row.get("Q7_listener_risk") == "high":
        weaknesses.append("high_listener_risk")

    if "missing_evidence" in weaknesses and "missing_kill_condition" in weaknesses:
        status = "FAIL_REVIEW"
    elif len(weaknesses) >= 3:
        status = "WEAKENED"
    elif weaknesses:
        status = "SURVIVES_WITH_REPAIRS"
    else:
        status = "SURVIVES"
    return {"status": status, "weaknesses": weaknesses}


def openai_verify_claim(row: dict, hits: list[dict], forward: dict, reverse: dict, model: str) -> dict:
    packet = {
        "claim": row.get("one_sentence_claim", ""),
        "section": row.get("section", ""),
        "nearby_equation": row.get("nearby_equation", ""),
        "maturity": row.get("claim_maturity_label", ""),
        "deterministic_axiom_hits": hits,
        "deterministic_7q_forward": forward,
        "deterministic_7q_reverse": reverse,
        "output_schema": OPENAI_JSON_INSTRUCTIONS,
    }
    try:
        from openai import OpenAI

        client = OpenAI()
        request = {
            "model": model,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": OPENAI_SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": (
                        "Verify this claim for the Axiom + 7Q station. "
                        "Use the requested output schema exactly.\n\n"
                        + json.dumps(packet, ensure_ascii=False, indent=2)
                    ),
                },
            ],
        }
        if not model.lower().startswith("o"):
            request["temperature"] = 0
        response = client.chat.completions.create(
            **request
        )
        content = response.choices[0].message.content or "{}"
        parsed = json.loads(content)
        parsed["enabled"] = True
        parsed["model"] = model
        return parsed
    except Exception as exc:
        return {
            "enabled": True,
            "model": model,
            "error": f"{type(exc).__name__}: {exc}",
        }


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def _verifier_md(verifier: dict | None) -> list[str]:
    if not verifier:
        return ["- OpenAI verifier: not run"]
    if verifier.get("error"):
        return [f"- OpenAI verifier: ERROR ({verifier.get('error')})"]
    return [
        f"- OpenAI verifier: {verifier.get('candidate', 'unsure')} "
        f"(confidence {verifier.get('confidence', '')}, model {verifier.get('model', '')})",
        f"- OpenAI axiom ids: {', '.join(verifier.get('axiom_ids') or []) or 'none'}",
        f"- OpenAI suggested registry terms: {', '.join(verifier.get('suggested_registry_terms') or []) or 'none'}",
        f"- OpenAI required evidence: {', '.join(verifier.get('required_evidence') or []) or 'none listed'}",
        f"- OpenAI failure conditions: {', '.join(verifier.get('failure_conditions') or []) or 'none listed'}",
        f"- OpenAI rationale: {verifier.get('rationale', '')}",
    ]


def write_md(path: Path, payload: dict) -> None:
    lines = [
        f"# {payload['paper_id']} - Axiom + 7Q Station Pass",
        "",
        f"Generated: {payload['generated_at']}",
        "",
        "## Summary",
        "",
        f"- Claims: {payload['claim_count']}",
        f"- Average 7Q forward score: {payload['forward_summary']['average_score']}/7",
        f"- Reverse verdicts: {payload['reverse_summary']['status_counts']}",
        f"- Chain-node / concept hits: {payload['axiom_summary']['hit_counts']}",
        "",
        "## Top Repairs",
        "",
    ]
    for item, count in payload["reverse_summary"]["weakness_counts"].most_common(10):
        lines.append(f"- {item}: {count}")
    lines += ["", "## Claim Rows", ""]
    for row in payload["claims"]:
        lines += [
            f"### Claim {row['index']}: {row['section']}",
            "",
            f"- Forward: {row['forward']['score']}/7",
            f"- Reverse: {row['reverse']['status']} ({', '.join(row['reverse']['weaknesses']) or 'no major weakness'})",
            f"- Chain-node hits: {', '.join(h['label'] for h in row['axiom_hits']) or 'none'}",
            *_verifier_md(row.get("openai_verifier")),
            f"- Claim: {row['claim'][:500]}",
            "",
        ]
    path.write_text("\n".join(lines), encoding="utf-8")


def _join_or_empty(items: list | None) -> str:
    return ", ".join(str(item) for item in (items or []))


def _status_class(status: str) -> str:
    return {
        "SURVIVES": "ok",
        "SURVIVES_WITH_REPAIRS": "repair",
        "WEAKENED": "weak",
        "FAIL_REVIEW": "fail",
    }.get(status, "neutral")


def _candidate_class(candidate: str) -> str:
    return {
        "accept": "ok",
        "repair": "repair",
        "reject": "fail",
        "unsure": "weak",
    }.get((candidate or "").lower(), "neutral")


def _html_page(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      --bg: #f6f7f3;
      --panel: #ffffff;
      --ink: #1d2524;
      --muted: #586260;
      --line: #d8ded8;
      --navy: #1f4e78;
      --ok: #dcebd8;
      --ok-ink: #24522b;
      --repair: #fff0bf;
      --repair-ink: #6a4a00;
      --weak: #f7d6c7;
      --weak-ink: #743716;
      --fail: #f3c4c4;
      --fail-ink: #7d2020;
      --neutral: #e8eceb;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font: 15px/1.55 "Segoe UI", Arial, sans-serif;
    }}
    header {{
      background: #172b33;
      color: #fff;
      padding: 24px clamp(18px, 4vw, 48px);
      border-bottom: 5px solid #b78f3b;
    }}
    header h1 {{ margin: 0 0 6px; font-size: clamp(24px, 3vw, 38px); letter-spacing: 0; }}
    header p {{ margin: 0; color: #d7e2e0; max-width: 1100px; }}
    main {{ width: min(1360px, calc(100vw - 32px)); margin: 22px auto 44px; }}
    h2 {{ margin: 26px 0 10px; font-size: 22px; }}
    h3 {{ margin: 0 0 8px; font-size: 18px; }}
    a {{ color: var(--navy); }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(210px, 1fr)); gap: 12px; }}
    .metric, .claim, .table-wrap {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 1px 2px rgba(0,0,0,.05);
    }}
    .metric {{ padding: 14px; }}
    .metric .label {{ color: var(--muted); font-size: 12px; text-transform: uppercase; letter-spacing: .06em; }}
    .metric .value {{ font-size: 26px; font-weight: 700; margin-top: 3px; }}
    .claim {{ padding: 16px; margin: 14px 0; }}
    .claim-text {{ white-space: pre-wrap; }}
    .meta {{ color: var(--muted); font-size: 13px; }}
    .badges {{ display: flex; flex-wrap: wrap; gap: 6px; margin: 8px 0 10px; }}
    .badge {{
      display: inline-flex;
      align-items: center;
      min-height: 24px;
      padding: 2px 8px;
      border-radius: 999px;
      border: 1px solid rgba(0,0,0,.12);
      font-size: 12px;
      font-weight: 650;
    }}
    .ok {{ background: var(--ok); color: var(--ok-ink); }}
    .repair {{ background: var(--repair); color: var(--repair-ink); }}
    .weak {{ background: var(--weak); color: var(--weak-ink); }}
    .fail {{ background: var(--fail); color: var(--fail-ink); }}
    .neutral {{ background: var(--neutral); color: var(--ink); }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ padding: 9px 10px; border-bottom: 1px solid var(--line); vertical-align: top; text-align: left; }}
    th {{ background: #edf2f0; font-size: 12px; text-transform: uppercase; letter-spacing: .04em; }}
    .table-wrap {{ overflow-x: auto; }}
    .two-col {{ display: grid; grid-template-columns: minmax(0, 1fr) minmax(0, 1fr); gap: 12px; }}
    .note {{ background: #eef3f6; border-left: 4px solid var(--navy); padding: 10px 12px; margin: 10px 0; }}
    code {{ background: #eef0ed; padding: 1px 4px; border-radius: 4px; }}
    @media (max-width: 820px) {{
      main {{ width: min(100% - 18px, 1360px); }}
      .two-col {{ grid-template-columns: 1fr; }}
      th, td {{ padding: 8px; }}
    }}
  </style>
</head>
<body>
{body}
</body>
</html>
"""


def write_paper_html(path: Path, payload: dict) -> None:
    rows = []
    for row in payload["claims"]:
        reverse = row["reverse"]
        verifier = row.get("openai_verifier") or {}
        axiom_rows = []
        for hit in row.get("axiom_hits", []):
            axiom_rows.append(f"""
              <tr>
                <td>{html.escape(str(hit.get('sequence', '')))}</td>
                <td>{html.escape(str(hit.get('chain_position') or ''))}</td>
                <td>{html.escape(hit.get('display_id', hit.get('node_id', '')))}</td>
                <td>{html.escape(hit.get('label', ''))}</td>
                <td>{html.escape(hit.get('source_layer', ''))}</td>
                <td>{html.escape(hit.get('node_type', ''))}</td>
                <td>{html.escape(hit.get('source', ''))}</td>
                <td>{html.escape(_join_or_empty(hit.get('matched_terms')))}</td>
              </tr>
""")
        axiom_table = f"""
        <div class="table-wrap">
          <table>
            <thead>
              <tr><th>Seq</th><th>Chain</th><th>ID</th><th>Node</th><th>Layer</th><th>Type</th><th>Source</th><th>Matched</th></tr>
            </thead>
            <tbody>{''.join(axiom_rows)}</tbody>
          </table>
        </div>
""" if axiom_rows else '<p><strong>Axiom hits:</strong> none</p>'
        verifier_block = ""
        if verifier.get("error"):
            verifier_block = f"<p class=\"note\">OpenAI verifier error: {html.escape(verifier['error'])}</p>"
        elif verifier:
            verifier_block = f"""
          <div class="two-col">
            <div>
              <h3>O3 Verifier</h3>
              <div class="badges">
                <span class="badge {_candidate_class(verifier.get('candidate', ''))}">{html.escape(verifier.get('candidate', 'unsure'))}</span>
                <span class="badge neutral">confidence {html.escape(str(verifier.get('confidence', '')))}</span>
                <span class="badge neutral">{html.escape(verifier.get('model', ''))}</span>
              </div>
              <p><strong>Axiom ids:</strong> {html.escape(_join_or_empty(verifier.get('axiom_ids')) or 'none')}</p>
              <p><strong>Suggested registry terms:</strong> {html.escape(_join_or_empty(verifier.get('suggested_registry_terms')) or 'none')}</p>
              <p><strong>Rationale:</strong> {html.escape(verifier.get('rationale', ''))}</p>
            </div>
            <div>
              <h3>Verifier Pressure</h3>
              <p><strong>Required evidence:</strong> {html.escape(_join_or_empty(verifier.get('required_evidence')) or 'none listed')}</p>
              <p><strong>Failure conditions:</strong> {html.escape(_join_or_empty(verifier.get('failure_conditions')) or 'none listed')}</p>
            </div>
          </div>
"""
        else:
            verifier_block = '<p class="note">OpenAI verifier not run for this claim.</p>'

        rows.append(f"""
      <article class="claim">
        <div class="meta">Claim {row['index']} / {html.escape(row.get('section', ''))}</div>
        <h3>{html.escape(row.get('section', '') or 'Untitled section')}</h3>
        <div class="badges">
          <span class="badge neutral">7Q {row['forward']['score']}/7</span>
          <span class="badge {_status_class(reverse['status'])}">{html.escape(reverse['status'])}</span>
          <span class="badge neutral">{html.escape(row.get('maturity', '') or 'unlabeled')}</span>
        </div>
        <p><strong>Weaknesses:</strong> {html.escape(_join_or_empty(reverse.get('weaknesses')) or 'none')}</p>
        <p class="claim-text">{html.escape(row.get('claim', ''))}</p>
        <h3>Matched Chain Sequence</h3>
        {axiom_table}
        {verifier_block}
      </article>
""")

    body = f"""
<header>
  <h1>{html.escape(payload['paper_id'])}</h1>
  <p>Axiom + 7Q station report. Generated {html.escape(payload['generated_at'])}.</p>
</header>
<main>
  <section class="grid">
    <div class="metric"><div class="label">Claims</div><div class="value">{payload['claim_count']}</div></div>
    <div class="metric"><div class="label">Average 7Q Forward</div><div class="value">{payload['forward_summary']['average_score']}/7</div></div>
    <div class="metric"><div class="label">Reverse Verdicts</div><div class="value">{html.escape(str(payload['reverse_summary']['status_counts']))}</div></div>
    <div class="metric"><div class="label">Chain-Node Hits</div><div class="value">{len(payload['axiom_summary']['hit_counts'])}</div></div>
    <div class="metric"><div class="label">Registry Nodes Loaded</div><div class="value">{len(AXIOM_REGISTRY)}</div></div>
  </section>
  <section>
    <h2>Claim Review</h2>
    {''.join(rows)}
  </section>
</main>
"""
    path.write_text(_html_page(f"{payload['paper_id']} - Axiom + 7Q", body), encoding="utf-8")


def write_batch_html(path: Path, index: dict, manifests: list[dict]) -> None:
    paper_by_id = {m["paper_id"]: m for m in manifests}
    rows = []
    for paper in index["papers"]:
        manifest = paper_by_id.get(paper["paper_id"], {})
        verifier_counts = Counter()
        verifier_calls = 0
        for claim in manifest.get("claims", []):
            verifier = claim.get("openai_verifier") or {}
            if verifier:
                verifier_calls += 1
                verifier_counts[verifier.get("candidate", "error" if verifier.get("error") else "unknown")] += 1
        href = f"{paper['paper_id']}/axiom-7q-stations.html"
        rows.append(f"""
          <tr>
            <td><a href="{html.escape(href)}">{html.escape(paper['paper_id'])}</a></td>
            <td>{paper['claim_count']}</td>
            <td>{paper['average_forward_score']}/7</td>
            <td>{html.escape(str(paper['reverse_status_counts']))}</td>
            <td>{html.escape(str(paper['axiom_hit_counts']))}</td>
            <td>{verifier_calls}</td>
            <td>{html.escape(str(dict(verifier_counts)))}</td>
          </tr>
""")

    body = f"""
<header>
  <h1>Axiom + 7Q Station Batch</h1>
  <p>Generated {html.escape(index['generated_at'])}. Deterministic station outputs plus optional O3 verifier lane.</p>
</header>
<main>
  <section class="grid">
    <div class="metric"><div class="label">Papers</div><div class="value">{index['paper_count']}</div></div>
    <div class="metric"><div class="label">Station Root</div><div class="value" style="font-size:16px;overflow-wrap:anywhere">{html.escape(index['station_root'])}</div></div>
  </section>
  <section>
    <h2>Batch Review</h2>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>Paper</th>
            <th>Claims</th>
            <th>Avg 7Q</th>
            <th>Reverse</th>
            <th>Axiom Hits</th>
            <th>O3 Calls</th>
            <th>O3 Candidates</th>
          </tr>
        </thead>
        <tbody>
          {''.join(rows)}
        </tbody>
      </table>
    </div>
  </section>
  <section>
    <h2>Artifacts</h2>
    <p><a href="axiom-7q-review.xlsx">Excel review workbook</a> / <a href="batch-index.json">JSON index</a> / <a href="batch-index.md">Markdown index</a></p>
  </section>
</main>
"""
    path.write_text(_html_page("Axiom + 7Q Station Batch", body), encoding="utf-8")


def write_batch_workbook(path: Path, index: dict, manifests: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    wb = Workbook()
    summary = wb.active
    summary.title = "Batch Summary"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    weak_fill = PatternFill("solid", fgColor="F4CCCC")
    repair_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="D9EAD3")

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    summary.append([
        "paper_id",
        "claim_count",
        "average_forward_score",
        "reverse_status_counts",
        "axiom_hit_counts",
    ])
    for paper in index["papers"]:
        summary.append([
            paper["paper_id"],
            paper["claim_count"],
            paper["average_forward_score"],
            json.dumps(paper["reverse_status_counts"], ensure_ascii=False),
            json.dumps(paper["axiom_hit_counts"], ensure_ascii=False),
        ])
    style_header(summary)

    claims = wb.create_sheet("Claim Review")
    claims.append([
        "paper_id",
        "claim_index",
        "section",
        "forward_score",
        "reverse_status",
        "weaknesses",
        "axiom_sequences",
        "axiom_display_ids",
        "axiom_node_ids",
        "axiom_labels",
        "axiom_layers",
        "node_types",
        "node_families",
        "kill_conditions",
        "axiom_sources",
        "matched_terms",
        "openai_enabled",
        "openai_model",
        "openai_candidate",
        "openai_confidence",
        "openai_axiom_ids",
        "openai_suggested_registry_terms",
        "openai_required_evidence",
        "openai_failure_conditions",
        "openai_rationale",
        "openai_error",
        "claim",
        "review_note",
    ])
    for manifest in manifests:
        for row in manifest["claims"]:
            hits = row.get("axiom_hits", [])
            verifier = row.get("openai_verifier") or {}
            claims.append([
                manifest["paper_id"],
                row["index"],
                row["section"],
                row["forward"]["score"],
                row["reverse"]["status"],
                ", ".join(row["reverse"]["weaknesses"]),
                ", ".join(str(hit.get("sequence", "")) for hit in hits),
                ", ".join(hit.get("display_id", hit["node_id"]) for hit in hits),
                ", ".join(hit["node_id"] for hit in hits),
                ", ".join(hit["label"] for hit in hits),
                ", ".join(hit.get("source_layer", "") for hit in hits),
                ", ".join(hit.get("node_type", "") for hit in hits),
                ", ".join(hit.get("family", "") for hit in hits),
                " | ".join(hit.get("kill_condition", "") for hit in hits if hit.get("kill_condition")),
                ", ".join(hit.get("source", "") for hit in hits),
                ", ".join(
                    f"{hit.get('display_id', hit['node_id'])}:{'|'.join(hit['matched_terms'])}"
                    for hit in hits
                ),
                bool(verifier.get("enabled")),
                verifier.get("model", ""),
                verifier.get("candidate", ""),
                verifier.get("confidence", ""),
                ", ".join(verifier.get("axiom_ids") or []),
                ", ".join(verifier.get("suggested_registry_terms") or []),
                ", ".join(verifier.get("required_evidence") or []),
                ", ".join(verifier.get("failure_conditions") or []),
                verifier.get("rationale", ""),
                verifier.get("error", ""),
                row["claim"],
                "",
            ])
    style_header(claims)
    for row in claims.iter_rows(min_row=2):
        status = row[4].value
        fill = weak_fill if status == "FAIL_REVIEW" else repair_fill if status == "WEAKENED" else ok_fill if status == "SURVIVES" else None
        if fill:
            for cell in row:
                cell.fill = fill

    rules = wb.create_sheet("Chain Nodes")
    rules.append(["sequence", "chain_position", "display_id", "node_id", "label", "node_type", "level", "family", "depends_on", "kill_condition", "source", "source_file", "matched_terms"])
    if AXIOM_REGISTRY:
        for item in AXIOM_REGISTRY:
            rules.append([
                item.get("sequence"),
                item.get("chain_position"),
                item.get("display_id"),
                item.get("node_id"),
                item.get("label"),
                item.get("node_type", ""),
                item.get("source_layer"),
                item.get("family", ""),
                item.get("depends_on", ""),
                item.get("kill_condition", ""),
                item.get("source", ""),
                item.get("source_file"),
                ", ".join(item.get("terms", [])),
            ])
    else:
        for node_id, label, terms in AXIOM_RULES:
            rules.append(["", "", node_id, node_id, label, "", "Fallback", "", "", "", "fallback", "", ", ".join(terms)])
    style_header(rules)

    tags = wb.create_sheet("Tag Graph Review")
    tags.append([
        "paper_id",
        "claim_index",
        "claim_tag_id",
        "primary_family",
        "secondary_families",
        "node_display_id",
        "node_type",
        "node_level",
        "relationship_type",
        "confidence",
        "status",
        "tags",
        "graph_source",
        "evidence_excerpt",
        "kill_condition",
        "review_note",
    ])
    for manifest in manifests:
        for row in manifest["claims"]:
            hits = row.get("axiom_hits", [])
            if hits:
                for hit in hits:
                    tags.append([
                        manifest["paper_id"],
                        row["index"],
                        f"{manifest['paper_id']}::claim-{row['index']:03d}::{hit.get('display_id', hit.get('node_id', 'node'))}",
                        (hit.get("family", "").split(",")[0] if hit.get("family") else ""),
                        hit.get("family", ""),
                        hit.get("display_id", hit.get("node_id", "")),
                        hit.get("node_type", ""),
                        hit.get("source_layer", ""),
                        "supports_or_maps_to",
                        "",
                        "draft",
                        hit.get("family", ""),
                        hit.get("source", ""),
                        row.get("claim", "")[:500],
                        hit.get("kill_condition", ""),
                        "",
                    ])
            else:
                tags.append([
                    manifest["paper_id"],
                    row["index"],
                    f"{manifest['paper_id']}::claim-{row['index']:03d}::unmapped",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "unmapped",
                    "",
                    "needs_review",
                    "",
                    "",
                    row.get("claim", "")[:500],
                    "",
                    "",
                ])
    style_header(tags)

    graph = wb.create_sheet("Graph Edges")
    graph.append([
        "edge_id",
        "source_id",
        "source_type",
        "relationship_type",
        "target_id",
        "target_type",
        "weight",
        "status",
        "evidence_excerpt",
        "notes",
    ])
    for manifest in manifests:
        for row in manifest["claims"]:
            claim_id = f"{manifest['paper_id']}::claim-{row['index']:03d}"
            for hit in row.get("axiom_hits", []):
                target_id = hit.get("display_id", hit.get("node_id", ""))
                graph.append([
                    f"{claim_id}--maps_to--{target_id}",
                    claim_id,
                    "claim",
                    "maps_to",
                    target_id,
                    hit.get("node_type", "chain_node"),
                    "",
                    "draft",
                    row.get("claim", "")[:500],
                    "",
                ])
    style_header(graph)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)

    wb.save(path)


def process_file(path: Path, use_openai: bool = False, openai_model: str = "o3", openai_limit: int | None = None) -> dict:
    rows = read_claims(path)
    paper_id = rows[0]["paper_id"] if rows else path.stem.replace(".claim-audit", "")
    paper_dir = STATION_ROOT / paper_id
    paper_dir.mkdir(parents=True, exist_ok=True)

    claims = []
    forward_scores = []
    reverse_statuses = Counter()
    weakness_counts = Counter()
    hit_counts = Counter()
    openai_calls = 0

    for i, row in enumerate(rows, 1):
        text = f"{row.get('section','')} {row.get('one_sentence_claim','')} {row.get('nearby_equation','')}"
        hits = axiom_hits(text)
        for h in hits:
            hit_counts[h["node_id"]] += 1
        forward = forward_score(row)
        reverse = reverse_verdict(row)
        forward_scores.append(forward["score"])
        reverse_statuses[reverse["status"]] += 1
        weakness_counts.update(reverse["weaknesses"])
        verifier = None
        if use_openai and (openai_limit is None or openai_calls < openai_limit):
            verifier = openai_verify_claim(row, hits, forward, reverse, openai_model)
            openai_calls += 1
        claims.append({
            "index": i,
            "section": row.get("section", ""),
            "claim": row.get("one_sentence_claim", ""),
            "maturity": row.get("claim_maturity_label", ""),
            "axiom_hits": hits,
            "forward": forward,
            "reverse": reverse,
            "openai_verifier": verifier,
        })

    payload = {
        "schema_version": "paper-proof-grader.axiom_7q_station.v1",
        "paper_id": paper_id,
        "source_claim_audit": str(path),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "claim_count": len(claims),
        "axiom_summary": {"hit_counts": dict(hit_counts)},
        "forward_summary": {
            "average_score": round(sum(forward_scores) / len(forward_scores), 2) if forward_scores else 0,
            "score_counts": dict(Counter(forward_scores)),
        },
        "reverse_summary": {
            "status_counts": dict(reverse_statuses),
            "weakness_counts": weakness_counts,
        },
        "claims": claims,
    }
    # JSON cannot serialize Counter in nested payload directly.
    payload["reverse_summary"]["weakness_counts"] = dict(weakness_counts)

    write_json(paper_dir / "axiom-7q-stations.json", payload)
    # For markdown, restore a Counter-like object for most_common convenience.
    payload_for_md = dict(payload)
    payload_for_md["reverse_summary"] = dict(payload["reverse_summary"])
    payload_for_md["reverse_summary"]["weakness_counts"] = Counter(payload["reverse_summary"]["weakness_counts"])
    write_md(paper_dir / "axiom-7q-stations.md", payload_for_md)
    write_paper_html(paper_dir / "axiom-7q-stations.html", payload)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Run Axiom + 7Q station pass over paper grader claim audits.")
    parser.add_argument("--openai", action="store_true", help="Run optional OpenAI verifier beside deterministic station scoring.")
    parser.add_argument("--openai-model", default=os.environ.get("AXIOM_7Q_OPENAI_MODEL", "o3"))
    parser.add_argument("--openai-limit", type=int, default=None, help="Maximum OpenAI verifier calls per paper.")
    parser.add_argument("--file-limit", type=int, default=None, help="Maximum claim-audit CSV files to process.")
    args = parser.parse_args()

    STATION_ROOT.mkdir(parents=True, exist_ok=True)
    files = sorted(OUTPUT.glob("0*.claim-audit.csv"))
    if args.file_limit is not None:
        files = files[:args.file_limit]
    manifests = [
        process_file(
            p,
            use_openai=args.openai,
            openai_model=args.openai_model,
            openai_limit=args.openai_limit,
        )
        for p in files
    ]
    index = {
        "schema_version": "paper-proof-grader.axiom_7q_batch.v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "station_root": str(STATION_ROOT),
        "paper_count": len(manifests),
        "papers": [
            {
                "paper_id": m["paper_id"],
                "claim_count": m["claim_count"],
                "average_forward_score": m["forward_summary"]["average_score"],
                "reverse_status_counts": m["reverse_summary"]["status_counts"],
                "axiom_hit_counts": m["axiom_summary"]["hit_counts"],
            }
            for m in manifests
        ],
    }
    write_json(STATION_ROOT / "batch-index.json", index)
    lines = ["# Axiom + 7Q Batch Index", "", f"Generated: {index['generated_at']}", ""]
    for p in index["papers"]:
        lines += [
            f"## {p['paper_id']}",
            f"- Claims: {p['claim_count']}",
            f"- Average 7Q forward: {p['average_forward_score']}/7",
            f"- Reverse: {p['reverse_status_counts']}",
            f"- Chain-node hits: {p['axiom_hit_counts']}",
            "",
        ]
    (STATION_ROOT / "batch-index.md").write_text("\n".join(lines), encoding="utf-8")
    write_batch_workbook(STATION_ROOT / "axiom-7q-review.xlsx", index, manifests)
    write_batch_html(STATION_ROOT / "index.html", index, manifests)
    print(STATION_ROOT)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
