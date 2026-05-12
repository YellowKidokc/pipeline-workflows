"""Rubric exporter — assembles per-paper scoring sheets from station sidecars.

Reads every JSON sidecar that the pipeline drops next to a paper, builds a real
Excel workbook (openpyxl) with one sheet per rubric category plus an OVERVIEW
sheet, and emits a companion HTML report. Falls back to a JSON-shaped .xlsx
only if openpyxl is unavailable — but openpyxl is in requirements.txt now.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


SHEET_ORDER = [
    "OVERVIEW",
    "CLASSIFICATION",
    "FRAMEWORK",
    "FACT CHECK",
    "MATH CHECK",
    "CONTRADICTION CHECK",
    "TIMELINE CHECK",
    "AXIOM MAPPING",
    "QUALITY GRADE",
    "MEDIA ROUTING",
    "VECTOR INDEX",
    "COMPOSITE SCORES",
]

CATEGORY_WEIGHTS = {
    "classification": 0.10,
    "framework": 0.15,
    "fact_check": 0.15,
    "math": 0.10,
    "contradiction": 0.10,
    "timeline": 0.05,
    "axiom": 0.10,
    "quality": 0.15,
    "voice": 0.05,
    "publish_readiness": 0.05,
}


class RubricExporter:
    """Collects sidecars and writes <paper>_RUBRIC.xlsx + <paper>_REPORT.html."""

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def collect_sidecars(self, paper_dir: Path, paper_id: str) -> dict[str, Any]:
        """Read every JSON sidecar matching <paper_id>* into a name→payload dict."""
        data: dict[str, Any] = {}
        for f in paper_dir.glob(f"{paper_id}*"):
            if f.suffix == ".json":
                try:
                    data[f.name] = json.loads(f.read_text(encoding="utf-8"))
                except json.JSONDecodeError:
                    data[f.name] = {"_error": "invalid json", "raw": f.read_text(encoding="utf-8")[:500]}
        return data

    def export(self, paper_dir: Path, paper_id: str) -> tuple[Path, Path]:
        data = self.collect_sidecars(paper_dir, paper_id)
        excel_path = self.build_excel(paper_id, data)
        html_path = self.build_html(paper_id, data)
        return excel_path, html_path

    def build_excel(self, paper_id: str, data: dict[str, Any]) -> Path:
        out = self.output_dir / f"{paper_id}_RUBRIC.xlsx"
        if not HAS_OPENPYXL:
            payload = self._sheet_payload(paper_id, data)
            out.write_text(json.dumps(payload, indent=2), encoding="utf-8")
            return out

        wb = Workbook()
        wb.remove(wb.active)
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="FF1F3A5F")
        wrap = Alignment(wrap_text=True, vertical="top")

        composite, breakdown = self._composite(data)
        verdict = self._verdict(composite)

        for sheet_name in SHEET_ORDER:
            ws = wb.create_sheet(sheet_name[:31])
            rows = self._rows_for_sheet(sheet_name, paper_id, data, composite, verdict, breakdown)
            for row in rows:
                ws.append(row)
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = wrap
            for col in ws.columns:
                width = 12
                for cell in col:
                    value = "" if cell.value is None else str(cell.value)
                    width = min(60, max(width, len(value) + 2))
                ws.column_dimensions[col[0].column_letter].width = width

        wb.save(out)
        return out

    def build_html(self, paper_id: str, data: dict[str, Any]) -> Path:
        composite, breakdown = self._composite(data)
        verdict = self._verdict(composite)
        rows = "".join(
            f"<tr><td>{k}</td><td>{v:.3f}</td></tr>"
            for k, v in breakdown.items()
        )
        sidecar_links = "".join(
            f"<li>{name}</li>" for name in sorted(data)
        ) or "<li><em>no sidecars yet</em></li>"
        arc = self._arc_length(composite)
        html = f"""<!doctype html>
<html><head><meta charset='utf-8'>
<title>{paper_id} — Rubric Report</title>
<style>
body{{background:#0d1117;color:#e6d5a8;font-family:Georgia,serif;padding:24px}}
table{{border-collapse:collapse;margin:12px 0}}
td,th{{border:1px solid #c9a227;padding:6px 12px}}
th{{background:#1f3a5f;color:#fff}}
.verdict{{font-size:1.6em;color:#c9a227}}
</style></head><body>
<h1>{paper_id}</h1>
<div class='verdict'>{verdict} — composite {composite:.3f}</div>
<svg width='220' height='220' viewBox='0 0 220 220'>
<circle cx='110' cy='110' r='90' stroke='#33333a' stroke-width='12' fill='none'/>
<circle cx='110' cy='110' r='90' stroke='#c9a227' stroke-width='12' fill='none'
        stroke-dasharray='{arc} 999' transform='rotate(-90 110 110)'/>
<text x='110' y='118' text-anchor='middle' fill='#e6d5a8' font-size='28'>{composite:.2f}</text>
</svg>
<h2>Category scores</h2>
<table><tr><th>Category</th><th>Score</th></tr>{rows}</table>
<h2>Sidecars collected</h2>
<ul>{sidecar_links}</ul>
<p><a style='color:#c9a227' href='{paper_id}_RUBRIC.xlsx'>Open Excel rubric</a></p>
</body></html>"""
        path = self.output_dir / f"{paper_id}_REPORT.html"
        path.write_text(html, encoding="utf-8")
        return path

    def _arc_length(self, composite: float) -> float:
        circumference = 2 * 3.14159 * 90
        return max(0.0, min(circumference, circumference * composite))

    def _rows_for_sheet(
        self,
        sheet_name: str,
        paper_id: str,
        data: dict[str, Any],
        composite: float,
        verdict: str,
        breakdown: dict[str, float],
    ) -> list[list[Any]]:
        if sheet_name == "OVERVIEW":
            return [
                ["paper_id", paper_id],
                ["composite_score", composite],
                ["verdict", verdict],
                ["sidecars_found", len(data)],
                ["sidecar_files", ", ".join(sorted(data))],
            ]
        if sheet_name == "COMPOSITE SCORES":
            rows = [["category", "score", "weight"]]
            for cat, weight in CATEGORY_WEIGHTS.items():
                rows.append([cat, breakdown.get(cat, 0.0), weight])
            rows.append(["composite", composite, sum(CATEGORY_WEIGHTS.values())])
            rows.append(["verdict", verdict, ""])
            return rows
        if sheet_name == "CLASSIFICATION":
            return self._flatten_payload(self._find(data, ".fap.json"), ["doc_type", "laws", "confidence", "word_count"])
        if sheet_name == "FRAMEWORK":
            return self._flatten_payload(
                self._find(data, ".framework.json"),
                [
                    "framework_coverage_score",
                    "framework_depth",
                    "laws_referenced",
                    "axiom_schemata",
                    "equations_present",
                    "fruits_referenced",
                    "anti_fruits_referenced",
                    "trinity_aspects",
                    "moral_conservation",
                    "seven_q",
                    "confidence",
                ],
            )
        if sheet_name == "AXIOM MAPPING":
            return self._flatten_payload(self._find(data, ".axioms.json"), ["claims", "mappings", "gaps", "contradictions", "score"])
        if sheet_name == "QUALITY GRADE":
            grade = self._find(data, ".grade.json")
            payload = grade.get("payload", grade)
            return self._flatten_payload(payload, ["overall_score", "score", "voice", "rigor", "cross_domain", "notes"])
        if sheet_name == "MEDIA ROUTING":
            return self._flatten_payload(self._find(data, ".media.json"), ["lane", "confidence", "reasons", "recommended_next_station"])
        if sheet_name == "VECTOR INDEX":
            vectors = self._find(data, ".vectors.json")
            chunks = vectors.get("chunks", []) if isinstance(vectors, dict) else []
            rows = [["chunk_index", "tokens", "heading"]]
            metadata = vectors.get("metadata", []) if isinstance(vectors, dict) else []
            for i, chunk in enumerate(chunks):
                heading = metadata[i].get("heading_context", "") if i < len(metadata) and isinstance(metadata[i], dict) else ""
                rows.append([i, len(chunk.split()), heading])
            if len(rows) == 1:
                rows.append(["—", "—", "no vector sidecar yet"])
            return rows
        if sheet_name in {"FACT CHECK", "MATH CHECK", "CONTRADICTION CHECK", "TIMELINE CHECK"}:
            key = sheet_name.split()[0].lower()
            payload = self._find(data, f".{key}.json")
            if not payload:
                return [["status", "not yet processed"]]
            return self._flatten_payload(payload, list(payload.keys())[:20])
        return [["status", "no rows"]]

    def _find(self, data: dict[str, Any], suffix: str) -> dict[str, Any]:
        for name, payload in data.items():
            if name.endswith(suffix) and isinstance(payload, dict):
                return payload
        return {}

    def _flatten_payload(self, payload: dict[str, Any], keys: list[str]) -> list[list[Any]]:
        if not payload:
            return [["status", "not yet processed"]]
        rows: list[list[Any]] = [["field", "value"]]
        for key in keys:
            value = payload.get(key, "")
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)[:500]
            rows.append([key, value])
        return rows

    def _composite(self, data: dict[str, Any]) -> tuple[float, dict[str, float]]:
        """Derive per-category scores from sidecars and combine via CATEGORY_WEIGHTS."""
        breakdown: dict[str, float] = {k: 0.0 for k in CATEGORY_WEIGHTS}

        fap = self._find(data, ".fap.json")
        if fap:
            breakdown["classification"] = float(fap.get("confidence", 0.0))

        framework = self._find(data, ".framework.json")
        if framework:
            breakdown["framework"] = float(framework.get("confidence", framework.get("framework_coverage_score", 0.0)))

        axioms = self._find(data, ".axioms.json")
        if axioms:
            breakdown["axiom"] = float(axioms.get("score", 0.0))

        grade = self._find(data, ".grade.json")
        payload = grade.get("payload", grade) if grade else {}
        if payload:
            breakdown["quality"] = float(payload.get("overall_score", payload.get("score", 0.0)))
            breakdown["fact_check"] = float(payload.get("fact_check", payload.get("rigor", 0.0)))
            breakdown["math"] = float(payload.get("math_check", payload.get("math", 0.0)))
            breakdown["contradiction"] = float(payload.get("contradiction_check", 0.0))
            breakdown["timeline"] = float(payload.get("timeline_check", 0.0))
            breakdown["voice"] = float(payload.get("voice", 0.0))
            breakdown["publish_readiness"] = float(payload.get("publish_readiness", 0.0))

        composite = sum(breakdown[k] * w for k, w in CATEGORY_WEIGHTS.items())
        return round(composite, 3), breakdown

    def _verdict(self, score: float) -> str:
        if score >= 0.8:
            return "PUBLISH"
        if score >= 0.6:
            return "REVISE"
        if score >= 0.4:
            return "RESTRUCTURE"
        return "HOLD"

    def _sheet_payload(self, paper_id: str, data: dict[str, Any]) -> dict[str, list]:
        composite, breakdown = self._composite(data)
        verdict = self._verdict(composite)
        return {
            name: self._rows_for_sheet(name, paper_id, data, composite, verdict, breakdown)
            for name in SHEET_ORDER
        }
