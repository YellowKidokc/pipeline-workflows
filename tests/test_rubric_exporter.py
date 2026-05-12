import json

from openpyxl import load_workbook

from engines.pipeline.rubric_exporter import RubricExporter


def test_rubric_exporter_outputs(tmp_path):
    (tmp_path / "P1.fap.json").write_text(json.dumps({"doc_type": "paper", "confidence": 0.82, "laws": ["L1", "L7"]}))
    (tmp_path / "P1.framework.json").write_text(json.dumps({"framework_coverage_score": 0.6, "confidence": 0.75, "laws_referenced": ["L1"], "axiom_schemata": ["AS-001"]}))
    (tmp_path / "P1.grade.json").write_text(json.dumps({"payload": {"overall_score": 0.78, "voice": 0.7, "rigor": 0.8}}))

    ex = RubricExporter(tmp_path / "out")
    excel_path, html_path = ex.export(tmp_path, "P1")

    assert excel_path.exists() and html_path.exists()

    wb = load_workbook(excel_path)
    assert "OVERVIEW" in wb.sheetnames
    assert "FRAMEWORK" in wb.sheetnames
    assert "COMPOSITE SCORES" in wb.sheetnames
    overview = {row[0]: row[1] for row in wb["OVERVIEW"].iter_rows(values_only=True) if row[0]}
    assert overview["paper_id"] == "P1"
    assert overview["sidecars_found"] == 3
    assert isinstance(overview["composite_score"], (int, float))
    assert overview["composite_score"] > 0

    composite_rows = list(wb["COMPOSITE SCORES"].iter_rows(values_only=True))
    cat_to_score = {row[0]: row[1] for row in composite_rows[1:] if row[0] and row[0] not in {"composite", "verdict"}}
    assert "classification" in cat_to_score
    assert cat_to_score["classification"] == 0.82

    fact_rows = list(wb["FACT CHECK"].iter_rows(values_only=True))
    assert fact_rows[0] == ("status", "not yet processed")

    html_text = html_path.read_text(encoding="utf-8")
    assert "P1" in html_text and "composite" in html_text.lower()
